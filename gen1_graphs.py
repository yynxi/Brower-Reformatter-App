import json
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side  # Add to existing imports
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

class TimingSystemApp:
    def __init__(self, root):
        # Initialization and GUI setup
        self.root = root
        self.root.title("Timing System Data Formatter")

        # Data structures
        self.selected_file = None
        self.athletes = {"SQAH": [], "SQAF": [], "OTHER": []}  # Store athletes by team
        self.temp_guests = []  # Temporary guests
        self.current_team = "SQAH"
        self.team_var = tk.StringVar()
        self.event_var = tk.StringVar()
        self.hill_var = tk.StringVar()
        self.name_entry_width = 0  # Will store the width of the entry
        self.animation_height = 0  # Current animation height
        self.target_height = 0     # Target height for animation
        self.animation_speed = 2   # Pixels per frame
        self.is_animating = False
        self.line_height = 20  # Default line height - will be updated after widget creation
        self.snow_condition_var = tk.StringVar()
        self.sky_condition_var = tk.StringVar()
        self.precipitation_var = tk.StringVar()
        self.wind_condition_var = tk.StringVar()
        self.date_var = tk.StringVar()
        self.time_var = tk.StringVar()
        self.session_var = tk.StringVar()
        self.outlier_threshold = 2  # Number of standard deviations for outlier detection
        # Hill-related attributes
        self.recent_hills = []
        self.hill_animation_height = 0
        self.hill_target_height = 0
        self.hill_is_animating = False

        # Version and author
        self.version = "1.0"
        self.author = "Julian H. Brunet"
        
        # Recent names memory (max 2000 names)
        self.recent_names = []
        
        # Initialize data
        self.load_recent_names()
        self.load_recent_hills()
        self.load_athletes_from_json()

        # Build GUI
        self.build_gui()

        # Set the default team and bindings
        self.set_team("SQAH")
        self.athlete_name_entry.bind('<KeyRelease>', self.autocomplete_athlete_name)
        self.autocomplete_listbox.bind("<<ListboxSelect>>", self.on_suggestion_select)
        self.bind_deletion_keys()

        # Bind the resize event to update the suggestion box width
        self.athlete_name_entry.bind('<Configure>', self.update_suggestion_box_width)

    def build_gui(self):
        """Builds the GUI components."""
        # Title
        self.title_label = tk.Label(self.root, text="Brower Timing Reformatted", font=("Arial", 20, "bold"))
        self.title_label.grid(row=0, column=0, columnspan=3, pady=(10, 0))

        # File Selection Section (Select Browser CSV)
        self.file_frame = tk.Frame(self.root)
        self.file_frame.grid(row=1, column=0, pady=(5, 0), sticky="n")

        # File Selection Section with a thick border
        self.button_label_border_frame = tk.Frame(self.file_frame, bd=4, relief="solid")
        self.button_label_border_frame.pack(padx=5, pady=(5, 5))

        self.select_file_button = tk.Button(self.button_label_border_frame, text="Select Brower CSV", command=self.select_file)
        self.select_file_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.file_label = tk.Label(self.button_label_border_frame, text="No file selected", anchor="w", width=40)
        self.file_label.pack(side=tk.LEFT, padx=5, pady=5)

        # Training Parameters Section
        self.training_parameters_frame = tk.Frame(self.root, bd=4, relief="solid", padx=50, pady=10)
        self.training_parameters_frame.grid(row=1, column=0, pady=(72, 10), sticky="n")

        self.training_parameters_label = tk.Label(self.training_parameters_frame, text="Training Parameters", font=("Arial", 12, "bold"))
        self.training_parameters_label.grid(row=0, column=0, pady=(0, 10))

        # Add Reformat Button Section below Training Parameters
        self.reformat_frame = tk.Frame(self.root, bd=4, relief="solid", padx=10, pady=10)
        self.reformat_frame.grid(row=2, column=0, pady=10, padx=10, sticky="n")

        self.reformat_button = tk.Button(
            self.reformat_frame, 
            text="Reformat Selected File",
            command=self.reformat_file,
            font=("Arial", 12, "bold"),
            pady=10,
            padx=85
        )
        self.reformat_button.pack(padx=20, pady=10)

        # Add a new frame at the bottom for version and author info
        self.bottom_frame = tk.Frame(self.root)
        self.bottom_frame.grid(row=999, column=0, columnspan=3, sticky="ew", pady=(10,5))
        
        # Configure column weights to create proper spacing
        self.bottom_frame.grid_columnconfigure(1, weight=1)  # This creates space between the labels
        
        # Add author label (left aligned)
        self.author_label = tk.Label(
            self.bottom_frame, 
            text=self.author, 
            font=("Arial", 8)
        )
        self.author_label.grid(row=0, column=0, sticky="w", padx=5)
        
        # Add version label (right aligned)
        self.version_label = tk.Label(
            self.bottom_frame, 
            text=f"Version Number: {self.version}", 
            font=("Arial", 8)
        )
        self.version_label.grid(row=0, column=2, sticky="e", padx=5)

        # Location & Event Section inside Training Parameters
        self.team_event_frame = tk.LabelFrame(self.training_parameters_frame, text="Event & Location", padx=52, pady=10)
        self.team_event_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        # Event Selection (Dropdown Menu)
        tk.Label(self.team_event_frame, text="Event: ").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.event_menu = tk.OptionMenu(self.team_event_frame, self.event_var, "SL", "GS", "SG", "DH", "SX")
        self.event_menu.grid(row=0, column=1, padx=5, pady=5, sticky='w')

        # Hill Entry with Autocomplete
        tk.Label(self.team_event_frame, text="Hill: ").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.hill_entry = tk.Entry(self.team_event_frame, textvariable=self.hill_var)
        self.hill_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        # Create the autocomplete listbox for hills
        self.hill_autocomplete_listbox = tk.Listbox(
            self.team_event_frame,
            height=0,
            borderwidth=1,
            relief="solid"
        )
        self.hill_autocomplete_listbox.grid(row=2, column=1, padx=(5, 5), pady=(0, 5), sticky="w")
        self.hill_autocomplete_listbox.grid_remove()  # Hide by default

        # Bind the events
        self.hill_entry.bind('<KeyRelease>', self.autocomplete_hill_name)
        self.hill_autocomplete_listbox.bind("<<ListboxSelect>>", self.on_hill_suggestion_select)

        # Right Side: Section with a thick border
        self.right_frame_with_border = tk.Frame(self.root, bd=4, relief="solid", padx=10, pady=10)
        self.right_frame_with_border.grid(row=1, column=2, rowspan=3, padx=10, pady=10, sticky="n")

        # Move everything into the bordered frame
        self.right_frame = tk.Frame(self.right_frame_with_border)
        self.right_frame.grid(row=0, column=0, padx=10, pady=10, sticky="n")

        # New Team Selector
        self.new_team_label = tk.Label(self.right_frame, text="Select Team", font=("Arial", 12, "bold"))
        self.new_team_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))

        self.team_button_frame = tk.Frame(self.right_frame)
        self.team_button_frame.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="w")

        self.sqah_button = tk.Button(self.team_button_frame, text="SQAH", width=10, command=lambda: self.set_team("SQAH"))
        self.sqah_button.pack(side=tk.LEFT, padx=5)

        self.sqaf_button = tk.Button(self.team_button_frame, text="SQAF", width=10, command=lambda: self.set_team("SQAF"))
        self.sqaf_button.pack(side=tk.LEFT, padx=5)

        self.other_button = tk.Button(self.team_button_frame, text="OTHER", width=10, command=lambda: self.set_team("OTHER"))
        self.other_button.pack(side=tk.LEFT, padx=5)

        # Athlete Listbox
        self.athlete_list_label = tk.Label(self.right_frame, text="Current Athletes:")
        self.athlete_list_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")

        self.athlete_listbox = tk.Listbox(self.right_frame, width=40, height=10)
        self.athlete_listbox.grid(row=3, column=0, padx=10, pady=5)

        # Temporary Guests Listbox
        self.guest_list_label = tk.Label(self.right_frame, text="Temporary Guests:")
        self.guest_list_label.grid(row=4, column=0, padx=10, pady=5, sticky="w")

        self.guest_listbox = tk.Listbox(self.right_frame, width=40, height=5)
        self.guest_listbox.grid(row=5, column=0, padx=10, pady=5)

        # Manage Athletes Section
        self.manage_athletes_frame = tk.Frame(self.right_frame, bd=2, relief="groove", padx=10, pady=10)
        self.manage_athletes_frame.grid(row=6, column=0, padx=10, pady=10, sticky="n")

        self.manage_athletes_label = tk.Label(self.manage_athletes_frame, text="Manage Athletes", font=("Arial", 12, "bold"))
        self.manage_athletes_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))

        self.athlete_name_label = tk.Label(self.manage_athletes_frame, text="Athlete Name:")
        self.athlete_name_label.grid(row=1, column=0, padx=(10, 5), pady=5, sticky="e")

        self.athlete_name_entry = tk.Entry(self.manage_athletes_frame)
        self.athlete_name_entry.grid(row=1, column=1, padx=(0, 5), pady=5, sticky="w")

        # Update the autocomplete listbox creation
        self.autocomplete_listbox = tk.Listbox(
            self.manage_athletes_frame,
            height=0,
            borderwidth=1,
            relief="solid"
        )
        self.autocomplete_listbox.grid(row=2, column=1, padx=(0, 5), pady=(0, 5), sticky="ew")  # Changed sticky to "ew"
        self.autocomplete_listbox.grid_remove()  # Hide by default

        self.bib_number_label = tk.Label(self.manage_athletes_frame, text="Bib Number:")
        self.bib_number_label.grid(row=2, column=0, padx=(10, 5), pady=5, sticky="e")

        self.bib_number_entry = tk.Entry(self.manage_athletes_frame)
        self.bib_number_entry.grid(row=2, column=1, padx=(0, 5), pady=5, sticky="w")

        # Buttons for athlete management
        self.button_frame = tk.Frame(self.manage_athletes_frame)
        self.button_frame.grid(row=3, column=0, columnspan=2, pady=10)

        button_width = 20
        self.add_athlete_button = tk.Button(self.button_frame, text="Add Athlete to Team", command=self.add_athlete, width=button_width)
        self.add_athlete_button.grid(row=0, column=0, pady=5)

        self.add_guest_button = tk.Button(self.button_frame, text="Add Temporary Guest", command=self.add_guest, width=button_width)
        self.add_guest_button.grid(row=1, column=0, pady=5)

        self.remove_athlete_button = tk.Button(self.button_frame, text="Remove Selected Athlete", command=self.remove_selected_athlete, width=button_width)
        self.remove_athlete_button.grid(row=2, column=0, pady=5)

        # Weather Conditions Section inside Training Parameters
        self.weather_frame = tk.LabelFrame(self.training_parameters_frame, text="Weather Conditions", padx=10, pady=10)
        self.weather_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

        # Snow Conditions
        tk.Label(self.weather_frame, text="Snow: ").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        snow_conditions = ["Very Soft", "Soft", "Medium", "Hard", "Injected"]
        self.snow_condition_menu = Combobox(self.weather_frame, textvariable=self.snow_condition_var, values=snow_conditions, width=17)
        self.snow_condition_menu.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.snow_condition_menu.set("Hard")  # Default value

        # Sky Conditions
        tk.Label(self.weather_frame, text="Sky: ").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        sky_conditions = ["Clear Sky", "Partly Cloudy", "Overcast", "Low Visibility", "Fog"]
        self.sky_condition_menu = Combobox(self.weather_frame, textvariable=self.sky_condition_var, values=sky_conditions, width=17)
        self.sky_condition_menu.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        self.sky_condition_menu.set("Clear Sky")  # Default value

        # Precipitation
        tk.Label(self.weather_frame, text="Precipitation: ").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        precipitation_types = ["No Precipitation", "Light Snow", "Medium Snow", "Snowstorm", "Light Rain", "Moderate Rain", "Downpour"]
        self.precipitation_menu = Combobox(self.weather_frame, textvariable=self.precipitation_var, values=precipitation_types, width=17)
        self.precipitation_menu.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        self.precipitation_menu.set("No Precipitation")  # Corrected from "No Precipitations"

        # Wind Conditions
        tk.Label(self.weather_frame, text="Wind: ").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        wind_conditions = ["No Wind", "Light Wind", "Moderate Wind", "Heavy Wind"]
        self.wind_condition_menu = Combobox(self.weather_frame, textvariable=self.wind_condition_var, values=wind_conditions, width=17)
        self.wind_condition_menu.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        self.wind_condition_menu.set("No Wind")  # Default value

        # Add File Details Section inside Training Parameters
        self.file_details_frame = tk.LabelFrame(self.training_parameters_frame, text="File Details", padx=10, pady=10)
        self.file_details_frame.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

        # Date Field
        tk.Label(self.file_details_frame, text="Date:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.date_entry = tk.Entry(self.file_details_frame, textvariable=self.date_var, width=20, state='readonly')
        self.date_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')

        # Start Time Field
        tk.Label(self.file_details_frame, text="Start Time:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.time_entry = tk.Entry(self.file_details_frame, textvariable=self.time_var, width=20, state='readonly')
        self.time_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        # Session Number Field
        tk.Label(self.file_details_frame, text="Session #:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.session_entry = tk.Entry(self.file_details_frame, textvariable=self.session_var, width=20, state='readonly')
        self.session_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')

    def detect_outliers(self, times):
        """
        Detects outlier times using standard deviation method.
        Returns a tuple of (mean, lower_bound, upper_bound).
        """
        if not times or len(times) < 3:  # Need at least 3 times for meaningful statistics
            return None, None, None
            
        mean_time = sum(times) / len(times)
        std_dev = (sum((x - mean_time) ** 2 for x in times) / len(times)) ** 0.5
        
        lower_bound = mean_time - (self.outlier_threshold * std_dev)
        upper_bound = mean_time + (self.outlier_threshold * std_dev)
        
        return mean_time, lower_bound, upper_bound

    def is_outlier(self, time_value, times):
        """
        Determines if a time is an outlier.
        """
        mean, lower_bound, upper_bound = self.detect_outliers(times)
        if mean is None:  # Not enough data points
            return False
        return time_value < lower_bound or time_value > upper_bound

    def parse_csv_file(self, file_path):
        """
        Parses the CSV file to extract session details.
        Updated to handle the custom separator.
        """
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                
                # Initialize variables
                session_num = ""
                date_str = ""
                time_str = ""
                
                # Parse the header section with custom separator
                for line in lines:
                    parts = line.strip().split('>')
                    
                    if len(parts) < 2:
                        continue
                        
                    if "Session #" in parts[0]:
                        session_num = parts[1].strip()
                    elif "Date" == parts[0]:
                        date_str = parts[1].strip()
                    elif "Time" == parts[0]:
                        time_str = parts[1].strip()
                    
                    # Break once we've found all needed information
                    if session_num and date_str and time_str:
                        break
                
                # Convert date format if found
                if date_str:
                    try:
                        # Parse the date (assuming mm/dd/yy format)
                        date_obj = datetime.strptime(date_str, '%m/%d/%y')
                        # Convert to dd/mm/yyyy format
                        formatted_date = date_obj.strftime('%d/%m/%Y')
                    except ValueError:
                        formatted_date = date_str
                else:
                    formatted_date = ""
                
                return {
                    'session': session_num,
                    'date': formatted_date,
                    'time': time_str
                }
                
        except Exception as e:
            messagebox.showerror("Error", f"Error reading file: {str(e)}")
            return None

    def select_file(self):
        """Opens a file dialog to select a CSV file and updates file details."""
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if file_path:
            self.selected_file = file_path  # Store the full path
            display_path = file_path if len(file_path) <= 35 else f"...{file_path[-33:]}"
            self.file_label.config(text=f"Selected File: {display_path}")
            
            # Parse the file and extract details
            file_details = self.parse_csv_file(file_path)
            if file_details:
                self.date_var.set(file_details['date'])
                self.time_var.set(file_details['time'])
                self.session_var.set(file_details['session'])
            else:
                # Clear the fields if parsing failed
                self.date_var.set("")
                self.time_var.set("")
                self.session_var.set("")

    # Athlete Data Management Methods
    def add_athlete_to_memory(self, athlete_name):
        """Adds an athlete name to recent names memory, ensuring a maximum of 2000 entries."""
        if athlete_name in self.recent_names:
            self.recent_names.remove(athlete_name)
        self.recent_names.append(athlete_name)
        if len(self.recent_names) > 2000:
            self.recent_names = self.recent_names[-2000:]  # Keep the last 2000 names
        self.save_recent_names()

    def save_recent_names(self):
        """Saves the list of recent names to a text file."""
        with open("recent_names.txt", "w", encoding='utf-8') as f:
            for name in self.recent_names:
                f.write(name + '\n')

    def load_recent_names(self):
        """Loads the list of recent names from a text file."""
        try:
            with open("recent_names.txt", "r", encoding='utf-8') as f:
                self.recent_names = [line.strip() for line in f.readlines()]
        except FileNotFoundError:
            self.recent_names = []

    def save_athletes_to_json(self):
        """Saves the athlete data to a JSON file."""
        with open("athletes_data.json", "w") as f:
            json.dump(self.athletes, f, indent=4)

    def load_athletes_from_json(self):
        """Loads the athlete data from a JSON file."""
        try:
            with open("athletes_data.json", "r") as f:
                self.athletes = json.load(f)
        except FileNotFoundError:
            self.athletes = {"SQAH": [], "SQAF": [], "OTHER": []}
        self.athletes["OTHER"] = []
        self.temp_guests = []

    # GUI Update Methods
    def update_athlete_listbox(self):
        """Update the athlete listbox with the current team's athletes, sorted by bib number."""
        self.athlete_listbox.delete(0, tk.END)  # Clear the listbox first

        # Check if the current team exists and has athletes
        if self.current_team in self.athletes:
            # Sort the athletes by bib number
            sorted_athletes = sorted(self.athletes[self.current_team], key=lambda x: int(x['bib']))

            # Insert sorted athletes into the listbox
            for athlete in sorted_athletes:
                self.athlete_listbox.insert(tk.END, f"{athlete['name']} (Bib {athlete['bib']})")

    def update_guest_listbox(self):
        """Update the guest listbox with the current temporary guests, marking inactive guests."""
        self.guest_listbox.delete(0, tk.END)
        for guest in self.temp_guests:
            guest_status = "(Inactive)" if guest.get("inactive", False) else ""
            self.guest_listbox.insert(tk.END, f"{guest['name']} (Bib {guest['bib']}) {guest_status}")

    def bind_deletion_keys(self):
        """Bind deletion keys to listboxes for removing athletes or guests."""
        self.athlete_listbox.bind('<Delete>', self.remove_selected_athlete)
        self.athlete_listbox.bind('<BackSpace>', self.remove_selected_athlete)
        self.guest_listbox.bind('<Delete>', self.remove_selected_athlete)
        self.guest_listbox.bind('<BackSpace>', self.remove_selected_athlete)

    def update_suggestion_box_width(self, event=None):
        """Updates the width of the suggestion box to match the Athlete Name entry."""
        # Configure the grid column to match the entry width
        self.manage_athletes_frame.grid_columnconfigure(1, minsize=self.athlete_name_entry.winfo_width())
    
   
    # Event Handlers and Animations
    def animate_listbox(self):
        """Handles the animation for the autocomplete listbox."""
        if self.is_animating:
            current_height = self.animation_height
            target_height = self.target_height

            if current_height < target_height:
                # Opening animation
                self.animation_height = min(current_height + self.animation_speed, target_height)
                self.autocomplete_listbox.configure(height=int(self.animation_height / self.line_height))
            elif current_height > target_height:
                # Closing animation
                self.animation_height = max(current_height - self.animation_speed, target_height)
                self.autocomplete_listbox.configure(height=int(self.animation_height / self.line_height))

            # Continue animation if not reached target
            if self.animation_height != target_height:
                self.root.after(10, self.animate_listbox)
            else:
                self.is_animating = False
                if target_height == 0:
                    self.autocomplete_listbox.grid_remove()

    def autocomplete_athlete_name(self, event):
        """Provides autocomplete suggestions for athlete names based on recent entries."""
        input_text = self.athlete_name_entry.get().lower()

        # Get entry widget width if we haven't stored it
        if not self.name_entry_width:
            self.update_suggestion_box_width()
            # Get line height for calculations
            self.line_height = self.athlete_name_entry.winfo_reqheight()
            self.autocomplete_listbox.configure(width=self.name_entry_width)

        if not input_text:
            # Animate closing if input is empty
            self.target_height = 0
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()
            return

        # Filter recent names that start with the input text
        suggestions = [name for name in reversed(self.recent_names) if name.lower().startswith(input_text)]
        suggestions = suggestions[:2]  # Limit to 2 suggestions

        # Update the Listbox with suggestions
        if suggestions:
            self.autocomplete_listbox.delete(0, tk.END)
            for suggestion in suggestions:
                self.autocomplete_listbox.insert(tk.END, suggestion)

            # Update the suggestion box width
            self.update_suggestion_box_width()

            # Calculate target height based on number of suggestions
            num_suggestions = len(suggestions)
            self.target_height = num_suggestions * self.line_height

            # Show and position the listbox if it's hidden
            if not self.autocomplete_listbox.winfo_viewable():
                self.autocomplete_listbox.grid(row=2, column=1, padx=(0, 5), pady=(0, 5), sticky="w")
                self.animation_height = 0

            # Start animation if not already animating
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()

            self.autocomplete_listbox.lift()
        else:
            # Animate closing if no suggestions
            self.target_height = 0
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()

    def on_suggestion_select(self, event):
        """Handles the selection of a suggestion from the autocomplete listbox."""
        if self.autocomplete_listbox.curselection():
            selected_suggestion = self.autocomplete_listbox.get(self.autocomplete_listbox.curselection())
            self.athlete_name_entry.delete(0, tk.END)
            self.athlete_name_entry.insert(0, selected_suggestion)

            # Animate closing
            self.target_height = 0
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()

    # Athlete Management Methods
    def add_athlete(self):
        """Adds a new athlete to the current team."""
        athlete_name = self.athlete_name_entry.get()
        bib_number = self.bib_number_entry.get()

        if athlete_name and bib_number and self.current_team:
            if not bib_number.isdigit():
                messagebox.showwarning("Input Error", "Bib invalid")
                return

            bib_number = int(bib_number)  # Ensure bib_number is an integer

            # Check for duplicate bib in the current team's athlete list
            for athlete in self.athletes[self.current_team]:
                if int(athlete['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to an athlete in this team.")
                    return

            # Check for duplicate bib in the temporary guest list
            for guest in self.temp_guests:
                if int(guest['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to a temporary guest.")
                    return

            # Add the new athlete if no duplicates found
            new_athlete = {"name": athlete_name, "bib": bib_number}
            self.athletes[self.current_team].append(new_athlete)
            self.save_athletes_to_json()
            self.update_athlete_listbox()

            # Add athlete name to recent names
            self.add_athlete_to_memory(athlete_name)

            # Clear the fields after adding
            self.athlete_name_entry.delete(0, tk.END)
            self.bib_number_entry.delete(0, tk.END)
        else:
            messagebox.showwarning("Input Error", "Please fill all fields before adding an athlete.")

    def add_guest(self):
        """Adds a new temporary guest."""
        guest_name = self.athlete_name_entry.get()
        bib_number = self.bib_number_entry.get()

        if guest_name and bib_number:
            if not bib_number.isdigit():
                messagebox.showwarning("Input Error", "Bib invalid")
                return

            bib_number = int(bib_number)  # Ensure bib_number is an integer

            # Check for duplicate bib in the current team's athlete list
            for athlete in self.athletes[self.current_team]:
                if int(athlete['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to an athlete in this team.")
                    return

            # Check for duplicate bib in the temporary guest list
            for guest in self.temp_guests:
                if int(guest['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to a temporary guest.")
                    return

            # Add the new guest if no duplicates found
            new_guest = {"name": guest_name, "bib": bib_number}
            self.temp_guests.append(new_guest)
            self.update_guest_listbox()

            # Add guest name to recent names
            self.add_athlete_to_memory(guest_name)

            # Clear the fields after adding
            self.athlete_name_entry.delete(0, tk.END)
            self.bib_number_entry.delete(0, tk.END)
        else:
            messagebox.showwarning("Input Error", "Please fill all fields before adding a guest.")

    def remove_selected_athlete(self, event=None):
        """Removes the selected athlete or guest from the list."""
        selected_team_athlete = self.athlete_listbox.curselection()
        selected_guest = self.guest_listbox.curselection()

        if selected_team_athlete and self.current_team:
            if messagebox.askyesno("Confirm Deletion", "Are you sure you want to remove this athlete?"):
                del self.athletes[self.current_team][selected_team_athlete[0]]
                self.save_athletes_to_json()  # Save changes to JSON
                self.update_athlete_listbox()
        elif selected_guest:
            if messagebox.askyesno("Confirm Deletion", "Are you sure you want to remove this guest?"):
                del self.temp_guests[selected_guest[0]]
                self.update_guest_listbox()
        else:
            messagebox.showwarning("Selection Error", "No athlete or guest selected to remove.")

    def check_guest_conflicts_with_athletes(self, new_team):
        """Checks for duplicate bibs between temporary guests and the athletes in the selected team.
        Marks guests as inactive if duplicates are found."""
        conflicts = False
        for guest in self.temp_guests:
            guest_bib = int(guest["bib"])

            # Check for duplicate bibs in the new team's athlete list
            duplicate_found = any(int(athlete['bib']) == guest_bib for athlete in self.athletes[new_team])

            # If a duplicate is found, mark the guest as inactive
            if duplicate_found:
                guest["inactive"] = True
                conflicts = True
            else:
                # If no duplicates are found in the new team, remove the inactive status
                guest["inactive"] = False
        # If conflicts are detected, show a warning
        if conflicts:
            messagebox.showwarning("Bib Conflict", "Some Bibs are duplicates. Guests will not be active.")

    # Other Methods
    def set_team(self, team_name):
        """Changes the selected team and updates the GUI accordingly."""
        self.current_team = team_name
        self.sqah_button.config(bg="SystemButtonFace")
        self.sqaf_button.config(bg="SystemButtonFace")
        self.other_button.config(bg="SystemButtonFace")

        if team_name == "SQAH":
            self.sqah_button.config(bg="lightblue")
        elif team_name == "SQAF":
            self.sqaf_button.config(bg="lightblue")
        elif team_name == "OTHER":
            self.other_button.config(bg="lightblue")

        # Recheck guest conflicts and remove "Inactive" if no duplicates found
        self.check_guest_conflicts_with_athletes(team_name)

        # Update the athlete and guest listboxes
        self.update_athlete_listbox()
        self.update_guest_listbox()

    # Add these methods to the TimingSystemApp class:

    def add_hill_to_memory(self, hill_name):
        """Adds a hill name to recent hills memory, ensuring a maximum of 2000 entries."""
        if hill_name in self.recent_hills:
            self.recent_hills.remove(hill_name)
        self.recent_hills.append(hill_name)
        if len(self.recent_hills) > 2000:
            self.recent_hills = self.recent_hills[-2000:]  # Keep the last 2000 names
        self.save_recent_hills()

    def save_recent_hills(self):
        """Saves the list of recent hills to a text file."""
        with open("recent_hills.txt", "w", encoding='utf-8') as f:
            for name in self.recent_hills:
                f.write(name + '\n')

    def load_recent_hills(self):
        """Loads the list of recent hills from a text file."""
        try:
            with open("recent_hills.txt", "r", encoding='utf-8') as f:
                self.recent_hills = [line.strip() for line in f.readlines()]
        except FileNotFoundError:
            self.recent_hills = []

    def autocomplete_hill_name(self, event):
        """Provides autocomplete suggestions for hill names based on recent entries."""
        input_text = self.hill_entry.get().lower()

        if not input_text:
            # Animate closing if input is empty
            self.hill_target_height = 0
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()
            return

        # Filter recent names that start with the input text
        suggestions = [name for name in reversed(self.recent_hills) if name.lower().startswith(input_text)]
        suggestions = suggestions[:1]  # Limit to 1 suggestion

        # Update the Listbox with suggestions
        if suggestions:
            self.hill_autocomplete_listbox.delete(0, tk.END)
            for suggestion in suggestions:
                self.hill_autocomplete_listbox.insert(tk.END, suggestion)

            # Calculate target height based on number of suggestions
            self.hill_target_height = self.line_height

            # Show and position the listbox if it's hidden
            if not self.hill_autocomplete_listbox.winfo_viewable():
                self.hill_autocomplete_listbox.grid()
                self.hill_animation_height = 0

            # Start animation if not already animating
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()

            self.hill_autocomplete_listbox.lift()
        else:
            # Animate closing if no suggestions
            self.hill_target_height = 0
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()

    def animate_hill_listbox(self):
        """Handles the animation for the hill autocomplete listbox."""
        if self.hill_is_animating:
            current_height = self.hill_animation_height
            target_height = self.hill_target_height

            if current_height < target_height:
                # Opening animation
                self.hill_animation_height = min(current_height + self.animation_speed, target_height)
                self.hill_autocomplete_listbox.configure(height=int(self.hill_animation_height / self.line_height))
            elif current_height > target_height:
                # Closing animation
                self.hill_animation_height = max(current_height - self.animation_speed, target_height)
                self.hill_autocomplete_listbox.configure(height=int(self.hill_animation_height / self.line_height))

            # Continue animation if not reached target
            if self.hill_animation_height != target_height:
                self.root.after(10, self.animate_hill_listbox)
            else:
                self.hill_is_animating = False
                if target_height == 0:
                    self.hill_autocomplete_listbox.grid_remove()

    def on_hill_suggestion_select(self, event):
        """Handles the selection of a suggestion from the hill autocomplete listbox."""
        if self.hill_autocomplete_listbox.curselection():
            selected_suggestion = self.hill_autocomplete_listbox.get(self.hill_autocomplete_listbox.curselection())
            self.hill_entry.delete(0, tk.END)
            self.hill_entry.insert(0, selected_suggestion)
            self.add_hill_to_memory(selected_suggestion)

            # Animate closing
            self.hill_target_height = 0
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()

    def get_athlete_name(self, bib_number):
        """
        Given a bib number, returns the athlete's name from all teams' athletes or temp guests.
        """
        bib_number = int(bib_number)
        # Check in all teams
        for team in self.athletes:
            for athlete in self.athletes[team]:
                if int(athlete['bib']) == bib_number:
                    return athlete['name']
        # Then, check in temp guests
        for guest in self.temp_guests:
            if int(guest['bib']) == bib_number and not guest.get('inactive', False):
                return guest['name']
        # If not found, return 'Unknown Athlete'
        return 'Unknown Athlete'

    def parse_timing_data(self, file_path):
        """
        Parses the CSV file to extract timing data, organized by runs.
        Handles custom separator and header format.
        """
        timing_data = {}
        header_found = False
        column_indices = {}
        
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                
                # First determine the separator
                separator = '>'  # Based on the sample file format
                
                for line in lines:
                    line = line.strip()
                    if not line:  # Skip empty lines
                        continue
                    
                    # Look for the header line that contains both Bib# and Run#
                    if not header_found and "Bib#" in line and "Run#" in line:
                        header_found = True
                        # Get column indices from header
                        headers = line.split(separator)
                        for i, header in enumerate(headers):
                            header = header.lower().strip()
                            if "bib" in header:
                                column_indices['bib'] = i
                            elif "run" in header:
                                column_indices['run'] = i
                            elif "split 1" in header:
                                column_indices['split1'] = i
                            elif "finish time" in header:
                                column_indices['finish'] = i
                            elif "status" in header:
                                column_indices['status'] = i
                        continue
                    
                    # Process data lines after header is found
                    if header_found:
                        data = line.split(separator)
                        if len(data) <= 1:  # Skip empty lines
                            continue
                        
                        try:
                            # Get run number and clean it
                            run_number = data[column_indices['run']].strip()
                            if not run_number:  # Skip if run number is empty
                                continue
                            
                            if run_number not in timing_data:
                                timing_data[run_number] = []
                            
                            # For this format, Split 1 is the intermediate time
                            # and Finish Time is the total time
                            split1_time = data[column_indices['split1']].strip()
                            status = data[column_indices.get('status', -1)].strip() if 'status' in column_indices else ''
                            
                            # Create entry for this run
                            entry = {
                                'bib': data[column_indices['bib']].strip(),
                                'split1': self.parse_time(split1_time),
                                'finish': self.parse_time(data[column_indices.get('finish', -1)].strip() if 'finish' in column_indices else ''),
                                'status': status if status else '',
                                'run': run_number
                            }
                            
                            # Only add valid entries
                            if entry['bib'] and entry['run']:
                                timing_data[run_number].append(entry)
                                
                        except IndexError as e:
                            print(f"Error processing line: {line}")
                            print(f"Error details: {str(e)}")
                            continue
            
            # Sort data for each run
            for run_number in timing_data:
                timing_data[run_number] = self.sort_run_data(timing_data[run_number])
            
            if not timing_data:
                messagebox.showerror("Error", "No valid timing data found in the file.")
                return None
                
            return timing_data
            
        except Exception as e:
            messagebox.showerror("Error", f"Error parsing CSV file: {str(e)}")
            return None

    def parse_time(self, time_str):
        """
        Parses a time string into a float value in seconds.
        Returns None for invalid times or DNF/DSQ.
        """
        if not time_str or not isinstance(time_str, str):
            return None
            
        time_str = time_str.strip().upper()
        if time_str in ['DNF', 'DSQ', '', 'DNS']:
            return None
            
        try:
            if ':' in time_str:
                minutes, seconds = time_str.split(':')
                return float(minutes) * 60 + float(seconds)
            return float(time_str)
        except (ValueError, TypeError):
            return None

    def create_formatted_excel(self, output_path):
        """
        Creates a formatted Excel file with the specified header layout and timing data.
        """
        # Create new workbook and get active sheet
        wb = Workbook()
        ws = wb.active
        
        # Define fonts
        try:
            title_font = Font(name='Avenir Next LT Pro', size=18, bold=True)
        except:
            title_font = Font(name='Arial', size=18, bold=True)
        
        normal_font = Font(name='Arial', size=11)
        header_font = Font(name='Arial', size=11, bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 17
        ws.column_dimensions['C'].width = 21
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 6
        
        # Set row height
        ws.row_dimensions[1].height = 30
        
        # Title in B1
        ws['B1'] = "Training SQA Équipe du Québec"
        ws['B1'].font = title_font
        ws['B1'].alignment = Alignment(vertical='center', horizontal='left')

        # Define labels and their corresponding values
        label_value_pairs = {
            ('B2', 'Team:'): self.current_team,
            ('B3', 'Session #:'): self.session_var.get(),
            ('B4', 'Event:'): self.event_var.get(),
            ('B5', 'Snow Condition:'): self.snow_condition_var.get(),
            ('E2', 'Date:'): self.date_var.get(),
            ('E3', 'Start Time:'): self.time_var.get(),
            ('E4', 'Hill:'): self.hill_var.get(),
        }

        # Write labels and values with proper formatting
        for (cell_coord, label), value in label_value_pairs.items():
            # Write label
            ws[cell_coord] = label
            ws[cell_coord].font = header_font
            ws[cell_coord].alignment = Alignment(horizontal='left')
            
            # Write value in the next column
            value_col = chr(ord(cell_coord[0]) + 1)  # Next column
            value_cell = f"{value_col}{cell_coord[1]}"
            ws[value_cell] = value
            ws[value_cell].font = normal_font
            ws[value_cell].alignment = Alignment(horizontal='left')

        # Special handling for Weather row (combines multiple conditions)
        ws['E5'] = 'Weather:'
        ws['E5'].font = header_font
        ws['E5'].alignment = Alignment(horizontal='left')
        
        weather_value = f"{self.sky_condition_var.get()}, {self.precipitation_var.get()}, {self.wind_condition_var.get()}"
        ws['F5'] = weather_value
        ws['F5'].font = normal_font
        ws['F5'].alignment = Alignment(horizontal='left')


        # Now parse and add timing data
        if self.selected_file:
            timing_data = self.parse_timing_data(self.selected_file)
            if timing_data:
                current_row = 8  # Start after header section
                
                # Write each run's data
                for run_number in sorted(timing_data.keys(), key=int):
                    current_row = self.write_run_data(ws, timing_data[run_number], current_row)
                    current_row += 1  # Extra space between runs
        # After writing all run data, add the analysis graphs
        if self.selected_file:
            timing_data = self.parse_timing_data(self.selected_file)
            if timing_data:
                current_row = self.add_analysis_graphs(ws, timing_data, current_row)
        
        try:
            wb.save(output_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {str(e)}")
            return False      



    def format_time(self, time_value, as_difference=False):
        """
        Formats a time value (in seconds) as a string.
        For differences, includes + sign and ensures 2 decimal places.
        """
        if time_value is None:
            return ""
            
        # For differences
        if as_difference:
            sign = '+' if time_value >= 0 else ''
            if abs(time_value) < 60:
                return f"{sign}{time_value:.2f}"
            minutes = int(abs(time_value) // 60)
            seconds = abs(time_value) % 60
            return f"{sign}{minutes}:{seconds:05.2f}"
        
        # For regular times
        if time_value < 60:
            return f"{time_value:.2f}"
        minutes = int(time_value // 60)
        seconds = time_value % 60
        return f"{minutes}:{seconds:05.2f}"

    def sort_run_data(self, run_data):
        """
        Sorts run data first by status (valid times first), then by finish time.
        """
        def sort_key(entry):
            # Primary sort: DNF/DSQ at the end
            is_valid = entry['status'].upper() not in ['DNF', 'DSQ']
            # Secondary sort: Finish time (None values at the end)
            finish_time = entry['finish'] if entry['finish'] is not None else float('inf')
            return (-is_valid, finish_time)
        
        return sorted(run_data, key=sort_key)

    def write_run_data(self, ws, run_data, start_row):
        """
        Modified write_run_data method with outlier detection and highlighting.
        """
        # Define styles
        light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        light_red_fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
        header_border = Border(
            left=Side(style='thick'), 
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        normal_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write run number
        current_row = start_row + 2
        ws[f'B{current_row}'] = f"Run {run_data[0]['run']}"
        current_row += 1
        
        # Write headers
        headers = ['Bib #', 'Name', 'Split 1', 'Split Diff.', 'Rank', 'Finish Time', 'Finish Split', 'Status']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Collect valid times and detect outliers
        valid_splits = [entry['split1'] for entry in run_data if entry['split1'] is not None]
        valid_finishes = [entry['finish'] for entry in run_data if entry['finish'] is not None 
                         and entry['status'].upper() not in ['DNF', 'DSQ']]
        
        # Find best non-outlier times
        non_outlier_splits = [t for t in valid_splits if not self.is_outlier(t, valid_splits)]
        non_outlier_finishes = [t for t in valid_finishes if not self.is_outlier(t, valid_finishes)]
        
        best_split = min(non_outlier_splits) if non_outlier_splits else None
        best_finish = min(non_outlier_finishes) if non_outlier_finishes else None
        
        # Calculate split ranks (excluding outliers)
        split_ranks = {}
        sorted_splits = sorted((entry['split1'], i) for i, entry in enumerate(run_data) 
                             if entry['split1'] is not None and not self.is_outlier(entry['split1'], valid_splits))
        for rank, (_, index) in enumerate(sorted_splits, 1):
            split_ranks[index] = rank
        
        # Write data rows
        for i, entry in enumerate(run_data):
            split_time = entry['split1']
            finish_time = entry['finish']
            
            # Check for outliers
            split_is_outlier = split_time is not None and self.is_outlier(split_time, valid_splits)
            finish_is_outlier = finish_time is not None and self.is_outlier(finish_time, valid_finishes)
            
            row = [
                entry['bib'],
                self.get_athlete_name(entry['bib']),
                self.format_time(split_time),
                self.format_time(split_time - best_split if split_time is not None and best_split is not None else None, True),
                split_ranks.get(i, ''),
                self.format_time(finish_time),
                self.format_time(finish_time - best_finish if finish_time is not None and best_finish is not None else None, True),
                entry['status']
            ]
            
            for col, value in enumerate(row, start=2):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = normal_border
                cell.alignment = Alignment(horizontal='center')
                
                # Apply highlighting
                if col == 4:  # Split 1 column
                    if split_is_outlier:
                        cell.fill = light_red_fill
                    elif split_time == best_split:
                        cell.fill = light_green_fill
                elif col == 7:  # Finish time column
                    if finish_is_outlier:
                        cell.fill = light_red_fill
                    elif finish_time == best_finish:
                        cell.fill = light_green_fill
            
            current_row += 1
        
        return current_row

    def reformat_file(self):
        """Handle the reformatting of the selected file with validation checks."""
        # Check if a file is selected
        if not hasattr(self, 'file_label') or self.file_label.cget("text") == "No file selected":
            messagebox.showwarning("Error", "Please select a Brower CSV file first.")
            return

        # Check if an event is selected
        if not self.event_var.get():
            messagebox.showwarning("Error", "Please select an Event type (SL/GS/SG/DH/SX).")
            return

        # Check if a hill is specified
        if not self.hill_var.get().strip():
            messagebox.showwarning("Error", "Please specify a Hill name.")
            return

        # Add the hill to recent hills memory for future autocomplete
        hill_name = self.hill_var.get().strip()
        self.add_hill_to_memory(hill_name)

        try:
            # Ask user where to save the reformatted file
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="reformatted_timing_data.xlsx"
            )
            
            if output_file:  # If user didn't cancel the save dialog
                if self.create_formatted_excel(output_file):
                    messagebox.showinfo("Success", f"File has been reformatted and saved successfully.")
                else:
                    messagebox.showerror("Error", "Failed to create reformatted file.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def clean_and_analyze_timing_data(self, run_data):
        """
        Cleans timing data by removing outliers and returns valid data for analysis.
        
        Args:
            run_data (list): List of timing entries for a run
        
        Returns:
            tuple: (cleaned_data, outliers) - Lists of valid entries and detected outliers
        """
        # Collect all valid split and finish times
        valid_splits = []
        valid_finishes = []
        
        for entry in run_data:
            if entry['split1'] is not None and entry['status'].upper() not in ['DNF', 'DSQ', 'DNS']:
                valid_splits.append(entry['split1'])
            if entry['finish'] is not None and entry['status'].upper() not in ['DNF', 'DSQ', 'DNS']:
                valid_finishes.append(entry['finish'])
        
        # Calculate statistics for valid times
        if valid_splits:
            split_mean = sum(valid_splits) / len(valid_splits)
            split_std = (sum((x - split_mean) ** 2 for x in valid_splits) / len(valid_splits)) ** 0.5
        else:
            split_mean = split_std = 0
            
        if valid_finishes:
            finish_mean = sum(valid_finishes) / len(valid_finishes)
            finish_std = (sum((x - finish_mean) ** 2 for x in valid_finishes) / len(valid_finishes)) ** 0.5
        else:
            finish_mean = finish_std = 0
        
        # Define acceptable ranges (mean ± 2 standard deviations)
        split_range = (split_mean - 2 * split_std, split_mean + 2 * split_std)
        finish_range = (finish_mean - 2 * finish_std, finish_mean + 2 * finish_std)
        
        # Separate valid data and outliers
        cleaned_data = []
        outliers = []
        
        for entry in run_data:
            is_outlier = False
            if entry['split1'] is not None:
                if entry['split1'] < split_range[0] or entry['split1'] > split_range[1]:
                    is_outlier = True
            if entry['finish'] is not None:
                if entry['finish'] < finish_range[0] or entry['finish'] > finish_range[1]:
                    is_outlier = True
            
            if is_outlier:
                outliers.append(entry)
            else:
                cleaned_data.append(entry)
        
        return cleaned_data, outliers

    def add_analysis_graphs(self, ws, timing_data, start_row):
        """
        Adds analysis graphs to the Excel worksheet showing split and finish time comparisons.
        """
        from openpyxl.chart import LineChart, Reference, Series
        from openpyxl.chart.marker import Marker

        current_row = start_row + 2

        # Write the analysis title
        title_cell = ws.cell(row=current_row, column=2)
        title_cell.value = "Timing Analysis"
        title_cell.font = Font(size=14, bold=True)
        current_row += 2

        # Prepare data for graphs
        run_numbers = sorted(timing_data.keys(), key=int)
        athletes = {}

        # First pass: collect all athlete data
        for run_num in run_numbers:
            run_data = timing_data[run_num]
            cleaned_data, _ = self.clean_and_analyze_timing_data(run_data)
            
            for entry in cleaned_data:
                if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS']:
                    bib = entry['bib']
                    athlete_name = self.get_athlete_name(bib)
                    
                    if athlete_name not in athletes:
                        athletes[athlete_name] = {'splits': {}, 'finishes': {}}
                    
                    if entry['split1'] is not None:
                        athletes[athlete_name]['splits'][run_num] = entry['split1']
                    if entry['finish'] is not None:
                        athletes[athlete_name]['finishes'][run_num] = entry['finish']

        if not athletes:
            print("No valid athlete data found for graphs")
            return current_row

        # Write data headers
        data_start_row = current_row
        ws.cell(row=current_row, column=2, value="Run")
        col = 3
        for athlete in athletes:
            ws.cell(row=current_row, column=col, value=f"{athlete} Split")
            ws.cell(row=current_row, column=col + 1, value=f"{athlete} Finish")
            col += 2
        current_row += 1

        # Write data rows
        for run_num in run_numbers:
            ws.cell(row=current_row, column=2, value=int(run_num))
            col = 3
            for athlete in athletes:
                split_time = athletes[athlete]['splits'].get(run_num)
                finish_time = athletes[athlete]['finishes'].get(run_num)
                
                ws.cell(row=current_row, column=col, value=split_time)
                ws.cell(row=current_row, column=col + 1, value=finish_time)
                
                # Format cells as numbers
                if split_time is not None:
                    ws.cell(row=current_row, column=col).number_format = '0.00'
                if finish_time is not None:
                    ws.cell(row=current_row, column=col + 1).number_format = '0.00'
                
                col += 2
            current_row += 1

        data_end_row = current_row - 1

        # Create split times chart
        split_chart = LineChart()
        split_chart.title = "Split Times Comparison"
        split_chart.style = 2
        split_chart.height = 15
        split_chart.width = 25
        split_chart.x_axis.title = "Run Number"
        split_chart.y_axis.title = "Split Time (seconds)"

        # Create finish times chart
        finish_chart = LineChart()
        finish_chart.title = "Finish Times Comparison"
        finish_chart.style = 2
        finish_chart.height = 15
        finish_chart.width = 25
        finish_chart.x_axis.title = "Run Number"
        finish_chart.y_axis.title = "Finish Time (seconds)"

        # Colors for different athletes
        colors = ['FF0000', '00FF00', '0000FF', 'FF00FF', '00FFFF', 'FFA500', '800080']

        # Add data series for each athlete
        col = 3
        for i, athlete in enumerate(athletes):
            # Split times series
            split_values = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=col)
            split_series = Series(split_values, title=f"{athlete} Split")
            split_series.marker = Marker(symbol='circle', size=7)
            split_series.smooth = True
            split_series.graphicalProperties.line.width = 20000
            split_series.graphicalProperties.line.solidFill = colors[i % len(colors)]
            split_chart.series.append(split_series)

            # Finish times series
            finish_values = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=col + 1)
            finish_series = Series(finish_values, title=f"{athlete} Finish")
            finish_series.marker = Marker(symbol='circle', size=7)
            finish_series.smooth = True
            finish_series.graphicalProperties.line.width = 20000
            finish_series.graphicalProperties.line.solidFill = colors[i % len(colors)]
            finish_chart.series.append(finish_series)

            col += 2

        # Add x-axis labels (run numbers)
        run_labels = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=2)
        split_chart.set_categories(run_labels)
        finish_chart.set_categories(run_labels)

        # Position the charts in the worksheet
        ws.add_chart(split_chart, f"B{current_row + 2}")
        ws.add_chart(finish_chart, f"B{current_row + 21}")

        return current_row + 40

    def reformat_file(self):
        """Handle the reformatting of the selected file with validation checks."""
        # Check if a file is selected
        if not hasattr(self, 'file_label') or self.file_label.cget("text") == "No file selected":
            messagebox.showwarning("Error", "Please select a Brower CSV file first.")
            return

        # Check if an event is selected
        if not self.event_var.get():
            messagebox.showwarning("Error", "Please select an Event type (SL/GS/SG/DH/SX).")
            return

        # Check if a hill is specified
        if not self.hill_var.get().strip():
            messagebox.showwarning("Error", "Please specify a Hill name.")
            return

        # Add the hill to recent hills memory for future autocomplete
        hill_name = self.hill_var.get().strip()
        self.add_hill_to_memory(hill_name)

        try:
            # Ask user where to save the reformatted file
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="reformatted_timing_data.xlsx"
            )
            
            if output_file:  # If user didn't cancel the save dialog
                if self.create_formatted_excel(output_file):
                    messagebox.showinfo("Success", f"File has been reformatted and saved successfully.")
                else:
                    messagebox.showerror("Error", "Failed to create reformatted file.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    # ... [Rest of the class methods] ...

if __name__ == "__main__":
    root = tk.Tk()
    app = TimingSystemApp(root)
    root.mainloop()