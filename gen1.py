import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side  # Add to existing imports
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

class TimingSystemApp:
    def __init__(self, root):
        # Initialization and GUI setup
        self.root = root
        self.root.title("Timing System Data Formatter")

        # Data structures
        self.selected_file = None
        self.athletes = {"SQAH": [], "SQAF": [], "OTHER": []}  # Store athletes by team
        self.temp_guests = []  # Temporary guests
        self.current_team = "SQAH"
        self.team_var = tk.StringVar()
        self.event_var = tk.StringVar()
        self.hill_var = tk.StringVar()
        self.name_entry_width = 0  # Will store the width of the entry
        self.animation_height = 0  # Current animation height
        self.target_height = 0     # Target height for animation
        self.animation_speed = 2   # Pixels per frame
        self.is_animating = False
        self.line_height = 20  # Default line height - will be updated after widget creation
        self.snow_condition_var = tk.StringVar()
        self.sky_condition_var = tk.StringVar()
        self.precipitation_var = tk.StringVar()
        self.wind_condition_var = tk.StringVar()
        self.date_var = tk.StringVar()
        self.time_var = tk.StringVar()
        self.session_var = tk.StringVar()
        self.outlier_threshold = 2  # Number of standard deviations for outlier detection
        # Hill-related attributes
        self.recent_hills = []
        self.hill_animation_height = 0
        self.hill_target_height = 0
        self.hill_is_animating = False


        # Version and author
        self.version = "1.0"
        self.author = "Julian H. Brunet"
        
        # Add new attributes for settings
        self.excel_title = "Training SQA Équipe du Québec"  # Default title
        self.team_names = {
            "SQAH": "SQAH",
            "SQAF": "SQAF"
        }
        self.default_hill = ""
        self.load_settings()
        self.selected_file = None
        self.athletes = {"SQAH": [], "SQAF": [], "OTHER": []}
        self.temp_guests = []
        self.current_team = "SQAH"
        # In the __init__ method, add these attributes
        self.top_buttons_frame = tk.Frame(self.root)
        self.top_buttons_frame.grid(row=0, column=2, sticky="ne", padx=10, pady=10)

        self.settings_button = tk.Button(
            self.top_buttons_frame,
            text="S",
            width=2,
            height=1,
            command=self.open_settings,
            font=("Arial", 10, "bold")
        )
        self.settings_button.pack(side=tk.RIGHT)


        # Recent names memory (max 2000 names)
        self.recent_names = []
        
        # Initialize data
        self.load_recent_names()
        self.load_recent_hills()
        self.load_athletes_from_json()

        # Build GUI
        self.build_gui()
        self.apply_settings_to_gui()
        # Set the default team and bindings
        self.set_team("SQAH")
        self.athlete_name_entry.bind('<KeyRelease>', self.autocomplete_athlete_name)
        self.autocomplete_listbox.bind("<<ListboxSelect>>", self.on_suggestion_select)
        self.bind_deletion_keys()

        # Bind the resize event to update the suggestion box width
        self.athlete_name_entry.bind('<Configure>', self.update_suggestion_box_width)

    def build_gui(self):
        """Builds the GUI components."""
        # Title
        self.title_label = tk.Label(self.root, text="Brower Timing Reformatted", font=("Arial", 20, "bold"))
        self.title_label.grid(row=0, column=0, columnspan=3, pady=(10, 0))

        # File Selection Section (Select Browser CSV)
        self.file_frame = tk.Frame(self.root)
        self.file_frame.grid(row=1, column=0, pady=(5, 0), sticky="n")

        # File Selection Section with a thick border
        self.button_label_border_frame = tk.Frame(self.file_frame, bd=4, relief="solid")
        self.button_label_border_frame.pack(padx=5, pady=(5, 5))

        self.select_file_button = tk.Button(self.button_label_border_frame, text="Select Brower CSV", command=self.select_file)
        self.select_file_button.pack(side=tk.LEFT, padx=5, pady=5)

        self.file_label = tk.Label(self.button_label_border_frame, text="No file selected", anchor="w", width=40)
        self.file_label.pack(side=tk.LEFT, padx=5, pady=5)

        # Training Parameters Section
        self.training_parameters_frame = tk.Frame(self.root, bd=4, relief="solid", padx=50, pady=10)
        self.training_parameters_frame.grid(row=1, column=0, pady=(72, 10), sticky="n")

        self.training_parameters_label = tk.Label(self.training_parameters_frame, text="Training Parameters", font=("Arial", 12, "bold"))
        self.training_parameters_label.grid(row=0, column=0, pady=(0, 10))

        # Add Reformat Button Section below Training Parameters
        self.reformat_frame = tk.Frame(self.root, bd=4, relief="solid", padx=10, pady=10)
        self.reformat_frame.grid(row=2, column=0, pady=10, padx=10, sticky="n")

        self.reformat_button = tk.Button(
            self.reformat_frame, 
            text="Reformat Selected File",
            command=self.reformat_file,
            font=("Arial", 12, "bold"),
            pady=10,
            padx=85
        )
        self.reformat_button.pack(padx=20, pady=10)

        # Add a new frame at the bottom for version and author info
        self.bottom_frame = tk.Frame(self.root)
        self.bottom_frame.grid(row=999, column=0, columnspan=3, sticky="ew", pady=(10,5))
        
        # Configure column weights to create proper spacing
        self.bottom_frame.grid_columnconfigure(1, weight=1)  # This creates space between the labels
        
        # Add author label (left aligned)
        self.author_label = tk.Label(
            self.bottom_frame, 
            text=self.author, 
            font=("Arial", 8)
        )
        self.author_label.grid(row=0, column=0, sticky="w", padx=5)
        
        # Add version label (right aligned)
        self.version_label = tk.Label(
            self.bottom_frame, 
            text=f"Version Number: {self.version}", 
            font=("Arial", 8)
        )
        self.version_label.grid(row=0, column=2, sticky="e", padx=5)

        # Location & Event Section inside Training Parameters
        self.team_event_frame = tk.LabelFrame(self.training_parameters_frame, text="Event & Location", padx=52, pady=10)
        self.team_event_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        # Event Selection (Dropdown Menu)
        tk.Label(self.team_event_frame, text="Event: ").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.event_menu = tk.OptionMenu(self.team_event_frame, self.event_var, "SL", "GS", "SG", "DH", "SX")
        self.event_menu.grid(row=0, column=1, padx=5, pady=5, sticky='w')

        # Hill Entry with Autocomplete
        tk.Label(self.team_event_frame, text="Hill: ").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.hill_entry = tk.Entry(self.team_event_frame, textvariable=self.hill_var)
        self.hill_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        # Create the autocomplete listbox for hills
        self.hill_autocomplete_listbox = tk.Listbox(
            self.team_event_frame,
            height=0,
            borderwidth=1,
            relief="solid"
        )
        self.hill_autocomplete_listbox.grid(row=2, column=1, padx=(5, 5), pady=(0, 5), sticky="w")
        self.hill_autocomplete_listbox.grid_remove()  # Hide by default

        # Bind the events
        self.hill_entry.bind('<KeyRelease>', self.autocomplete_hill_name)
        self.hill_autocomplete_listbox.bind("<<ListboxSelect>>", self.on_hill_suggestion_select)

        # Right Side: Section with a thick border
        self.right_frame_with_border = tk.Frame(self.root, bd=4, relief="solid", padx=10, pady=10)
        self.right_frame_with_border.grid(row=1, column=2, rowspan=3, padx=10, pady=10, sticky="n")

        # Move everything into the bordered frame
        self.right_frame = tk.Frame(self.right_frame_with_border)
        self.right_frame.grid(row=0, column=0, padx=10, pady=10, sticky="n")

        # New Team Selector
        self.new_team_label = tk.Label(self.right_frame, text="Select Team", font=("Arial", 12, "bold"))
        self.new_team_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))

        self.team_button_frame = tk.Frame(self.right_frame)
        self.team_button_frame.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="w")

        self.sqah_button = tk.Button(self.team_button_frame, text="SQAH", width=10, command=lambda: self.set_team("SQAH"))
        self.sqah_button.pack(side=tk.LEFT, padx=5)

        self.sqaf_button = tk.Button(self.team_button_frame, text="SQAF", width=10, command=lambda: self.set_team("SQAF"))
        self.sqaf_button.pack(side=tk.LEFT, padx=5)

        self.other_button = tk.Button(self.team_button_frame, text="OTHER", width=10, command=lambda: self.set_team("OTHER"))
        self.other_button.pack(side=tk.LEFT, padx=5)

        # Athlete Listbox
        self.athlete_list_label = tk.Label(self.right_frame, text="Current Athletes:")
        self.athlete_list_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")

        self.athlete_listbox = tk.Listbox(self.right_frame, width=40, height=10)
        self.athlete_listbox.grid(row=3, column=0, padx=10, pady=5)

        # Temporary Guests Listbox
        self.guest_list_label = tk.Label(self.right_frame, text="Temporary Guests:")
        self.guest_list_label.grid(row=4, column=0, padx=10, pady=5, sticky="w")

        self.guest_listbox = tk.Listbox(self.right_frame, width=40, height=5)
        self.guest_listbox.grid(row=5, column=0, padx=10, pady=5)

        # Manage Athletes Section
        self.manage_athletes_frame = tk.Frame(self.right_frame, bd=2, relief="groove", padx=10, pady=10)
        self.manage_athletes_frame.grid(row=6, column=0, padx=10, pady=10, sticky="n")

        self.manage_athletes_label = tk.Label(self.manage_athletes_frame, text="Manage Athletes", font=("Arial", 12, "bold"))
        self.manage_athletes_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))

        self.athlete_name_label = tk.Label(self.manage_athletes_frame, text="Athlete Name:")
        self.athlete_name_label.grid(row=1, column=0, padx=(10, 5), pady=5, sticky="e")

        self.athlete_name_entry = tk.Entry(self.manage_athletes_frame)
        self.athlete_name_entry.grid(row=1, column=1, padx=(0, 5), pady=5, sticky="w")

        # Update the autocomplete listbox creation
        self.autocomplete_listbox = tk.Listbox(
            self.manage_athletes_frame,
            height=0,
            borderwidth=1,
            relief="solid"
        )
        self.autocomplete_listbox.grid(row=2, column=1, padx=(0, 5), pady=(0, 5), sticky="ew")  # Changed sticky to "ew"
        self.autocomplete_listbox.grid_remove()  # Hide by default

        self.bib_number_label = tk.Label(self.manage_athletes_frame, text="Bib Number:")
        self.bib_number_label.grid(row=2, column=0, padx=(10, 5), pady=5, sticky="e")

        self.bib_number_entry = tk.Entry(self.manage_athletes_frame)
        self.bib_number_entry.grid(row=2, column=1, padx=(0, 5), pady=5, sticky="w")

        # Buttons for athlete management
        self.button_frame = tk.Frame(self.manage_athletes_frame)
        self.button_frame.grid(row=3, column=0, columnspan=2, pady=10)

        button_width = 20
        self.add_athlete_button = tk.Button(self.button_frame, text="Add Athlete to Team", command=self.add_athlete, width=button_width)
        self.add_athlete_button.grid(row=0, column=0, pady=5)

        self.add_guest_button = tk.Button(self.button_frame, text="Add Temporary Guest", command=self.add_guest, width=button_width)
        self.add_guest_button.grid(row=1, column=0, pady=5)

        self.remove_athlete_button = tk.Button(self.button_frame, text="Remove Selected Athlete", command=self.remove_selected_athlete, width=button_width)
        self.remove_athlete_button.grid(row=2, column=0, pady=5)

        # Weather Conditions Section inside Training Parameters
        self.weather_frame = tk.LabelFrame(self.training_parameters_frame, text="Weather Conditions", padx=10, pady=10)
        self.weather_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

        # Snow Conditions
        tk.Label(self.weather_frame, text="Snow: ").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        snow_conditions = ["Very Soft", "Soft", "Medium", "Hard", "Injected"]
        self.snow_condition_menu = Combobox(self.weather_frame, textvariable=self.snow_condition_var, values=snow_conditions, width=17)
        self.snow_condition_menu.grid(row=0, column=1, padx=5, pady=5, sticky='w')
        self.snow_condition_menu.set("Hard")  # Default value

        # Sky Conditions
        tk.Label(self.weather_frame, text="Sky: ").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        sky_conditions = ["Clear Sky", "Partly Cloudy", "Overcast", "Low Visibility", "Fog"]
        self.sky_condition_menu = Combobox(self.weather_frame, textvariable=self.sky_condition_var, values=sky_conditions, width=17)
        self.sky_condition_menu.grid(row=1, column=1, padx=5, pady=5, sticky='w')
        self.sky_condition_menu.set("Clear Sky")  # Default value

        # Precipitation
        tk.Label(self.weather_frame, text="Precipitation: ").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        precipitation_types = ["No Precipitation", "Light Snow", "Medium Snow", "Snowstorm", "Light Rain", "Moderate Rain", "Downpour"]
        self.precipitation_menu = Combobox(self.weather_frame, textvariable=self.precipitation_var, values=precipitation_types, width=17)
        self.precipitation_menu.grid(row=2, column=1, padx=5, pady=5, sticky='w')
        self.precipitation_menu.set("No Precipitation")  # Corrected from "No Precipitations"

        # Wind Conditions
        tk.Label(self.weather_frame, text="Wind: ").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        wind_conditions = ["No Wind", "Light Wind", "Moderate Wind", "Heavy Wind"]
        self.wind_condition_menu = Combobox(self.weather_frame, textvariable=self.wind_condition_var, values=wind_conditions, width=17)
        self.wind_condition_menu.grid(row=3, column=1, padx=5, pady=5, sticky='w')
        self.wind_condition_menu.set("No Wind")  # Default value

        # Add File Details Section inside Training Parameters
        self.file_details_frame = tk.LabelFrame(self.training_parameters_frame, text="File Details", padx=10, pady=10)
        self.file_details_frame.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

        # Date Field
        tk.Label(self.file_details_frame, text="Date:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.date_entry = tk.Entry(self.file_details_frame, textvariable=self.date_var, width=20, state='readonly')
        self.date_entry.grid(row=0, column=1, padx=5, pady=5, sticky='w')

        # Start Time Field
        tk.Label(self.file_details_frame, text="Start Time:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.time_entry = tk.Entry(self.file_details_frame, textvariable=self.time_var, width=20, state='readonly')
        self.time_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        # Session Number Field
        tk.Label(self.file_details_frame, text="Session #:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        self.session_entry = tk.Entry(self.file_details_frame, textvariable=self.session_var, width=20, state='readonly')
        self.session_entry.grid(row=2, column=1, padx=5, pady=5, sticky='w')


    def save_current_state(self):
        """Saves the current state of athletes and settings."""
        try:
            self.save_athletes_to_json()
            self.save_recent_names()
            self.save_recent_hills()
            self.save_settings()
            messagebox.showinfo("Success", "Current state saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save current state: {str(e)}")


    # Update the save_settings method
    def save_settings(self):
        """Saves settings to a JSON file."""
        settings = {
            "excel_title": self.excel_title,
            "team_names": self.team_names,
            "default_hill": self.default_hill
        }
        try:
            with open("settings.json", "w", encoding='utf-8') as f:
                json.dump(settings, f, indent=4, ensure_ascii=False)
                
            # Update UI elements
            self.update_team_buttons()
        except Exception as e:
            print(f"Error saving settings: {str(e)}")  # Add debug print
            messagebox.showerror("Error", f"Failed to save settings: {str(e)}")

    def update_team_buttons(self):
        """Updates the team button text based on current settings."""
        self.sqah_button.config(text=self.team_names["SQAH"])
        self.sqaf_button.config(text=self.team_names["SQAF"])
        
    def apply_settings_to_gui(self):
        """Applies loaded settings to GUI elements after they're created."""
        # Update team buttons
        self.sqah_button.config(text=self.team_names["SQAH"])
        self.sqaf_button.config(text=self.team_names["SQAF"])
        
        # Apply default hill if set
        if self.default_hill:
            self.hill_var.set(self.default_hill)

    def load_settings(self):
        """Loads settings from JSON file."""
        try:
            with open("settings.json", "r", encoding='utf-8') as f:
                settings = json.load(f)
                
                # Load each setting with proper default fallback
                self.excel_title = settings.get("excel_title", "Training SQA Équipe du Québec")
                self.team_names = settings.get("team_names", {"SQAH": "SQAH", "SQAF": "SQAF"})
                self.default_hill = settings.get("default_hill", "")
                
        except FileNotFoundError:
            print("No settings file found, using defaults")
        except Exception as e:
            print(f"Error loading settings: {str(e)}")
    def open_settings(self):
        """Opens the settings window."""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("Settings")
        settings_window.geometry("400x600")
        
        # Make window modal
        settings_window.transient(self.root)
        settings_window.grab_set()
        
        # Title label
        title_label = tk.Label(
            settings_window,
            text="Settings",
            font=("Arial", 16, "bold")
        )
        title_label.pack(pady=20)
        
        # Excel Title Section
        excel_title_frame = tk.LabelFrame(settings_window, text="Excel Sheet Title", padx=10, pady=10)
        excel_title_frame.pack(fill="x", padx=20, pady=10)
        
        tk.Label(excel_title_frame, text="Current title:").pack(anchor="w")
        title_var = tk.StringVar(value=self.excel_title)
        title_entry = tk.Entry(excel_title_frame, textvariable=title_var, width=40)
        title_entry.pack(fill="x", pady=5)
        
        # Team Names Section
        team_frame = tk.LabelFrame(settings_window, text="Team Names", padx=10, pady=10)
        team_frame.pack(fill="x", padx=20, pady=10)
        
        team_vars = {}
        labels = {"SQAH": "Current Team 1", "SQAF": "Current Team 2"}
        for team in ["SQAH", "SQAF"]:
            team_label = tk.Label(team_frame, text=f"{labels[team]}:")
            team_label.pack(anchor="w")
            team_vars[team] = tk.StringVar(value=self.team_names[team])
            team_entry = tk.Entry(team_frame, textvariable=team_vars[team], width=40)
            team_entry.pack(fill="x", pady=5)

        
        # Default Hill Section
        hill_frame = tk.LabelFrame(settings_window, text="Default Training Hill", padx=10, pady=10)
        hill_frame.pack(fill="x", padx=20, pady=10)
        
        # Main hill section frame
        hill_section_frame = tk.Frame(hill_frame)
        hill_section_frame.pack(fill="x", pady=5)
        
        tk.Label(hill_section_frame, text="Current default hill:").pack(anchor="w")
        hill_var = tk.StringVar(value=self.default_hill)
        hill_entry = tk.Entry(hill_section_frame, textvariable=hill_var, width=40)
        hill_entry.pack(fill="x", pady=5)
        
        # Create listbox for hill suggestions
        hill_listbox = tk.Listbox(hill_section_frame, height=3)
        hill_listbox.pack(fill="x", pady=5)
        hill_listbox.pack_forget()  # Hide initially
        
        def show_hill_suggestions(*args):
            input_text = hill_var.get().lower()
            if input_text:
                suggestions = [hill for hill in self.recent_hills 
                            if hill.lower().startswith(input_text)][:3]
                if suggestions:
                    hill_listbox.delete(0, tk.END)
                    for hill in suggestions:
                        hill_listbox.insert(tk.END, hill)
                    hill_listbox.pack(fill="x", pady=5)
                else:
                    hill_listbox.pack_forget()
            else:
                hill_listbox.pack_forget()
        
        def use_suggestion(*args):
            if hill_listbox.curselection():
                selected = hill_listbox.get(hill_listbox.curselection())
                hill_var.set(selected)
                hill_listbox.pack_forget()
        
        hill_entry.bind('<KeyRelease>', show_hill_suggestions)
        hill_listbox.bind('<<ListboxSelect>>', use_suggestion)
        
        # Null button at the bottom of hill_frame
        null_button = tk.Button(hill_frame, text="Null", command=lambda: hill_var.set(""))
        null_button.pack(pady=5)
        
        # Create buttons frame
        buttons_frame = tk.Frame(settings_window)
        buttons_frame.pack(pady=20)
        
        def save_settings_only():
            """Saves settings without closing the window."""
            self.excel_title = title_var.get()
            for team in team_vars:
                self.team_names[team] = team_vars[team].get()
            self.default_hill = hill_var.get()
            
            # Save the settings to file
            self.save_settings()
            
            # Update GUI elements
            if self.default_hill:
                self.hill_var.set(self.default_hill)
            
            # Update team buttons
            self.update_team_buttons()
            
            messagebox.showinfo("Success", "Settings saved successfully!")
        # Add Save and Close buttons side by side
        save_button = tk.Button(
            buttons_frame,
            text="Save",
            command=save_settings_only,
            font=("Arial", 10, "bold"),
            width=10
        )
        save_button.pack(side=tk.LEFT, padx=5)
        
        close_button = tk.Button(
            buttons_frame,
            text="Close",
            command=settings_window.destroy,
            font=("Arial", 10, "bold"),
            width=10
        )
        close_button.pack(side=tk.LEFT, padx=5)








    def select_file(self):
        """Opens a file dialog to select a CSV file and updates file details."""
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if file_path:
            self.selected_file = file_path  # Store the full path
            display_path = file_path if len(file_path) <= 35 else f"...{file_path[-33:]}"
            self.file_label.config(text=f"Selected File: {display_path}")
            
            # Parse the file and extract details
            file_details = self.parse_csv_file(file_path)
            if file_details:
                # Update UI elements with the parsed details
                self.date_var.set(file_details['date'] if file_details['date'] else "")
                self.time_var.set(file_details['time'] if file_details['time'] else "")
                self.session_var.set(file_details['session'] if file_details['session'] else "")
                
                # Force update of the UI
                self.root.update_idletasks()
            else:
                # Clear the fields if parsing failed
                self.date_var.set("")
                self.time_var.set("")
                self.session_var.set("")
                messagebox.showwarning("Warning", "Could not read file details. Please check the file format.")

    # Athlete Data Management Methods
    def add_athlete_to_memory(self, athlete_name):
        """Adds an athlete name to recent names memory, ensuring a maximum of 2000 entries."""
        if athlete_name in self.recent_names:
            self.recent_names.remove(athlete_name)
        self.recent_names.append(athlete_name)
        if len(self.recent_names) > 2000:
            self.recent_names = self.recent_names[-2000:]  # Keep the last 2000 names
        self.save_recent_names()

    def save_recent_names(self):
        """Saves the list of recent names to a text file."""
        with open("recent_names.txt", "w", encoding='utf-8') as f:
            for name in self.recent_names:
                f.write(name + '\n')

    def load_recent_names(self):
        """Loads the list of recent names from a text file."""
        try:
            with open("recent_names.txt", "r", encoding='utf-8') as f:
                self.recent_names = [line.strip() for line in f.readlines()]
        except FileNotFoundError:
            self.recent_names = []

    def save_athletes_to_json(self):
        """Saves the athlete data to a JSON file."""
        with open("athletes_data.json", "w") as f:
            json.dump(self.athletes, f, indent=4)

    def load_athletes_from_json(self):
        """Loads the athlete data from a JSON file."""
        try:
            with open("athletes_data.json", "r") as f:
                self.athletes = json.load(f)
        except FileNotFoundError:
            self.athletes = {"SQAH": [], "SQAF": [], "OTHER": []}
        self.athletes["OTHER"] = []
        self.temp_guests = []

    # GUI Update Methods
    def update_athlete_listbox(self):
        """Update the athlete listbox with the current team's athletes, sorted by bib number."""
        self.athlete_listbox.delete(0, tk.END)  # Clear the listbox first

        # Check if the current team exists and has athletes
        if self.current_team in self.athletes:
            # Sort the athletes by bib number
            sorted_athletes = sorted(self.athletes[self.current_team], key=lambda x: int(x['bib']))

            # Insert sorted athletes into the listbox
            for athlete in sorted_athletes:
                self.athlete_listbox.insert(tk.END, f"{athlete['name']} (Bib {athlete['bib']})")

    def update_guest_listbox(self):
        """Update the guest listbox with the current temporary guests, marking inactive guests."""
        self.guest_listbox.delete(0, tk.END)
        for guest in self.temp_guests:
            guest_status = "(Inactive)" if guest.get("inactive", False) else ""
            self.guest_listbox.insert(tk.END, f"{guest['name']} (Bib {guest['bib']}) {guest_status}")

    def bind_deletion_keys(self):
        """Bind deletion keys to listboxes for removing athletes or guests."""
        self.athlete_listbox.bind('<Delete>', self.remove_selected_athlete)
        self.athlete_listbox.bind('<BackSpace>', self.remove_selected_athlete)
        self.guest_listbox.bind('<Delete>', self.remove_selected_athlete)
        self.guest_listbox.bind('<BackSpace>', self.remove_selected_athlete)

    def update_suggestion_box_width(self, event=None):
        """Updates the width of the suggestion box to match the Athlete Name entry."""
        # Configure the grid column to match the entry width
        self.manage_athletes_frame.grid_columnconfigure(1, minsize=self.athlete_name_entry.winfo_width())
    
   
    # Event Handlers and Animations
    def animate_listbox(self):
        """Handles the animation for the autocomplete listbox."""
        if self.is_animating:
            current_height = self.animation_height
            target_height = self.target_height

            if current_height < target_height:
                # Opening animation
                self.animation_height = min(current_height + self.animation_speed, target_height)
                self.autocomplete_listbox.configure(height=int(self.animation_height / self.line_height))
            elif current_height > target_height:
                # Closing animation
                self.animation_height = max(current_height - self.animation_speed, target_height)
                self.autocomplete_listbox.configure(height=int(self.animation_height / self.line_height))

            # Continue animation if not reached target
            if self.animation_height != target_height:
                self.root.after(10, self.animate_listbox)
            else:
                self.is_animating = False
                if target_height == 0:
                    self.autocomplete_listbox.grid_remove()

    def autocomplete_athlete_name(self, event):
        """Provides autocomplete suggestions for athlete names based on recent entries."""
        input_text = self.athlete_name_entry.get().lower()

        # Get entry widget width if we haven't stored it
        if not self.name_entry_width:
            self.update_suggestion_box_width()
            # Get line height for calculations
            self.line_height = self.athlete_name_entry.winfo_reqheight()
            self.autocomplete_listbox.configure(width=self.name_entry_width)

        if not input_text:
            # Animate closing if input is empty
            self.target_height = 0
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()
            return

        # Filter recent names that start with the input text
        suggestions = [name for name in reversed(self.recent_names) if name.lower().startswith(input_text)]
        suggestions = suggestions[:2]  # Limit to 2 suggestions

        # Update the Listbox with suggestions
        if suggestions:
            self.autocomplete_listbox.delete(0, tk.END)
            for suggestion in suggestions:
                self.autocomplete_listbox.insert(tk.END, suggestion)

            # Update the suggestion box width
            self.update_suggestion_box_width()

            # Calculate target height based on number of suggestions
            num_suggestions = len(suggestions)
            self.target_height = num_suggestions * self.line_height

            # Show and position the listbox if it's hidden
            if not self.autocomplete_listbox.winfo_viewable():
                self.autocomplete_listbox.grid(row=2, column=1, padx=(0, 5), pady=(0, 5), sticky="w")
                self.animation_height = 0

            # Start animation if not already animating
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()

            self.autocomplete_listbox.lift()
        else:
            # Animate closing if no suggestions
            self.target_height = 0
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()

    def on_suggestion_select(self, event):
        """Handles the selection of a suggestion from the autocomplete listbox."""
        if self.autocomplete_listbox.curselection():
            selected_suggestion = self.autocomplete_listbox.get(self.autocomplete_listbox.curselection())
            self.athlete_name_entry.delete(0, tk.END)
            self.athlete_name_entry.insert(0, selected_suggestion)

            # Animate closing
            self.target_height = 0
            if not self.is_animating:
                self.is_animating = True
                self.animate_listbox()

    # Athlete Management Methods
    def add_athlete(self):
        """Adds a new athlete to the current team."""
        athlete_name = self.athlete_name_entry.get()
        bib_number = self.bib_number_entry.get()

        if athlete_name and bib_number and self.current_team:
            if not bib_number.isdigit():
                messagebox.showwarning("Input Error", "Bib invalid")
                return

            bib_number = int(bib_number)  # Ensure bib_number is an integer

            # Check for duplicate bib in the current team's athlete list
            for athlete in self.athletes[self.current_team]:
                if int(athlete['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to an athlete in this team.")
                    return

            # Check for duplicate bib in the temporary guest list
            for guest in self.temp_guests:
                if int(guest['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to a temporary guest.")
                    return

            # Add the new athlete if no duplicates found
            new_athlete = {"name": athlete_name, "bib": bib_number}
            self.athletes[self.current_team].append(new_athlete)
            self.save_athletes_to_json()
            self.update_athlete_listbox()

            # Add athlete name to recent names
            self.add_athlete_to_memory(athlete_name)

            # Clear the fields after adding
            self.athlete_name_entry.delete(0, tk.END)
            self.bib_number_entry.delete(0, tk.END)
        else:
            messagebox.showwarning("Input Error", "Please fill all fields before adding an athlete.")

    def add_guest(self):
        """Adds a new temporary guest."""
        guest_name = self.athlete_name_entry.get()
        bib_number = self.bib_number_entry.get()

        if guest_name and bib_number:
            if not bib_number.isdigit():
                messagebox.showwarning("Input Error", "Bib invalid")
                return

            bib_number = int(bib_number)  # Ensure bib_number is an integer

            # Check for duplicate bib in the current team's athlete list
            for athlete in self.athletes[self.current_team]:
                if int(athlete['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to an athlete in this team.")
                    return

            # Check for duplicate bib in the temporary guest list
            for guest in self.temp_guests:
                if int(guest['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", "This bib number is already assigned to a temporary guest.")
                    return

            # Add the new guest if no duplicates found
            new_guest = {"name": guest_name, "bib": bib_number}
            self.temp_guests.append(new_guest)
            self.update_guest_listbox()

            # Add guest name to recent names
            self.add_athlete_to_memory(guest_name)

            # Clear the fields after adding
            self.athlete_name_entry.delete(0, tk.END)
            self.bib_number_entry.delete(0, tk.END)
        else:
            messagebox.showwarning("Input Error", "Please fill all fields before adding a guest.")

    def remove_selected_athlete(self, event=None):
        """Removes the selected athlete or guest from the list."""
        selected_team_athlete = self.athlete_listbox.curselection()
        selected_guest = self.guest_listbox.curselection()

        if selected_team_athlete and self.current_team:
            if messagebox.askyesno("Confirm Deletion", "Are you sure you want to remove this athlete?"):
                del self.athletes[self.current_team][selected_team_athlete[0]]
                self.save_athletes_to_json()  # Save changes to JSON
                self.update_athlete_listbox()
        elif selected_guest:
            if messagebox.askyesno("Confirm Deletion", "Are you sure you want to remove this guest?"):
                del self.temp_guests[selected_guest[0]]
                self.update_guest_listbox()
        else:
            messagebox.showwarning("Selection Error", "No athlete or guest selected to remove.")

    def check_guest_conflicts_with_athletes(self, new_team):
        """Checks for duplicate bibs between temporary guests and the athletes in the selected team.
        Marks guests as inactive if duplicates are found."""
        conflicts = False
        for guest in self.temp_guests:
            guest_bib = int(guest["bib"])

            # Check for duplicate bibs in the new team's athlete list
            duplicate_found = any(int(athlete['bib']) == guest_bib for athlete in self.athletes[new_team])

            # If a duplicate is found, mark the guest as inactive
            if duplicate_found:
                guest["inactive"] = True
                conflicts = True
            else:
                # If no duplicates are found in the new team, remove the inactive status
                guest["inactive"] = False
        # If conflicts are detected, show a warning
        if conflicts:
            messagebox.showwarning("Bib Conflict", "Some Bibs are duplicates. Guests will not be active.")

    # Other Methods
    def set_team(self, team_name):
        """Changes the selected team and updates the GUI accordingly."""
        self.current_team = team_name
        self.sqah_button.config(bg="SystemButtonFace")
        self.sqaf_button.config(bg="SystemButtonFace")
        self.other_button.config(bg="SystemButtonFace")

        if team_name == "SQAH":
            self.sqah_button.config(bg="lightblue")
        elif team_name == "SQAF":
            self.sqaf_button.config(bg="lightblue")
        elif team_name == "OTHER":
            self.other_button.config(bg="lightblue")

        # Recheck guest conflicts and remove "Inactive" if no duplicates found
        self.check_guest_conflicts_with_athletes(team_name)

        # Update the athlete and guest listboxes
        self.update_athlete_listbox()
        self.update_guest_listbox()

    # Add these methods to the TimingSystemApp class:

    def add_hill_to_memory(self, hill_name):
        """Adds a hill name to recent hills memory, ensuring a maximum of 2000 entries."""
        if hill_name in self.recent_hills:
            self.recent_hills.remove(hill_name)
        self.recent_hills.append(hill_name)
        if len(self.recent_hills) > 2000:
            self.recent_hills = self.recent_hills[-2000:]  # Keep the last 2000 names
        self.save_recent_hills()

    def save_recent_hills(self):
        """Saves the list of recent hills to a text file."""
        with open("recent_hills.txt", "w", encoding='utf-8') as f:
            for name in self.recent_hills:
                f.write(name + '\n')

    def load_recent_hills(self):
        """Loads the list of recent hills from a text file."""
        try:
            with open("recent_hills.txt", "r", encoding='utf-8') as f:
                self.recent_hills = [line.strip() for line in f.readlines()]
        except FileNotFoundError:
            self.recent_hills = []

    def get_hill_name_matches(self, input_text):
        """
        Returns matching hill names based on various search criteria.
        Handles partial matches after common prefixes and within compound names.
        
        Args:
            input_text (str): The search text
            
        Returns:
            list: Matching hill names
        """
        input_text = input_text.lower().strip()
        matches = set()  # Using set to avoid duplicates
        
        for hill in self.recent_hills:
            hill_lower = hill.lower()
            
            # Direct start match
            if hill_lower.startswith(input_text):
                matches.add(hill)
                
            # Split hill name into parts and check each part
            parts = hill_lower.replace('-', ' ').split()
            
            # Handle common prefixes like "Mont-" or "Mont "
            if len(parts) > 1:
                # If the hill starts with "Mont" or similar, check the part after it
                if parts[0] in ['mont']:
                    remaining_name = ' '.join(parts[1:])
                    if remaining_name.startswith(input_text):
                        matches.add(hill)
                    
                    # Also check individual parts after "Mont"
                    for part in parts[1:]:
                        if part.startswith(input_text):
                            matches.add(hill)
            
            # Check each part individually
            for part in parts:
                if part.startswith(input_text):
                    matches.add(hill)
        
        # Convert set back to list and sort by relevance
        matches_list = list(matches)
        matches_list.sort(key=lambda x: (
            # Prioritize exact matches
            not x.lower().startswith(input_text),
            # Then prioritize matches after "Mont-" or "Mont "
            not any(part.lower().startswith(input_text) 
                for part in x.lower().replace('-', ' ').split()[1:] 
                if x.lower().startswith('mont')),
            # Finally sort alphabetically
            x.lower()
        ))
        
        return matches_list[:3]  # Limit to top 3 matches



    def autocomplete_hill_name(self, event):
        """
        Enhanced autocomplete for hill names with improved matching logic.
        """
        input_text = self.hill_entry.get().lower()
        
        if not input_text:
            # Animate closing if input is empty
            self.hill_target_height = 0
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()
            return
        
        # Get matches using the new matching function
        suggestions = self.get_hill_name_matches(input_text)
        
        # Update the Listbox with suggestions
        if suggestions:
            self.hill_autocomplete_listbox.delete(0, tk.END)
            for suggestion in suggestions:
                self.hill_autocomplete_listbox.insert(tk.END, suggestion)
            
            # Calculate target height based on number of suggestions
            self.hill_target_height = self.line_height * len(suggestions)
            
            # Show and position the listbox if it's hidden
            if not self.hill_autocomplete_listbox.winfo_viewable():
                self.hill_autocomplete_listbox.grid()
                self.hill_animation_height = 0
            
            # Start animation if not already animating
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()
            
            self.hill_autocomplete_listbox.lift()
        else:
            # Animate closing if no suggestions
            self.hill_target_height = 0
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()

    def animate_hill_listbox(self):
        """Handles the animation for the hill autocomplete listbox."""
        if self.hill_is_animating:
            current_height = self.hill_animation_height
            target_height = self.hill_target_height

            if current_height < target_height:
                # Opening animation
                self.hill_animation_height = min(current_height + self.animation_speed, target_height)
                self.hill_autocomplete_listbox.configure(height=int(self.hill_animation_height / self.line_height))
            elif current_height > target_height:
                # Closing animation
                self.hill_animation_height = max(current_height - self.animation_speed, target_height)
                self.hill_autocomplete_listbox.configure(height=int(self.hill_animation_height / self.line_height))

            # Continue animation if not reached target
            if self.hill_animation_height != target_height:
                self.root.after(10, self.animate_hill_listbox)
            else:
                self.hill_is_animating = False
                if target_height == 0:
                    self.hill_autocomplete_listbox.grid_remove()

    def on_hill_suggestion_select(self, event):
        """Handles the selection of a suggestion from the hill autocomplete listbox."""
        if self.hill_autocomplete_listbox.curselection():
            selected_suggestion = self.hill_autocomplete_listbox.get(self.hill_autocomplete_listbox.curselection())
            self.hill_entry.delete(0, tk.END)
            self.hill_entry.insert(0, selected_suggestion)
            self.add_hill_to_memory(selected_suggestion)

            # Animate closing
            self.hill_target_height = 0
            if not self.hill_is_animating:
                self.hill_is_animating = True
                self.animate_hill_listbox()

    def get_athlete_name(self, bib_number):
        """
        Given a bib number, returns the athlete's name from all teams' athletes or temp guests.
        """
        bib_number = int(bib_number)
        # Check in all teams
        for team in self.athletes:
            for athlete in self.athletes[team]:
                if int(athlete['bib']) == bib_number:
                    return athlete['name']
        # Then, check in temp guests
        for guest in self.temp_guests:
            if int(guest['bib']) == bib_number and not guest.get('inactive', False):
                return guest['name']
        # If not found, return 'Unknown Athlete'
        return 'Unknown Athlete'

    def parse_csv_file(self, file_path):
        """
        Parses the CSV file to extract session details.
        Updated to handle the custom separator and add debug logging.
        """
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                
                # Initialize variables
                session_num = ""
                date_str = ""
                time_str = ""
                
                print("Parsing file:", file_path)  # Debug logging
                
                # Parse the header section
                for line in lines[:20]:  # Only check first 20 lines for headers
                    line = line.strip()
                    print("Processing line:", line)  # Debug logging
                    
                    # Try both '>' and ':' as separators
                    if '>' in line:
                        parts = line.split('>')
                    elif ':' in line:
                        parts = line.split(':')
                    else:
                        continue
                    
                    # Clean up parts
                    parts = [p.strip() for p in parts]
                    
                    if len(parts) < 2:
                        continue
                        
                    # Check for each field
                    if "Session" in parts[0]:
                        session_num = parts[1].strip('#').strip()
                        print("Found session:", session_num)  # Debug logging
                    elif "Date" == parts[0]:
                        date_str = parts[1].strip()
                        print("Found date:", date_str)  # Debug logging
                    elif "Time" == parts[0]:
                        time_str = parts[1].strip()
                        print("Found time:", time_str)  # Debug logging
                    
                # Convert date format if found
                if date_str:
                    try:
                        # Parse the date (assuming mm/dd/yy format)
                        date_obj = datetime.strptime(date_str, '%m/%d/%y')
                        # Convert to dd/mm/yyyy format
                        formatted_date = date_obj.strftime('%d/%m/%Y')
                    except ValueError:
                        print("Date parsing failed, using original:", date_str)  # Debug logging
                        formatted_date = date_str
                else:
                    formatted_date = ""
                
                print("Final values:", {  # Debug logging
                    'session': session_num,
                    'date': formatted_date,
                    'time': time_str
                })
                
                return {
                    'session': session_num,
                    'date': formatted_date,
                    'time': time_str
                }
                
        except Exception as e:
            print(f"Error reading file: {str(e)}")  # Debug logging
            messagebox.showerror("Error", f"Error reading file: {str(e)}")
            return None

    def parse_timing_data(self, file_path):
        """
        Parses the CSV file to extract timing data, organized by runs.
        Handles custom separator and header format.
        """
        timing_data = {}
        header_found = False
        column_indices = {}
        
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                
                # First determine the separator
                separator = '>'  # Based on the sample file format
                
                for line in lines:
                    line = line.strip()
                    if not line:  # Skip empty lines
                        continue
                    
                    # Look for the header line that contains both Bib# and Run#
                    if not header_found and "Bib#" in line and "Run#" in line:
                        header_found = True
                        # Get column indices from header
                        headers = line.split(separator)
                        for i, header in enumerate(headers):
                            header = header.lower().strip()
                            if "bib" in header:
                                column_indices['bib'] = i
                            elif "run" in header:
                                column_indices['run'] = i
                            elif "split 1" in header:
                                column_indices['split1'] = i
                            elif "finish time" in header:
                                column_indices['finish'] = i
                            elif "status" in header:
                                column_indices['status'] = i
                        continue
                    
                    # Process data lines after header is found
                    if header_found:
                        data = line.split(separator)
                        if len(data) <= 1:  # Skip empty lines
                            continue
                        
                        try:
                            # Get run number and clean it
                            run_number = data[column_indices['run']].strip()
                            if not run_number:  # Skip if run number is empty
                                continue
                            
                            if run_number not in timing_data:
                                timing_data[run_number] = []
                            
                            # For this format, Split 1 is the intermediate time
                            # and Finish Time is the total time
                            split1_time = data[column_indices['split1']].strip()
                            status = data[column_indices.get('status', -1)].strip() if 'status' in column_indices else ''
                            
                            # Create entry for this run
                            entry = {
                                'bib': data[column_indices['bib']].strip(),
                                'split1': self.parse_time(split1_time),
                                'finish': self.parse_time(data[column_indices.get('finish', -1)].strip() if 'finish' in column_indices else ''),
                                'status': status if status else '',
                                'run': run_number
                            }
                            
                            # Only add valid entries
                            if entry['bib'] and entry['run']:
                                timing_data[run_number].append(entry)
                                
                        except IndexError as e:
                            print(f"Error processing line: {line}")
                            print(f"Error details: {str(e)}")
                            continue
            
            # Sort data for each run
            for run_number in timing_data:
                timing_data[run_number] = self.sort_run_data(timing_data[run_number])
            
            if not timing_data:
                messagebox.showerror("Error", "No valid timing data found in the file.")
                return None
                
            return timing_data
            
        except Exception as e:
            messagebox.showerror("Error", f"Error parsing CSV file: {str(e)}")
            return None

    def parse_time(self, time_str):
        """
        Parses a time string into a float value in seconds.
        Returns None for invalid times or DNF/DSQ.
        """
        if not time_str or not isinstance(time_str, str):
            return None
            
        time_str = time_str.strip().upper()
        if time_str in ['DNF', 'DSQ', '', 'DNS']:
            return None
            
        try:
            if ':' in time_str:
                minutes, seconds = time_str.split(':')
                return float(minutes) * 60 + float(seconds)
            return float(time_str)
        except (ValueError, TypeError):
            return None

    def create_formatted_excel(self, output_path):
        """
        Creates a formatted Excel file with the specified header layout and timing data.
        """
        # Create new workbook and get active sheet
        wb = Workbook()
        ws = wb.active
        
        # Define fonts
        try:
            title_font = Font(name='Avenir Next LT Pro', size=18, bold=True)
        except:
            title_font = Font(name='Arial', size=18, bold=True)
        
        normal_font = Font(name='Arial', size=11)
        header_font = Font(name='Arial', size=11, bold=True)
        
       # Set column widths
        ws.column_dimensions['A'].width = 6   # Index column
        ws.column_dimensions['B'].width = 16   # Bib #
        ws.column_dimensions['C'].width = 20  # Name
        ws.column_dimensions['D'].width = 10  # Split 1
        ws.column_dimensions['E'].width = 12  # Split Diff.
        ws.column_dimensions['F'].width = 8   # Rank
        ws.column_dimensions['G'].width = 12  # Split Finish
        ws.column_dimensions['H'].width = 12  # S-F Diff.
        ws.column_dimensions['I'].width = 8   # Rank
        ws.column_dimensions['J'].width = 12  # Finish Time
        ws.column_dimensions['K'].width = 12  # Finish Diff.
        ws.column_dimensions['L'].width = 10  # Status
            
        # Set row height
        ws.row_dimensions[1].height = 30
        # Get the season based on the date
        season = self.get_season(self.date_var.get())
        
        # Get the custom team name from settings and add season
        current_team_name = self.team_names[self.current_team]  # Get the custom team name
        team_with_season = f"{current_team_name} - {season}" if season else current_team_name
        
        # Title in B1
        ws['B1'] = self.excel_title
        ws['B1'].font = title_font
        ws['B1'].alignment = Alignment(vertical='center', horizontal='left')

        # Define labels and their corresponding values
        label_value_pairs = {
            ('B2', 'Team:'): team_with_season,  # Using the custom team name here
            ('B3', 'Session #:'): self.session_var.get(),
            ('B4', 'Event:'): self.event_var.get(),
            ('B5', 'Snow Condition:'): self.snow_condition_var.get(),
            ('E2', 'Date:'): self.date_var.get(),
            ('E3', 'Start Time:'): self.time_var.get(),
            ('E4', 'Hill:'): self.hill_var.get(),
        }
        

        # Write labels and values with proper formatting
        for (cell_coord, label), value in label_value_pairs.items():
            # Write label
            ws[cell_coord] = label
            ws[cell_coord].font = header_font
            ws[cell_coord].alignment = Alignment(horizontal='left')
            
            # Write value in the next column
            value_col = chr(ord(cell_coord[0]) + 1)  # Next column
            value_cell = f"{value_col}{cell_coord[1]}"
            ws[value_cell] = value
            ws[value_cell].font = normal_font
            ws[value_cell].alignment = Alignment(horizontal='left')

        # Special handling for Weather row (combines multiple conditions)
        ws['E5'] = 'Weather:'
        ws['E5'].font = header_font
        ws['E5'].alignment = Alignment(horizontal='left')
        
        weather_value = f"{self.sky_condition_var.get()}, {self.precipitation_var.get()}, {self.wind_condition_var.get()}"
        ws['F5'] = weather_value
        ws['F5'].font = normal_font
        ws['F5'].alignment = Alignment(horizontal='left')


        # Now parse and add timing data
        if self.selected_file:
            timing_data = self.parse_timing_data(self.selected_file)
            if timing_data:
                current_row = 8  # Start after header section
                
                # Write each run's data
                for run_number in sorted(timing_data.keys(), key=int):
                    current_row = self.write_run_data(ws, timing_data[run_number], current_row)
                    current_row += 1  # Extra space between runs
        # After writing all run data, add the analysis graphs
        if self.selected_file:
            timing_data = self.parse_timing_data(self.selected_file)
            if timing_data:
                current_row = self.add_analysis_graphs(ws, timing_data, current_row)
        
        try:
            wb.save(output_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {str(e)}")
            return False      

    def get_season(self, date_str):
        """
        Determines the season based on the date.
        Args:
            date_str: Date string in dd/mm/yyyy format
        Returns:
            Season string in yyyy/yyyy format
        """
        try:
            # Parse the date string (dd/mm/yyyy)
            date_parts = date_str.split('/')
            if len(date_parts) != 3:
                return ""
                
            day = int(date_parts[0])
            month = int(date_parts[1])
            year = int(date_parts[2])
            
            # If date is before June 1st, season is previous_year/current_year
            # If date is June 1st or later, season is current_year/next_year
            if month < 6 or (month == 6 and day == 1):
                return f"{year-1}/{year}"
            else:
                return f"{year}/{year+1}"
        except (ValueError, AttributeError, IndexError):
            return ""

    def format_time(self, time_value, as_difference=False):
        """
        Formats a time value (in seconds) as a string.
        For differences, includes + sign and ensures 2 decimal places.
        """
        if time_value is None:
            return ""
            
        # For differences
        if as_difference:
            sign = '+' if time_value >= 0 else ''
            if abs(time_value) < 60:
                return f"{sign}{time_value:.2f}"
            minutes = int(abs(time_value) // 60)
            seconds = abs(time_value) % 60
            return f"{sign}{minutes}:{seconds:05.2f}"
        
        # For regular times
        if time_value < 60:
            return f"{time_value:.2f}"
        minutes = int(time_value // 60)
        seconds = time_value % 60
        return f"{minutes}:{seconds:05.2f}"

    def sort_run_data(self, run_data):
        """
        Sorts run data first by status (valid times first), then by finish time.
        """
        def sort_key(entry):
            # Primary sort: DNF/DSQ at the end
            is_valid = entry['status'].upper() not in ['DNF', 'DSQ']
            # Secondary sort: Finish time (None values at the end)
            finish_time = entry['finish'] if entry['finish'] is not None else float('inf')
            return (-is_valid, finish_time)
        
        return sorted(run_data, key=sort_key)

 

    def generate_filename(self):
        """
        Generates a filename in the format Hill_Team_Event_Date or Hill-hill_Team_Event_Date.
        Date is formatted as dd-mm-yyyy.
        
        Returns:
            str: The formatted filename with .xlsx extension
        """
        try:
            # Get and sanitize hill name
            hill = self.hill_var.get().strip()
            if not hill:
                hill = "Unknown-Hill"
            # Replace spaces with hyphens and remove any invalid filename characters
            hill = '-'.join(hill.split())
            hill = ''.join(c for c in hill if c.isalnum() or c in '-_')
            
            # Get team name
            team = self.current_team if self.current_team else "Unknown-Team"
            
            # Get event
            event = self.event_var.get().strip() if self.event_var.get() else "Unknown-Event"
            
            # Format the date
            date_str = self.date_var.get().strip()
            if date_str and '/' in date_str:
                try:
                    # Parse the date from dd/mm/yyyy
                    day, month, year = date_str.split('/')
                    date_str = f"{day}-{month}-{year}"
                except:
                    date_str = "No-Date"
            else:
                date_str = "No-Date"
            
            # Combine all parts with underscores
            filename = f"{hill}_{team}_{event}_{date_str}"
            
            # Remove any remaining invalid characters
            filename = ''.join(c for c in filename if c.isalnum() or c in '-_')
            
            # Add extension
            return f"{filename}.xlsx"
        except Exception as e:
            print(f"Error generating filename: {str(e)}")
            return "reformatted_timing_data.xlsx"

   

    def detect_outliers(self, times, is_split=True):
        """
        Detects outliers based on fixed thresholds from the mean, with improved invalid time handling.
        
        Args:
            times (list): List of time values
            is_split (bool): True if analyzing split times, False for finish times
            
        Returns:
            tuple: (mean_time, lower_bound, upper_bound, valid_times)
        """
        if not times:
            return None, None, None, []
        
        # Initial filtering of valid times (>= 10.0 seconds)
        filtered_times = [t for t in times if t >= 10.0]
        
        if not filtered_times:
            return None, None, None, []
        
        # Calculate mean from filtered times
        mean_time = sum(filtered_times) / len(filtered_times)
        
        # Set threshold based on type
        threshold = 3 if is_split else 5
        
        # Calculate bounds
        lower_bound = mean_time - threshold
        upper_bound = mean_time + threshold
        
        # Filter valid times within bounds
        valid_times = [t for t in filtered_times if lower_bound <= t <= upper_bound]
        
        return mean_time, lower_bound, upper_bound, valid_times

    
    def write_run_data(self, ws, run_data, start_row):
        """
        Writes run data to worksheet with gradient highlighting from fastest to slowest times.
        """
        # Define base styles
        header_border = Border(
            left=Side(style='thick'), 
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        normal_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def get_gradient_color(index, total, start_color=(0, 128, 0), end_color=(139, 0, 0)):
            """
            Generates a color along a gradient from dark green to dark red.
            """
            if total <= 1:
                return PatternFill(start_color='008000', end_color='008000', fill_type='solid')
                
            factor = index / (total - 1)
            r = int(start_color[0] + (end_color[0] - start_color[0]) * factor)
            g = int(start_color[1] + (end_color[1] - start_color[1]) * factor)
            b = int(start_color[2] + (end_color[2] - start_color[2]) * factor)
            
            hex_color = f"{r:02x}{g:02x}{b:02x}".upper()
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        # Pre-process and organize entries
        valid_entries = []
        dnf_entries = []
        err_entries = []
        dns_entries = []
        
        for entry in run_data:
            split_time = entry['split1']
            finish_time = entry['finish']
            status = entry['status'].upper()
            
            if status not in ['DNF', 'DSQ', 'DNS']:
                split_invalid = split_time is not None and split_time < 10.0
                finish_invalid = finish_time is not None and finish_time < 10.0
                if split_invalid or finish_invalid:
                    status = 'ERR'
                    entry['status'] = 'ERR'
            
            if status == 'DNS':
                dns_entries.append(entry)
            elif status == 'ERR':
                err_entries.append(entry)
            elif status == 'DNF':
                dnf_entries.append(entry)
            else:
                valid_entries.append(entry)
        
        ordered_run_data = valid_entries + dnf_entries + err_entries + dns_entries
        
        # Write run number
        current_row = start_row + 2
        ws[f'B{current_row}'] = f"Run {run_data[0]['run']}"
        current_row += 1
        
        # Write headers
        headers = ['Bib #', 'Name', 'Split 1', 'Split Diff.', 'Rank', 
                'Split Finish', 'S-F Diff.', 'Rank', 
                'Finish Time', 'Finish Diff.', 'Status']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1

        # Collect valid times for gradient coloring
        valid_split_times = []
        valid_finish_times = []
        valid_split_finish_times = []

        for entry in run_data:
            bib = entry['bib']
            # Include split times even if DNF, as long as the time is valid
            if entry['split1'] is not None and entry['split1'] >= 10.0 and entry['status'].upper() not in ['DNS', 'ERR']:
                valid_split_times.append((entry['split1'], bib))
            
            # For finish and split-finish times, only include complete runs
            if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']:
                if entry['finish'] is not None and entry['finish'] >= 10.0:
                    valid_finish_times.append((entry['finish'], bib))
                if (entry['finish'] is not None and entry['split1'] is not None and 
                    entry['finish'] >= 10.0 and entry['split1'] >= 10.0):
                    split_finish_time = entry['finish'] - entry['split1']
                    if split_finish_time > 0:
                        valid_split_finish_times.append((split_finish_time, bib))

        # Sort times and create color lookups
        valid_split_times.sort(key=lambda x: x[0])
        valid_finish_times.sort(key=lambda x: x[0])
        valid_split_finish_times.sort(key=lambda x: x[0])

        split_colors = {bib: get_gradient_color(i, len(valid_split_times)) 
                    for i, (_, bib) in enumerate(valid_split_times)}
        finish_colors = {bib: get_gradient_color(i, len(valid_finish_times)) 
                        for i, (_, bib) in enumerate(valid_finish_times)}
        split_finish_colors = {bib: get_gradient_color(i, len(valid_split_finish_times)) 
                            for i, (_, bib) in enumerate(valid_split_finish_times)}

        # Calculate best times for differences
        best_split = min(valid_split_times)[0] if valid_split_times else None
        best_finish = min(valid_finish_times)[0] if valid_finish_times else None
        best_split_finish = min(valid_split_finish_times)[0] if valid_split_finish_times else None

        # Write data rows
        for entry in ordered_run_data:
            split_time = entry['split1']
            finish_time = entry['finish']
            status = entry['status'].upper()
            bib = entry['bib']
            
            # Calculate split-to-finish time
            split_finish_time = None
            split_finish_diff = None
            if (status not in ['DNF', 'DSQ', 'DNS', 'ERR'] and 
                finish_time is not None and split_time is not None):
                split_finish_time = finish_time - split_time
                if split_finish_time > 0 and best_split_finish is not None:
                    split_finish_diff = split_finish_time - best_split_finish

            row = [
                bib,
                self.get_athlete_name(bib),
                self.format_time(split_time if split_time is not None and split_time >= 10.0 else None),
                self.format_time(split_time - best_split if split_time is not None and best_split is not None else None, True),
                '',  # Rank will be filled in later
                self.format_time(split_finish_time) if split_finish_time is not None and split_finish_time > 0 else '',
                self.format_time(split_finish_diff, True) if split_finish_diff is not None else '',
                '',  # Rank will be filled in later
                self.format_time(finish_time if finish_time is not None and finish_time >= 10.0 else None),
                self.format_time(finish_time - best_finish if finish_time is not None and best_finish is not None else None, True),
                status
            ]
            
            # Write row data and apply formatting
            for col, value in enumerate(row, start=2):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = normal_border
                cell.alignment = Alignment(horizontal='center')
                
                # Apply gradient highlighting
                if status not in ['DNS', 'ERR']:  # Allow Split 1 highlighting for DNF
                    if col in [3, 4]:  # Split 1 times and differences
                        if bib in split_colors:
                            cell.fill = split_colors[bib]
                
                if status not in ['DNS', 'ERR', 'DNF', 'DSQ']:  # Finish times only for complete runs
                    if col == 7:  # S-F Diff.
                        if bib in split_finish_colors:
                            cell.fill = split_finish_colors[bib]
                    elif col == 10:  # Finish times
                        if bib in finish_colors:
                            cell.fill = finish_colors[bib]
            
            current_row += 1
        
        return current_row



    
    def clean_and_analyze_timing_data(self, run_data):
        """
        Improved timing data analysis with separate handling for splits and finishes,
        and proper handling of ERR status.
        
        Args:
            run_data (list): List of timing entries for a run
        
        Returns:
            tuple: (cleaned_data, outliers)
        """
        valid_splits = []
        valid_finishes = []
        
        # First pass: collect valid times
        for entry in run_data:
            if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']:
                if entry['split1'] is not None and entry['split1'] >= 10.0:
                    valid_splits.append(entry['split1'])
                if entry['finish'] is not None and entry['finish'] >= 10.0:
                    valid_finishes.append(entry['finish'])
        
        # Get outlier bounds for both split and finish times
        split_stats = self.detect_outliers(valid_splits)
        finish_stats = self.detect_outliers(valid_finishes)
        
        cleaned_data = []
        outliers = []
        
        # Second pass: classify entries
        for entry in run_data:
            is_outlier = False
            status = entry['status'].upper()
            
            if status not in ['DNF', 'DSQ', 'DNS', 'ERR']:
                # Check for invalid times
                if (entry['split1'] is not None and entry['split1'] < 10.0) or \
                (entry['finish'] is not None and entry['finish'] < 10.0):
                    entry['status'] = 'ERR'
                    cleaned_data.append(entry)
                    continue
                
                # Check for outliers in valid times
                if entry['split1'] is not None and split_stats[0] is not None:
                    if (entry['split1'] < split_stats[1] or 
                        entry['split1'] > split_stats[2]):
                        is_outlier = True
                
                if entry['finish'] is not None and finish_stats[0] is not None:
                    if (entry['finish'] < finish_stats[1] or 
                        entry['finish'] > finish_stats[2]):
                        is_outlier = True
            
            if is_outlier:
                outliers.append(entry)
            else:
                cleaned_data.append(entry)
        
        return cleaned_data, outliers
   
    def add_analysis_graphs(self, ws, timing_data, start_row):
        """
        Adds analysis graphs to the Excel worksheet showing split and finish time comparisons.
        Positions the first graph 2 rows after the last run data.
        """
        from openpyxl.chart import LineChart, Reference, Series
        from openpyxl.chart.marker import Marker

        current_row = start_row + 2

        # Prepare data for graphs
        run_numbers = sorted(timing_data.keys(), key=int)
        athletes = {}

        # First pass: collect all athlete data
        for run_num in run_numbers:
            run_data = timing_data[run_num]
            cleaned_data, _ = self.clean_and_analyze_timing_data(run_data)
            
            for entry in cleaned_data:
                if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS']:
                    bib = entry['bib']
                    athlete_name = self.get_athlete_name(bib)
                    
                    if athlete_name not in athletes:
                        athletes[athlete_name] = {'splits': {}, 'finishes': {}}
                    
                    if entry['split1'] is not None:
                        athletes[athlete_name]['splits'][run_num] = entry['split1']
                    if entry['finish'] is not None:
                        athletes[athlete_name]['finishes'][run_num] = entry['finish']

        if not athletes:
            print("No valid athlete data found for graphs")
            return current_row

        # Write data for graphs in a less visible area (column J onwards)
        data_start_row = current_row
        ws.cell(row=current_row, column=10, value="Run")  # Column J
        col = 11
        for athlete in athletes:
            ws.cell(row=current_row, column=col, value=f"{athlete} Split")
            ws.cell(row=current_row, column=col + 1, value=f"{athlete} Finish")
            col += 2
        current_row += 1

        # Write data rows
        for run_num in run_numbers:
            ws.cell(row=current_row, column=10, value=int(run_num))
            col = 11
            for athlete in athletes:
                split_time = athletes[athlete]['splits'].get(run_num)
                finish_time = athletes[athlete]['finishes'].get(run_num)
                
                ws.cell(row=current_row, column=col, value=split_time)
                ws.cell(row=current_row, column=col + 1, value=finish_time)
                
                # Format cells as numbers
                if split_time is not None:
                    ws.cell(row=current_row, column=col).number_format = '0.00'
                if finish_time is not None:
                    ws.cell(row=current_row, column=col + 1).number_format = '0.00'
                
                col += 2
            current_row += 1

        data_end_row = current_row - 1

        # Create split times chart
        split_chart = LineChart()
        split_chart.title = "Split Times Comparison"
        split_chart.style = 2
        split_chart.height = 15
        split_chart.width = 25
        split_chart.x_axis.title = "Run Number"
        split_chart.y_axis.title = "Split Time (seconds)"

        # Create finish times chart
        finish_chart = LineChart()
        finish_chart.title = "Finish Times Comparison"
        finish_chart.style = 2
        finish_chart.height = 15
        finish_chart.width = 25
        finish_chart.x_axis.title = "Run Number"
        finish_chart.y_axis.title = "Finish Time (seconds)"

        # Colors for different athletes
        colors = ['FF0000', '00FF00', '0000FF', 'FF00FF', '00FFFF', 'FFA500', '800080']

        # Add data series for each athlete
        col = 11  # Start from column K
        for i, athlete in enumerate(athletes):
            # Split times series
            split_values = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=col)
            split_series = Series(split_values, title=f"{athlete} Split")
            split_series.marker = Marker(symbol='circle', size=7)
            split_series.smooth = True
            split_series.graphicalProperties.line.width = 20000
            split_series.graphicalProperties.line.solidFill = colors[i % len(colors)]
            split_chart.series.append(split_series)

            # Finish times series
            finish_values = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=col + 1)
            finish_series = Series(finish_values, title=f"{athlete} Finish")
            finish_series.marker = Marker(symbol='circle', size=7)
            finish_series.smooth = True
            finish_series.graphicalProperties.line.width = 20000
            finish_series.graphicalProperties.line.solidFill = colors[i % len(colors)]
            finish_chart.series.append(finish_series)

            col += 2

        # Add x-axis labels (run numbers)
        run_labels = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=10)  # Column J
        split_chart.set_categories(run_labels)
        finish_chart.set_categories(run_labels)

        # Position the charts - first chart just 2 rows after the last run data
        first_chart_row = start_row + 2  # This puts it 2 rows after the last run data
        ws.add_chart(split_chart, f"B{first_chart_row}")
        ws.add_chart(finish_chart, f"B{first_chart_row + 19}")  # Position second chart after the first

        return first_chart_row + 38  # Return the row after both charts


    def add_analysis_graphs(self, ws, timing_data, start_row):
        """
        Adds analysis graphs with custom axis labels and intervals.
        Now includes a third graph for split-finish times.
        """
        from openpyxl.chart import LineChart, Reference, Series
        from openpyxl.chart.marker import Marker
        from openpyxl.chart.axis import ChartLines
        import math

        # Graph positioning and sizing variables
        CHART_WIDTH = 25  # Width of charts
        CHART_HEIGHT = 15  # Height of charts
        VERTICAL_SPACING = 19  # Spacing between charts
        START_COLUMN = "B"  # Starting column for charts
        
        current_row = start_row + 2

        # Prepare data for graphs
        run_numbers = sorted(timing_data.keys(), key=int)
        athletes = {}

        # First pass: collect all athlete data including DNF with valid splits
        for run_num in run_numbers:
            run_data = timing_data[run_num]
            
            for entry in run_data:
                bib = entry['bib']
                athlete_name = self.get_athlete_name(bib)
                status = entry['status'].upper()
                
                if athlete_name not in athletes:
                    athletes[athlete_name] = {'splits': {}, 'finishes': {}, 'split_finishes': {}}
                
                # Include split times even for DNF entries if they have valid split times
                if entry['split1'] is not None and entry['split1'] >= 10.0:
                    athletes[athlete_name]['splits'][run_num] = entry['split1']
                
                # Calculate and store split-finish times (time between split and finish)
                if (status not in ['DNF', 'DSQ', 'DNS', 'ERR'] and 
                    entry['finish'] is not None and entry['split1'] is not None and 
                    entry['finish'] >= 10.0 and entry['split1'] >= 10.0):
                    split_finish_time = entry['finish'] - entry['split1']
                    if split_finish_time > 0:  # Only store positive time differences
                        athletes[athlete_name]['split_finishes'][run_num] = split_finish_time
                        athletes[athlete_name]['finishes'][run_num] = entry['finish']

        if not athletes:
            print("No valid athlete data found for graphs")
            return current_row

        # Add an extra run number for spacing
        max_run = max(run_numbers, key=int)
        extra_run = str(int(max_run) + 1)
        run_numbers.append(extra_run)

        # Find min/max times for axis scaling
        min_split = float('inf')
        max_split = float('-inf')
        min_finish = float('inf')
        max_finish = float('-inf')
        min_split_finish = float('inf')
        max_split_finish = float('-inf')

        for athlete_data in athletes.values():
            # Split times
            split_times = [time for time in athlete_data['splits'].values() if time is not None]
            if split_times:
                min_split = min(min_split, min(split_times))
                max_split = max(max_split, max(split_times))
            
            # Finish times
            finish_times = [time for time in athlete_data['finishes'].values() if time is not None]
            if finish_times:
                min_finish = min(min_finish, min(finish_times))
                max_finish = max(max_finish, max(finish_times))
                
            # Split-finish times
            split_finish_times = [time for time in athlete_data['split_finishes'].values() if time is not None]
            if split_finish_times:
                min_split_finish = min(min_split_finish, min(split_finish_times))
                max_split_finish = max(max_split_finish, max(split_finish_times))

        # Write data for graphs
        data_start_row = current_row
        ws.cell(row=current_row, column=10, value="Run")  # Column J
        col = 11
        for athlete in athletes:
            ws.cell(row=current_row, column=col, value=f"{athlete} Split")
            ws.cell(row=current_row, column=col + 1, value=f"{athlete} Finish")
            ws.cell(row=current_row, column=col + 2, value=f"{athlete} Split-Finish")
            col += 3
        current_row += 1

        # Write data rows including the extra run
        for run_num in run_numbers:
            ws.cell(row=current_row, column=10, value=int(run_num))
            col = 11
            for athlete in athletes:
                split_time = athletes[athlete]['splits'].get(run_num) if run_num != extra_run else None
                finish_time = athletes[athlete]['finishes'].get(run_num) if run_num != extra_run else None
                split_finish_time = athletes[athlete]['split_finishes'].get(run_num) if run_num != extra_run else None
                
                ws.cell(row=current_row, column=col, value=split_time)
                ws.cell(row=current_row, column=col + 1, value=finish_time)
                ws.cell(row=current_row, column=col + 2, value=split_finish_time)
                
                for offset in range(3):
                    cell = ws.cell(row=current_row, column=col + offset)
                    if cell.value is not None:
                        cell.number_format = '0.00'
                
                col += 3
            current_row += 1

        data_end_row = current_row - 1

        def create_chart(title, y_axis_title, min_val, max_val, unit):
            """Helper function to create a chart with consistent formatting"""
            chart = LineChart()
            chart.title = title
            chart.style = 2
            chart.height = CHART_HEIGHT
            chart.width = CHART_WIDTH
            chart.x_axis.title = "Run Number"
            chart.y_axis.title = y_axis_title

            if min_val != float('inf') and max_val != float('-inf'):
                y_min = max(0, min_val - 0.3)
                y_max = max_val + 0.3
                
                # Round to nearest unit
                y_min = math.floor(y_min * (1/unit)) / (1/unit)
                y_max = math.ceil(y_max * (1/unit)) / (1/unit)
                
                chart.y_axis.scaling.min = y_min
                chart.y_axis.scaling.max = y_max
                chart.y_axis.majorUnit = unit
                chart.y_axis.minorUnit = unit
                chart.y_axis.majorGridlines = ChartLines()
                chart.y_axis.minorGridlines = None
                chart.y_axis.numFmt = '0.00'

            # Configure x-axis
            chart.x_axis.majorUnit = 1
            chart.x_axis.minorUnit = 1
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.crosses = "min"
            chart.x_axis.numFmt = '0'

            return chart

        # Create all three charts
        split_chart = create_chart("Split Times Comparison", "Split Time (seconds)", min_split, max_split, 0.05)
        finish_chart = create_chart("Finish Times Comparison", "Finish Time (seconds)", min_finish, max_finish, 0.1)
        split_finish_chart = create_chart("Split-Finish Times Comparison", "Split-Finish Time (seconds)", 
                                        min_split_finish, max_split_finish, 0.05)

        # Colors for different athletes
        colors = ['FF0000', '00FF00', '0000FF', 'FF00FF', '00FFFF', 'FFA500', '800080']

        # Add data series for each athlete
        col = 11  # Start from column K
        for i, athlete in enumerate(athletes):
            athlete_color = colors[i % len(colors)]
            
            # Helper function to create series
            def create_series(values, title):
                series = Series(values, title=title)
                series.marker = Marker(symbol='circle', size=7)
                series.smooth = True
                series.graphicalProperties.line.width = 9525
                series.graphicalProperties.line.solidFill = athlete_color
                series.graphicalProperties.line.alpha = 40000
                series.marker.graphicalProperties.solidFill = athlete_color
                series.marker.graphicalProperties.line.solidFill = athlete_color
                return series

            # Add series to each chart
            split_values = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=col)
            split_chart.series.append(create_series(split_values, f"{athlete} Split"))

            finish_values = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=col + 1)
            finish_chart.series.append(create_series(finish_values, f"{athlete} Finish"))

            split_finish_values = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=col + 2)
            split_finish_chart.series.append(create_series(split_finish_values, f"{athlete} Split-Finish"))

            col += 3

        # Add x-axis labels to all charts
        run_labels = Reference(ws, min_row=data_start_row + 1, max_row=data_end_row, min_col=10)
        split_chart.set_categories(run_labels)
        finish_chart.set_categories(run_labels)
        split_finish_chart.set_categories(run_labels)

        # Position all three charts
        first_chart_row = start_row + 2
        ws.add_chart(split_chart, f"{START_COLUMN}{first_chart_row}")
        ws.add_chart(finish_chart, f"{START_COLUMN}{first_chart_row + VERTICAL_SPACING}")
        ws.add_chart(split_finish_chart, f"{START_COLUMN}{first_chart_row + (2 * VERTICAL_SPACING)}")

        return first_chart_row + (3 * VERTICAL_SPACING)



    def reformat_file(self):
        """Handle the reformatting of the selected file with validation checks."""
        # Check if a file is selected
        if not self.selected_file:  # Changed condition to check selected_file
            messagebox.showwarning("Error", "Please select a Brower CSV file first.")
            return

        # Check if an event is selected
        if not self.event_var.get():
            messagebox.showwarning("Error", "Please select an Event type (SL/GS/SG/DH/SX).")
            return

        # Check if a hill is specified
        if not self.hill_var.get().strip():
            messagebox.showwarning("Error", "Please specify a Hill name.")
            return

        try:
            # Generate the default filename
            default_filename = self.generate_filename()
            print(f"Generated filename: {default_filename}")  # Debug print
            
            # Add the hill to recent hills memory for future autocomplete
            hill_name = self.hill_var.get().strip()
            self.add_hill_to_memory(hill_name)
            
            # Ask user where to save the reformatted file
            initial_dir = os.path.dirname(self.selected_file) if self.selected_file else "."
            output_file = filedialog.asksaveasfilename(
                initialdir=initial_dir,
                initialfile=default_filename,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if output_file:  # If user didn't cancel the save dialog
                if self.create_formatted_excel(output_file):
                    messagebox.showinfo("Success", "File has been reformatted and saved successfully.")
                else:
                    messagebox.showerror("Error", "Failed to create reformatted file.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            print(f"Error in reformat_file: {str(e)}")  # Debug print


   

if __name__ == "__main__":
    root = tk.Tk()
    app = TimingSystemApp(root)
    root.mainloop()
