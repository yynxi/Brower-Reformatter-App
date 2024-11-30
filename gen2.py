# This is the new code, updated but not complete.
import os
import json
import tkinter as tk
import tkinter as ttk
import sv_ttk
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

class TimingSystemApp:
    def __init__(self, root):
        """Initialize the application with enhanced timing capabilities."""
        self.root = root
        self.root.title("Timing System Data Formatter")
        
        # Apply Sun Valley theme
        sv_ttk.set_theme("dark")

        # Configuration
        self.VERSION = "1.1"  # Updated for multi-split support
        self.AUTHOR = "Julian H. Brunet"
        self.MIN_REGULAR_SPLIT_TIME = 3.0
        self.MIN_ACCELERATION_SPLIT_TIME = 0.5
        self.MAX_SPLIT_TIME = 35.0
        
        # Data structures
        self.selected_file = None
        self.athletes = {"SQAH": [], "SQAF": [], "OTHER": []}
        self.temp_guests = []
        self.current_team = "SQAH"
        self.num_splits = 0  # Will be determined from data
        
        # UI state variables
        self.team_var = tk.StringVar()
        self.event_var = tk.StringVar()
        self.hill_var = tk.StringVar()
        self.snow_condition_var = tk.StringVar()
        self.sky_condition_var = tk.StringVar()
        self.precipitation_var = tk.StringVar()
        self.wind_condition_var = tk.StringVar()
        self.date_var = tk.StringVar()
        self.time_var = tk.StringVar()
        self.session_var = tk.StringVar()
        
        # Load saved data
        self.load_settings()
        self.load_recent_names()
        self.load_recent_hills()
        self.load_athletes_from_json()
        
        # Build GUI
        self.build_gui()
        
    def parse_timing_data(self, file_path):
        """
        Parse CSV timing data with enhanced split time handling.
        Detects number of splits automatically and validates data.
        """
        timing_data = {}
        header_found = False
        column_indices = {}
        split_columns = []
        
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                separator = '>'
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    if not header_found and "Bib#" in line:
                        header_found = True
                        headers = line.split(separator)
                        
                        # Get all column indices
                        for i, header in enumerate(headers):
                            header = header.lower().strip()
                            if "bib" in header:
                                column_indices['bib'] = i
                            elif "run" in header:
                                column_indices['run'] = i
                            elif "split" in header:
                                split_columns.append(i)
                            elif "finish time" in header:
                                column_indices['finish'] = i
                            elif "status" in header:
                                column_indices['status'] = i
                        
                        self.num_splits = len(split_columns)
                        continue
                    
                    if header_found:
                        data = line.split(separator)
                        if len(data) <= 1:
                            continue
                        
                        try:
                            run_number = data[column_indices['run']].strip()
                            if not run_number:
                                continue
                            
                            if run_number not in timing_data:
                                timing_data[run_number] = []
                            
                            # Process all split times
                            splits = []
                            for split_col in split_columns:
                                split_time = self.validate_time(data[split_col].strip())
                                splits.append(split_time)
                            
                            # Create entry
                            entry = {
                                'bib': data[column_indices['bib']].strip(),
                                'splits': splits,
                                'finish': self.validate_time(
                                    data[column_indices.get('finish', -1)].strip() 
                                    if 'finish' in column_indices else ''
                                ),
                                'status': data[column_indices.get('status', -1)].strip() 
                                         if 'status' in column_indices else '',
                                'run': run_number,
                                'error_details': []  # For storing validation messages
                            }
                            
                            timing_data[run_number].append(entry)
                            
                        except IndexError as e:
                            print(f"Error processing line: {line}")
                            print(f"Error details: {str(e)}")
                            continue
                
                # Validate and clean data for each run
                for run_number in timing_data:
                    timing_data[run_number] = self.validate_run_data(timing_data[run_number])
                
                if not timing_data:
                    messagebox.showerror("Error", "No valid timing data found in the file.")
                    return None
                
                return timing_data
                
        except Exception as e:
            messagebox.showerror("Error", f"Error parsing CSV file: {str(e)}")
            return None

    def validate_time(self, time_str):
        """
        Basic time string validation and conversion.
        No minimum time requirement - that's handled in run validation.
        """
        if not time_str or not isinstance(time_str, str):
            return None
            
        time_str = time_str.strip().upper()
        if time_str in ['DNF', 'DSQ', '', 'DNS', '0']:
            return None
            
        try:
            if ':' in time_str:
                minutes, seconds = time_str.split(':')
                return float(minutes) * 60 + float(seconds)
            return float(time_str)
        except (ValueError, TypeError):
            return None

    def validate_run_data(self, run_data):
        """
        Validates timing data for a complete run with enhanced split time handling.
        """
        # First pass: collect times for each split position
        split_times = [[] for _ in range(self.num_splits)]
        finish_times = []
        
        for entry in run_data:
            if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']:
                # Collect all non-zero split times
                for i, split in enumerate(entry['splits']):
                    if split is not None and split > 0:
                        split_times[i].append(split)
                
                if entry['finish'] is not None and entry['finish'] > 0:
                    finish_times.append(entry['finish'])
        
        # Analyze characteristics of each split
        split_characteristics = []
        for split_position, times in enumerate(split_times):
            if times:
                mean = sum(times) / len(times)
                std_dev = (sum((x - mean) ** 2 for x in times) / len(times)) ** 0.5
                
                is_acceleration_split = mean < 5.0 and std_dev < 0.5
                
                bounds = self.calculate_adaptive_bounds(
                    times,
                    z_score=3.0 if is_acceleration_split else 2.5,
                    min_allowed=self.MIN_ACCELERATION_SPLIT_TIME if is_acceleration_split 
                              else self.MIN_REGULAR_SPLIT_TIME
                )
                
                split_characteristics.append({
                    'mean': mean,
                    'std_dev': std_dev,
                    'is_acceleration_split': is_acceleration_split,
                    'bounds': bounds
                })
            else:
                split_characteristics.append(None)
        
        # Calculate finish time bounds
        finish_bounds = self.calculate_adaptive_bounds(finish_times) if finish_times else None
        
        # Second pass: validate entries
        valid_entries = []
        for entry in run_data:
            status = entry['status'].upper()
            
            # Skip DNS entries
            if status == 'DNS':
                valid_entries.append(entry)
                continue
                
            is_valid = True
            error_details = []
            
            # Validate split times
            valid_splits = []
            for i, split in enumerate(entry['splits']):
                if split is not None and split > 0:
                    char = split_characteristics[i]
                    if char and char['bounds']:
                        bounds = char['bounds']
                        if bounds[0] <= split <= bounds[1]:
                            valid_splits.append(split)
                        else:
                            error_details.append(
                                f"Split {i+1}: {split:.2f}s outside bounds "
                                f"[{bounds[0]:.2f}, {bounds[1]:.2f}]"
                            )
                            is_valid = False
            
            # Check split progression
            if len(valid_splits) > 1:
                for j in range(1, len(valid_splits)):
                    if valid_splits[j] <= valid_splits[j-1]:
                        error_details.append(
                            f"Invalid progression: {valid_splits[j-1]:.2f} → {valid_splits[j]:.2f}"
                        )
                        is_valid = False
            
            # Validate finish time
            if entry['finish'] is not None and entry['finish'] > 0:
                if finish_bounds:
                    if not (finish_bounds[0] <= entry['finish'] <= finish_bounds[1]):
                        error_details.append(
                            f"Finish: {entry['finish']:.2f}s outside bounds "
                            f"[{finish_bounds[0]:.2f}, {finish_bounds[1]:.2f}]"
                        )
                        is_valid = False
                if valid_splits and entry['finish'] <= valid_splits[-1]:
                    error_details.append(
                        f"Finish ({entry['finish']:.2f}) ≤ last split ({valid_splits[-1]:.2f})"
                    )
                    is_valid = False
            
            # Update entry status
            if not is_valid and status not in ['DNF', 'DSQ']:
                entry['status'] = 'ERR'
            entry['error_details'] = error_details
            valid_entries.append(entry)
        
        return sorted(valid_entries, key=lambda x: (
            x['status'].upper() in ['DNF', 'DSQ', 'DNS', 'ERR'],
            float('inf') if x['finish'] is None else x['finish']
        ))

    def calculate_adaptive_bounds(self, times, z_score=2.5, min_allowed=3.0, max_allowed=180.0):
        """
        Calculate statistical bounds with adaptive thresholds.
        """
        if not times:
            return None
            
        mean = sum(times) / len(times)
        std_dev = (sum((x - mean) ** 2 for x in times) / len(times)) ** 0.5
        
        # For very consistent times, widen bounds slightly
        if std_dev < 0.1:
            z_score = z_score * 1.5
        
        lower_bound = max(min_allowed, mean - (z_score * std_dev))
        upper_bound = min(max_allowed, mean + (z_score * std_dev))
        
        # Adjust bounds for acceleration splits
        if mean < 5.0:
            relative_range = 0.5  # 50% variation allowed
            lower_bound = max(min_allowed, mean * (1 - relative_range))
            upper_bound = mean * (1 + relative_range)
        
        return (lower_bound, upper_bound)

    def format_time(self, time_value, as_difference=False):
        """Format time values with appropriate precision."""
        if time_value is None:
            return ""
            
        if as_difference:
            sign = '+' if time_value >= 0 else ''
            if abs(time_value) < 60:
                return f"{sign}{time_value:.3f}" if abs(time_value) < 1 else f"{sign}{time_value:.2f}"
            minutes = int(abs(time_value) // 60)
            seconds = abs(time_value) % 60
            return f"{sign}{minutes}:{seconds:05.2f}"
        
        if time_value < 60:
            return f"{time_value:.3f}" if time_value < 1 else f"{time_value:.2f}"
        minutes = int(time_value // 60)
        seconds = time_value % 60
        return f"{minutes}:{seconds:05.2f}"
    

    def create_formatted_excel(self, output_path):
            """Creates a formatted Excel file with enhanced split time handling."""
            wb = Workbook()
            ws = wb.active
            
            # Define styles
            try:
                title_font = Font(name='Avenir Next LT Pro', size=18, bold=True)
            except:
                title_font = Font(name='Arial', size=18, bold=True)
            
            normal_font = Font(name='Arial', size=11)
            header_font = Font(name='Arial', size=11, bold=True)
            
            # Define borders
            header_border = Border(
                left=Side(style='thick'), 
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            normal_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
                
            # Set dynamic column widths based on number of splits
            basic_columns = {
                'A': 6,   # Index
                'B': 18,  # Bib #
                'C': 20,  # Name
            }
            
            # Add columns for each split
            split_columns_per_set = 3  # Time, Diff, Rank for each split
            current_col = 'D'
            for i in range(self.num_splits):
                for _ in range(split_columns_per_set):
                    ws.column_dimensions[current_col] = 12
                    current_col = get_column_letter(ord(current_col) + 1)
            
            # Add finish columns
            finish_columns = {
                current_col: 12,      # Finish Time
                chr(ord(current_col) + 1): 12,  # Finish Diff
                chr(ord(current_col) + 2): 10   # Status
            }
            
            # Apply basic column widths
            for col, width in {**basic_columns, **finish_columns}.items():
                ws.column_dimensions[col].width = width
                
            # Set row height for title
            ws.row_dimensions[1].height = 30
            
            # Get season and team name
            season = self.get_season(self.date_var.get())
            current_team_name = self.team_names[self.current_team]
            team_with_season = f"{current_team_name} - {season}" if season else current_team_name
            
            # Write title and header information
            ws['B1'] = self.excel_title
            ws['B1'].font = title_font
            ws['B1'].alignment = Alignment(vertical='center', horizontal='left')

            # Header information
            header_info = {
                ('B2', 'Team:'): team_with_season,
                ('B3', 'Session #:'): self.session_var.get(),
                ('B4', 'Event:'): self.event_var.get(),
                ('B5', 'Snow Condition:'): self.snow_condition_var.get(),
                ('E2', 'Date:'): self.date_var.get(),
                ('E3', 'Start Time:'): self.time_var.get(),
                ('E4', 'Hill:'): self.hill_var.get(),
                ('E5', 'Weather:'): f"{self.sky_condition_var.get()}, {self.precipitation_var.get()}, {self.wind_condition_var.get()}"
            }

            for (cell_coord, label), value in header_info.items():
                # Write label
                ws[cell_coord] = label
                ws[cell_coord].font = header_font
                ws[cell_coord].alignment = Alignment(horizontal='left')
                
                # Write value in next column
                value_col = chr(ord(cell_coord[0]) + 1)
                value_cell = f"{value_col}{cell_coord[1]}"
                ws[value_cell] = value
                ws[value_cell].font = normal_font
                ws[value_cell].alignment = Alignment(horizontal='left')

            # Process timing data
            if self.selected_file:
                timing_data = self.parse_timing_data(self.selected_file)
                if timing_data:
                    current_row = 8  # Start after header section
                    
                    for run_number in sorted(timing_data.keys(), key=int):
                        current_row = self.write_run_data(ws, timing_data[run_number], current_row)
                        current_row += 1  # Space between runs
                    
                    # Add analysis graphs
                    current_row = self.add_analysis_graphs(ws, timing_data, current_row)
                    
                    # Create athlete analysis sheet
                    self.create_athlete_analysis_sheet(wb, timing_data)
            
            try:
                wb.save(output_path)
                return True
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save Excel file: {str(e)}")
                return False

    def write_run_data(self, ws, run_data, start_row):
        """Writes run data with enhanced split time handling and validation."""
        # Run header
        current_row = start_row + 2
        ws[f'B{current_row}'] = f"Run {run_data[0]['run']}"
        current_row += 1
        
        # Generate headers based on number of splits
        headers = ['Bib #', 'Name']
        for i in range(self.num_splits):
            headers.extend([f'Split {i+1}', 'Diff.', 'Rank'])
        headers.extend(['Finish Time', 'Diff.', 'Status'])
        
        # Write headers
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Arial', size=11, bold=True)
        
        current_row += 1
        
        # Collect and process valid times for each split
        split_data = [[] for _ in range(self.num_splits)]
        finish_data = []
        
        for entry in run_data:
            if entry['status'].upper() not in ['DNS', 'ERR']:
                for i, split in enumerate(entry['splits']):
                    if split is not None and split > 0:
                        split_data[i].append((split, entry['bib']))
                
                if (entry['status'].upper() not in ['DNF', 'DSQ'] and 
                    entry['finish'] is not None and entry['finish'] > 0):
                    finish_data.append((entry['finish'], entry['bib']))
        
        # Sort times and calculate ranks
        for i in range(self.num_splits):
            split_data[i].sort(key=lambda x: x[0])
        finish_data.sort(key=lambda x: x[0])
        
        # Write data rows with enhanced formatting
        for entry in run_data:
            bib = entry['bib']
            status = entry['status'].upper()
            
            # Base row data
            row_data = [bib, self.get_athlete_name(bib)]
            
            # Process each split
            for i, split in enumerate(entry['splits']):
                if split is not None and split > 0:
                    best_split = split_data[i][0][0] if split_data[i] else None
                    split_diff = split - best_split if best_split is not None else None
                    split_rank = next((j + 1 for j, (_, b) in enumerate(split_data[i]) if b == bib), '')
                    
                    row_data.extend([
                        self.format_time(split),
                        self.format_time(split_diff, True) if split_diff is not None else '',
                        split_rank
                    ])
                else:
                    row_data.extend(['', '', ''])
            
            # Add finish data
            if status not in ['DNF', 'DSQ', 'DNS', 'ERR']:
                best_finish = finish_data[0][0] if finish_data else None
                finish_diff = (entry['finish'] - best_finish 
                             if entry['finish'] is not None and best_finish is not None 
                             else None)
                
                row_data.extend([
                    self.format_time(entry['finish']),
                    self.format_time(finish_diff, True) if finish_diff is not None else '',
                    status
                ])
            else:
                row_data.extend(['', '', status])
            
            # Write row with formatting
            for col, value in enumerate(row_data, start=2):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
                
                # Apply gradient coloring for valid times
                if status not in ['DNS', 'ERR']:
                    # Color split times
                    if (col - 4) % 3 == 0 and (col - 4) // 3 < len(split_data):
                        split_idx = (col - 4) // 3
                        if split_data[split_idx]:
                            rank = next((i for i, (_, b) in enumerate(split_data[split_idx]) 
                                      if b == bib), None)
                            if rank is not None:
                                cell.fill = self.get_gradient_color(
                                    rank, len(split_data[split_idx]), 
                                    0.6 if split_idx == 0 else 0.5
                                )
                    
                    # Color finish time
                    elif col == len(row_data) - 1 and finish_data:
                        rank = next((i for i, (_, b) in enumerate(finish_data) if b == bib), None)
                        if rank is not None:
                            cell.fill = self.get_gradient_color(rank, len(finish_data), 0.5)
            
            current_row += 1
            
            # Add error details if present
            if entry['error_details']:
                error_cell = ws.cell(row=current_row, column=2, 
                                   value=f"Validation Issues for Bib {bib}:")
                error_cell.font = Font(color="FF0000")
                current_row += 1
                
                for error in entry['error_details']:
                    ws.cell(row=current_row, column=2, value=f"  • {error}").font = Font(color="FF0000")
                    current_row += 1
        
        return current_row
    
    def create_athlete_analysis_sheet(self, wb, timing_data):
            """Creates detailed athlete analysis sheet with enhanced split analysis."""
            analysis_sheet = wb.create_sheet(title="Athlete Analysis")
            
            # Apply column formatting
            base_width = 12
            analysis_sheet.column_dimensions['A'].width = 6   # Index
            analysis_sheet.column_dimensions['B'].width = 18  # Bib
            analysis_sheet.column_dimensions['C'].width = 25  # Name
            
            # Dynamic columns for splits
            current_col = 'D'
            for i in range(self.num_splits):
                for _ in range(3):  # Time, Diff, Rank for each split
                    analysis_sheet.column_dimensions[current_col] = base_width
                    current_col = get_column_letter(ord(current_col) + 1)
            
            # Finish columns
            analysis_sheet.column_dimensions[current_col] = base_width      # Finish Time
            analysis_sheet.column_dimensions[chr(ord(current_col) + 1)] = base_width  # Finish Diff
            analysis_sheet.column_dimensions[chr(ord(current_col) + 2)] = base_width  # Status

            current_row = 2
            
            # Process each athlete
            for athlete in sorted(self.athletes[self.current_team], key=lambda x: int(x['bib'])):
                athlete_bib = int(athlete['bib'])
                athlete_name = athlete['name']
                
                # Write athlete header
                athlete_header = f"{athlete_name} (Bib {athlete_bib})"
                cell = analysis_sheet.cell(row=current_row, column=2, value=athlete_header)
                cell.font = Font(name='Arial', size=11, bold=True)
                current_row += 1

                # Generate headers
                headers = ['Bib #', 'Run']
                for i in range(self.num_splits):
                    headers.extend([f'Split {i+1}', 'Diff.', 'Rank'])
                headers.extend(['Finish', 'Diff.', 'Status'])
                
                # Write column headers
                for col, header in enumerate(headers, start=2):
                    cell = analysis_sheet.cell(row=current_row, column=col, value=header)
                    cell.border = Border(
                        left=Side(style='thick'),
                        right=Side(style='thick'),
                        top=Side(style='thick'),
                        bottom=Side(style='thick')
                    )
                    cell.alignment = Alignment(horizontal='center')
                current_row += 1

                # Process each run
                athlete_data = []
                for run_number in sorted(timing_data.keys(), key=int):
                    run_data = timing_data[run_number]
                    
                    # Find athlete's entry and best times
                    athlete_entry = next((entry for entry in run_data 
                                        if int(entry['bib']) == athlete_bib), None)
                    if not athlete_entry:
                        continue

                    # Collect valid times for this run
                    run_splits = [[] for _ in range(self.num_splits)]
                    run_finishes = []
                    
                    for entry in run_data:
                        if entry['status'].upper() not in ['DNS', 'ERR']:
                            # Collect split times
                            for i, split in enumerate(entry['splits']):
                                if split is not None and split > 0:
                                    run_splits[i].append((split, entry['bib']))
                            
                            # Collect finish times
                            if (entry['status'].upper() not in ['DNF', 'DSQ'] and 
                                entry['finish'] is not None and entry['finish'] > 0):
                                run_finishes.append((entry['finish'], entry['bib']))

                    # Sort times
                    for split_list in run_splits:
                        split_list.sort(key=lambda x: x[0])
                    run_finishes.sort(key=lambda x: x[0])

                    # Prepare row data
                    row_data = [athlete_bib, f"Run {run_number}"]
                    
                    # Add split data
                    for i, split in enumerate(athlete_entry['splits']):
                        if split is not None and split > 0:
                            best_split = run_splits[i][0][0] if run_splits[i] else None
                            split_diff = split - best_split if best_split is not None else None
                            split_rank = next((j + 1 for j, (_, b) in enumerate(run_splits[i]) 
                                            if b == str(athlete_bib)), '')
                            
                            row_data.extend([
                                self.format_time(split),
                                self.format_time(split_diff, True) if split_diff is not None else '',
                                split_rank
                            ])
                        else:
                            row_data.extend(['', '', ''])

                    # Add finish data
                    if athlete_entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']:
                        best_finish = run_finishes[0][0] if run_finishes else None
                        finish_diff = (athlete_entry['finish'] - best_finish 
                                    if athlete_entry['finish'] is not None and best_finish is not None 
                                    else None)
                        
                        row_data.extend([
                            self.format_time(athlete_entry['finish']),
                            self.format_time(finish_diff, True) if finish_diff is not None else '',
                            athlete_entry['status']
                        ])
                    else:
                        row_data.extend(['', '', athlete_entry['status']])

                    # Store data for graphing
                    athlete_data.append({
                        'run': int(run_number),
                        'splits': athlete_entry['splits'],
                        'finish': athlete_entry['finish'],
                        'status': athlete_entry['status']
                    })

                    # Write row with formatting
                    for col, value in enumerate(row_data, start=2):
                        cell = analysis_sheet.cell(row=current_row, column=col, value=value)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center')
                    
                    current_row += 1

                # Add progression graphs for this athlete
                current_row = self.add_athlete_progression_graphs(
                    analysis_sheet, athlete_data, athlete_name, current_row
                )
                current_row += 2  # Space between athletes

            return current_row
    
    def add_analysis_to_sheet(self, ws, timing_data, start_row):
        """Creates comprehensive analysis of timing data with enhanced split handling."""
        current_row = start_row + 2
        
        # Prepare data for analysis
        run_numbers = sorted(timing_data.keys(), key=int)
        athletes = {}
        
        # Collect data for each athlete across all runs
        for run_num in run_numbers:
            run_data = timing_data[run_num]
            
            for entry in run_data:
                if entry['status'].upper() not in ['DNS', 'ERR']:
                    bib = entry['bib']
                    athlete_name = self.get_athlete_name(bib)
                    
                    if athlete_name not in athletes:
                        athletes[athlete_name] = {
                            'splits': [[] for _ in range(self.num_splits)],
                            'finishes': [],
                            'best_splits': [None] * self.num_splits,
                            'best_finish': None,
                            'improvements': {'splits': [0] * self.num_splits, 'finish': 0},
                            'consistency': {'splits': [0] * self.num_splits, 'finish': 0}
                        }
                    
                    # Process split times
                    for i, split in enumerate(entry['splits']):
                        if split is not None and split > 0:
                            athletes[athlete_name]['splits'][i].append((run_num, split))
                            
                            # Track best times and improvements
                            if (athletes[athlete_name]['best_splits'][i] is None or 
                                split < athletes[athlete_name]['best_splits'][i]):
                                athletes[athlete_name]['best_splits'][i] = split
                                athletes[athlete_name]['improvements']['splits'][i] += 1
                    
                    # Process finish times
                    if (entry['status'].upper() not in ['DNF', 'DSQ'] and 
                        entry['finish'] is not None and entry['finish'] > 0):
                        athletes[athlete_name]['finishes'].append((run_num, entry['finish']))
                        
                        # Track best finish and improvements
                        if (athletes[athlete_name]['best_finish'] is None or 
                            entry['finish'] < athletes[athlete_name]['best_finish']):
                            athletes[athlete_name]['best_finish'] = entry['finish']
                            athletes[athlete_name]['improvements']['finish'] += 1
        
        if not athletes:
            return current_row
        
        # Calculate consistency scores
        for athlete_name, data in athletes.items():
            # Split consistency
            for i in range(self.num_splits):
                times = [t for _, t in data['splits'][i]]
                if len(times) >= 2:
                    mean = sum(times) / len(times)
                    variance = sum((t - mean) ** 2 for t in times) / len(times)
                    data['consistency']['splits'][i] = (1 / (1 + variance)) * 100
            
            # Finish consistency
            times = [t for _, t in data['finishes']]
            if len(times) >= 2:
                mean = sum(times) / len(times)
                variance = sum((t - mean) ** 2 for t in times) / len(times)
                data['consistency']['finish'] = (1 / (1 + variance)) * 100

        # Write analysis headers
        headers = ['Athlete', 'Runs Completed']
        for i in range(self.num_splits):
            headers.extend([f'Best S{i+1}', f'Imp. S{i+1}', f'Cons. S{i+1}'])
        headers.extend(['Best Finish', 'Finish Imp.', 'Finish Cons.'])
        
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
        current_row += 1

        # Write analysis data
        for athlete_name, data in athletes.items():
            row_data = [
                athlete_name,
                len(data['finishes'])
            ]
            
            # Add split analysis
            for i in range(self.num_splits):
                row_data.extend([
                    self.format_time(data['best_splits'][i]),
                    data['improvements']['splits'][i],
                    f"{data['consistency']['splits'][i]:.1f}%"
                ])
            
            # Add finish analysis
            row_data.extend([
                self.format_time(data['best_finish']),
                data['improvements']['finish'],
                f"{data['consistency']['finish']:.1f}%"
            ])
            
            for col, value in enumerate(row_data, start=2):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1

        # Add statistical summary
        current_row = self.add_statistical_summary(ws, athletes, current_row + 2)
        
        return current_row

    def add_statistical_summary(self, ws, athletes_data, start_row):
        """Adds statistical summary of the session."""
        current_row = start_row
        
        # Write section header
        cell = ws.cell(row=current_row, column=2, value="Statistical Summary")
        cell.font = Font(bold=True, size=12)
        current_row += 2

        # Calculate summary statistics
        stats = {
            'splits': [[] for _ in range(self.num_splits)],
            'finishes': [],
            'improvements': {'splits': [0] * self.num_splits, 'finish': 0},
            'dnf_rate': 0,
            'completion_rate': 0
        }

        total_runs = 0
        completed_runs = 0
        
        for athlete_data in athletes_data.values():
            # Process splits
            for i in range(self.num_splits):
                stats['splits'][i].extend(t for _, t in athlete_data['splits'][i])
                stats['improvements']['splits'][i] += athlete_data['improvements']['splits'][i]
            
            # Process finishes
            stats['finishes'].extend(t for _, t in athlete_data['finishes'])
            stats['improvements']['finish'] += athlete_data['improvements']['finish']
            
            # Count runs
            if athlete_data['finishes']:
                total_runs += len(athlete_data['finishes'])
                completed_runs += len([t for _, t in athlete_data['finishes'] if t is not None])

        # Calculate completion and DNF rates
        if total_runs > 0:
            stats['completion_rate'] = (completed_runs / total_runs) * 100
            stats['dnf_rate'] = ((total_runs - completed_runs) / total_runs) * 100

        # Write summary statistics
        summary_rows = [
            ("Total Athletes", len(athletes_data)),
            ("Total Runs", total_runs),
            ("Completion Rate", f"{stats['completion_rate']:.1f}%"),
            ("DNF Rate", f"{stats['dnf_rate']:.1f}%")
        ]

        for title, value in summary_rows:
            ws.cell(row=current_row, column=2, value=title).font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1

        # Add split-specific statistics
        current_row += 1
        for i in range(self.num_splits):
            if stats['splits'][i]:
                split_times = stats['splits'][i]
                mean = sum(split_times) / len(split_times)
                std_dev = (sum((t - mean) ** 2 for t in split_times) / len(split_times)) ** 0.5
                
                ws.cell(row=current_row, column=2, value=f"Split {i+1} Statistics:").font = Font(bold=True)
                current_row += 1
                
                split_stats = [
                    ("Average", self.format_time(mean)),
                    ("Std Dev", f"{std_dev:.3f}"),
                    ("Best", self.format_time(min(split_times))),
                    ("Range", self.format_time(max(split_times) - min(split_times))),
                    ("Improvements", stats['improvements']['splits'][i])
                ]
                
                for title, value in split_stats:
                    ws.cell(row=current_row, column=3, value=title)
                    ws.cell(row=current_row, column=4, value=value)
                    current_row += 1
                current_row += 1

        return current_row
    
    def calculate_split_statistics(self, split_times, is_acceleration_split=False):
        """
        Calculates comprehensive statistics for a set of split times.
        Handles both regular and acceleration splits differently.
        """
        if not split_times:
            return None
            
        # Remove any None or zero values
        valid_times = [t for t in split_times if t is not None and t > 0]
        if not valid_times:
            return None
            
        # Basic statistics
        mean = sum(valid_times) / len(valid_times)
        variance = sum((t - mean) ** 2 for t in valid_times) / len(valid_times)
        std_dev = variance ** 0.5
        
        # Determine outlier bounds based on split type
        z_score = 3.0 if is_acceleration_split else 2.5
        min_allowed = self.MIN_ACCELERATION_SPLIT_TIME if is_acceleration_split else self.MIN_REGULAR_SPLIT_TIME
        
        lower_bound = max(min_allowed, mean - (z_score * std_dev))
        upper_bound = min(self.MAX_SPLIT_TIME, mean + (z_score * std_dev))
        
        return {
            'mean': mean,
            'std_dev': std_dev,
            'min': min(valid_times),
            'max': max(valid_times),
            'range': max(valid_times) - min(valid_times),
            'count': len(valid_times),
            'bounds': (lower_bound, upper_bound),
            'coefficient_of_variation': (std_dev / mean) * 100 if mean > 0 else None
        }

    def validate_run_consistency(self, run_data):
        """
        Validates the consistency of times within a run.
        Returns a list of potential issues found.
        """
        issues = []
        
        # Collect valid times for each split
        split_times = [[] for _ in range(self.num_splits)]
        finish_times = []
        
        for entry in run_data:
            if entry['status'].upper() not in ['DNS', 'ERR']:
                # Collect split times
                for i, split in enumerate(entry['splits']):
                    if split is not None and split > 0:
                        split_times[i].append((split, entry['bib']))
                
                # Collect finish times
                if (entry['status'].upper() not in ['DNF', 'DSQ'] and 
                    entry['finish'] is not None and entry['finish'] > 0):
                    finish_times.append((entry['finish'], entry['bib']))
        
        # Analyze each split
        for i, times in enumerate(split_times):
            if times:
                is_acceleration = i == 0 and all(t[0] < 5.0 for t in times)
                stats = self.calculate_split_statistics([t[0] for t in times], is_acceleration)
                
                if stats:
                    # Check for suspicious patterns
                    if stats['coefficient_of_variation'] > 15 and not is_acceleration:
                        issues.append(f"High variation in Split {i+1} times (CV: {stats['coefficient_of_variation']:.1f}%)")
                    
                    # Check for outliers
                    for time, bib in times:
                        if time < stats['bounds'][0] or time > stats['bounds'][1]:
                            issues.append(f"Suspicious time for Bib {bib} in Split {i+1}: {self.format_time(time)}")

        # Analyze finish times
        if finish_times:
            stats = self.calculate_split_statistics([t[0] for t in finish_times])
            if stats and stats['coefficient_of_variation'] > 15:
                issues.append(f"High variation in Finish times (CV: {stats['coefficient_of_variation']:.1f}%)")

        return issues

    def validate_athlete_progression(self, athlete_data):
        """
        Validates an athlete's time progression across runs.
        Returns a list of potential issues or anomalies.
        """
        issues = []
        
        # Analyze each split separately
        for split_index in range(self.num_splits):
            split_progression = []
            for run in athlete_data:
                if (run['status'].upper() not in ['DNS', 'ERR', 'DNF', 'DSQ'] and 
                    len(run['splits']) > split_index and 
                    run['splits'][split_index] is not None and 
                    run['splits'][split_index] > 0):
                    split_progression.append((run['run'], run['splits'][split_index]))
            
            if len(split_progression) >= 2:
                # Check for sudden time changes
                for i in range(1, len(split_progression)):
                    time_diff = split_progression[i][1] - split_progression[i-1][1]
                    if abs(time_diff) > 2.0:  # More than 2 seconds difference
                        issues.append(
                            f"Large time change in Split {split_index + 1} between "
                            f"runs {split_progression[i-1][0]} and {split_progression[i][0]}: "
                            f"{self.format_time(time_diff, True)}"
                        )

        # Analyze finish time progression
        finish_progression = []
        for run in athlete_data:
            if (run['status'].upper() not in ['DNS', 'ERR', 'DNF', 'DSQ'] and 
                run['finish'] is not None and run['finish'] > 0):
                finish_progression.append((run['run'], run['finish']))
        
        if len(finish_progression) >= 2:
            # Check for consistent progress
            time_differences = []
            for i in range(1, len(finish_progression)):
                time_diff = finish_progression[i][1] - finish_progression[i-1][1]
                time_differences.append(time_diff)
                
                if abs(time_diff) > 3.0:  # More than 3 seconds difference
                    issues.append(
                        f"Large finish time change between runs "
                        f"{finish_progression[i-1][0]} and {finish_progression[i][0]}: "
                        f"{self.format_time(time_diff, True)}"
                    )
            
            # Check for overall trend
            if len(time_differences) >= 2:
                improving = all(diff <= 0 for diff in time_differences)
                worsening = all(diff >= 0 for diff in time_differences)
                if worsening:
                    issues.append("Consistent pattern of increasing times - possible fatigue or equipment issue")

        return issues

    def analyze_split_relationships(self, run_data):
        """
        Analyzes relationships between splits and identifies potential timing issues.
        """
        issues = []
        
        for entry in run_data:
            if entry['status'].upper() not in ['DNS', 'ERR', 'DNF', 'DSQ']:
                valid_splits = [s for s in entry['splits'] if s is not None and s > 0]
                
                if len(valid_splits) >= 2:
                    # Check split time relationships
                    for i in range(1, len(valid_splits)):
                        time_diff = valid_splits[i] - valid_splits[i-1]
                        
                        # Flag suspicious split relationships
                        if time_diff <= 0:
                            issues.append(
                                f"Invalid split relationship for Bib {entry['bib']}: "
                                f"Split {i} ({self.format_time(valid_splits[i])}) <= "
                                f"Split {i-1} ({self.format_time(valid_splits[i-1])})"
                            )
                
                # Check finish time relationship
                if valid_splits and entry['finish'] is not None and entry['finish'] > 0:
                    if entry['finish'] <= valid_splits[-1]:
                        issues.append(
                            f"Invalid finish time for Bib {entry['bib']}: "
                            f"Finish ({self.format_time(entry['finish'])}) <= "
                            f"Last split ({self.format_time(valid_splits[-1])})"
                        )
        
        return issues
    
    def calculate_athlete_metrics(self, athlete_data):
        """
        Calculates comprehensive metrics for an athlete's performance.
        Handles both acceleration and regular splits.
        """
        metrics = {
            'splits': [[] for _ in range(self.num_splits)],
            'split_improvements': [0] * self.num_splits,
            'finish_times': [],
            'finish_improvements': 0,
            'consistency_scores': {'splits': [], 'finish': None},
            'progression_rate': {'splits': [], 'finish': None},
            'completion_rate': 0,
            'best_times': {'splits': [None] * self.num_splits, 'finish': None},
            'average_times': {'splits': [None] * self.num_splits, 'finish': None}
        }
        
        total_runs = len(athlete_data)
        completed_runs = 0
        
        # Process each run
        for run in athlete_data:
            if run['status'].upper() not in ['DNS', 'ERR']:
                # Process split times
                for i, split in enumerate(run['splits']):
                    if split is not None and split > 0:
                        metrics['splits'][i].append(split)
                        
                        # Update best times
                        if (metrics['best_times']['splits'][i] is None or 
                            split < metrics['best_times']['splits'][i]):
                            metrics['best_times']['splits'][i] = split
                            metrics['split_improvements'][i] += 1
                
                # Process finish times
                if run['status'].upper() not in ['DNF', 'DSQ'] and run['finish'] is not None:
                    metrics['finish_times'].append(run['finish'])
                    completed_runs += 1
                    
                    # Update best finish time
                    if (metrics['best_times']['finish'] is None or 
                        run['finish'] < metrics['best_times']['finish']):
                        metrics['best_times']['finish'] = run['finish']
                        metrics['finish_improvements'] += 1
        
        # Calculate completion rate
        metrics['completion_rate'] = (completed_runs / total_runs * 100) if total_runs > 0 else 0
        
        # Calculate averages and consistency scores
        for i in range(self.num_splits):
            if metrics['splits'][i]:
                metrics['average_times']['splits'][i] = sum(metrics['splits'][i]) / len(metrics['splits'][i])
                
                # Calculate consistency score (lower variance = higher consistency)
                variance = sum((t - metrics['average_times']['splits'][i]) ** 2 
                             for t in metrics['splits'][i]) / len(metrics['splits'][i])
                metrics['consistency_scores']['splits'].append(100 / (1 + variance))
                
                # Calculate progression rate
                if len(metrics['splits'][i]) >= 2:
                    progression = (metrics['splits'][i][0] - metrics['splits'][i][-1]) / len(metrics['splits'][i])
                    metrics['progression_rate']['splits'].append(progression)
        
        # Calculate finish time metrics
        if metrics['finish_times']:
            metrics['average_times']['finish'] = sum(metrics['finish_times']) / len(metrics['finish_times'])
            
            # Finish time consistency
            variance = sum((t - metrics['average_times']['finish']) ** 2 
                         for t in metrics['finish_times']) / len(metrics['finish_times'])
            metrics['consistency_scores']['finish'] = 100 / (1 + variance)
            
            # Finish time progression
            if len(metrics['finish_times']) >= 2:
                progression = (metrics['finish_times'][0] - metrics['finish_times'][-1]) / len(metrics['finish_times'])
                metrics['progression_rate']['finish'] = progression
        
        return metrics

    def analyze_section_times(self, run_data):
        """
        Analyzes times between splits to identify patterns and potential issues.
        """
        section_analysis = []
        
        for entry in run_data:
            if entry['status'].upper() not in ['DNS', 'ERR', 'DNF', 'DSQ']:
                valid_splits = [s for s in entry['splits'] if s is not None and s > 0]
                
                if len(valid_splits) >= 2:
                    section_times = []
                    for i in range(1, len(valid_splits)):
                        section_time = valid_splits[i] - valid_splits[i-1]
                        if section_time > 0:  # Only include valid section times
                            section_times.append({
                                'bib': entry['bib'],
                                'section': f"S{i}-S{i+1}",
                                'time': section_time
                            })
                    
                    # Add final section if finish time exists
                    if entry['finish'] is not None and entry['finish'] > valid_splits[-1]:
                        final_section = entry['finish'] - valid_splits[-1]
                        section_times.append({
                            'bib': entry['bib'],
                            'section': f"S{len(valid_splits)}-F",
                            'time': final_section
                        })
                    
                    section_analysis.append({
                        'bib': entry['bib'],
                        'sections': section_times
                    })
        
        return self.calculate_section_statistics(section_analysis)

    def calculate_section_statistics(self, section_analysis):
        """
        Calculates statistics for each section between splits.
        """
        section_stats = {}
        
        # Group times by section
        for entry in section_analysis:
            for section in entry['sections']:
                section_name = section['section']
                if section_name not in section_stats:
                    section_stats[section_name] = {
                        'times': [],
                        'best_time': None,
                        'average': None,
                        'std_dev': None,
                        'range': None
                    }
                section_stats[section_name]['times'].append({
                    'bib': section['bib'],
                    'time': section['time']
                })
        
        # Calculate statistics for each section
        for section_name, stats in section_stats.items():
            times = [t['time'] for t in stats['times']]
            if times:
                stats['best_time'] = min(times)
                stats['average'] = sum(times) / len(times)
                variance = sum((t - stats['average']) ** 2 for t in times) / len(times)
                stats['std_dev'] = variance ** 0.5
                stats['range'] = max(times) - min(times)
                
                # Mark suspicious times (beyond 2 standard deviations)
                for time_entry in stats['times']:
                    time = time_entry['time']
                    if abs(time - stats['average']) > 2 * stats['std_dev']:
                        time_entry['suspicious'] = True
        
        return section_stats

    def identify_performance_patterns(self, athlete_data):
        """
        Identifies patterns in athlete performance across runs.
        """
        patterns = {
            'split_patterns': [[] for _ in range(self.num_splits)],
            'finish_pattern': [],
            'fatigue_indicators': [],
            'improvement_trends': [],
            'consistency_issues': []
        }
        
        # Analyze each split separately
        for i in range(self.num_splits):
            split_times = []
            for run in athlete_data:
                if (run['status'].upper() not in ['DNS', 'ERR'] and 
                    len(run['splits']) > i and run['splits'][i] is not None):
                    split_times.append(run['splits'][i])
            
            if len(split_times) >= 3:  # Need at least 3 points to identify a pattern
                patterns['split_patterns'][i] = self.analyze_time_sequence(split_times)
        
        # Analyze finish times
        finish_times = []
        for run in athlete_data:
            if (run['status'].upper() not in ['DNS', 'ERR', 'DNF', 'DSQ'] and 
                run['finish'] is not None):
                finish_times.append(run['finish'])
        
        if len(finish_times) >= 3:
            patterns['finish_pattern'] = self.analyze_time_sequence(finish_times)
        
        return patterns

    def analyze_time_sequence(self, times):
        """
        Analyzes a sequence of times to identify trends and patterns.
        """
        if len(times) < 3:
            return {'trend': 'insufficient_data'}
            
        differences = [times[i] - times[i-1] for i in range(1, len(times))]
        
        # Calculate trend characteristics
        improving = all(d <= 0 for d in differences)
        worsening = all(d >= 0 for d in differences)
        
        # Calculate consistency
        avg_diff = sum(abs(d) for d in differences) / len(differences)
        
        if improving:
            return {
                'trend': 'improving',
                'avg_improvement': -sum(differences) / len(differences),
                'consistency': avg_diff
            }
        elif worsening:
            return {
                'trend': 'worsening',
                'avg_decline': sum(differences) / len(differences),
                'consistency': avg_diff
            }
        else:
            return {
                'trend': 'variable',
                'variation': avg_diff,
                'range': max(times) - min(times)
            }
        
    def save_excel_summary(self, wb, data, athletes_metrics):
        """Creates a summary sheet with comprehensive analysis."""
        summary_sheet = wb.create_sheet(title="Session Summary")
        current_row = 1

        # Session Overview
        summary_sheet.cell(row=current_row, column=1, value="Session Overview").font = Font(bold=True, size=14)
        current_row += 2

        overview_data = [
            ("Event", self.event_var.get()),
            ("Date", self.date_var.get()),
            ("Hill", self.hill_var.get()),
            ("Team", f"{self.team_names[self.current_team]} - {self.get_season(self.date_var.get())}"),
            ("Session #", self.session_var.get()),
            ("Conditions", f"Snow: {self.snow_condition_var.get()}, Weather: {self.sky_condition_var.get()}, "
                         f"Wind: {self.wind_condition_var.get()}")
        ]

        for label, value in overview_data:
            summary_sheet.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            summary_sheet.cell(row=current_row, column=2, value=value)
            current_row += 1

        current_row += 2

        # Performance Highlights
        self.add_performance_highlights(summary_sheet, athletes_metrics, current_row)
        current_row += len(athletes_metrics) + 3

        # Training Progression Analysis
        current_row = self.add_progression_analysis(summary_sheet, data, current_row)

        # Format the sheet
        for col in ['A', 'B', 'C', 'D', 'E']:
            summary_sheet.column_dimensions[col].width = 15

        return summary_sheet

    def add_performance_highlights(self, ws, athletes_metrics, start_row):
        """Adds key performance metrics for each athlete."""
        ws.cell(row=start_row, column=1, value="Performance Highlights").font = Font(bold=True, size=12)
        start_row += 2

        # Headers
        headers = ['Athlete', 'Best Performance', 'Improvement Rate', 'Consistency', 'Notes']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
        start_row += 1

        # Add metrics for each athlete
        for athlete_name, metrics in athletes_metrics.items():
            row_data = [
                athlete_name,
                f"{self.format_time(metrics['best_times']['finish'])} "
                f"(Run {self.get_best_run_number(metrics)})",
                f"{self.calculate_improvement_rate(metrics):.1f}%",
                f"{metrics['consistency_scores'].get('finish', 0):.1f}%",
                self.generate_performance_notes(metrics)
            ]

            for col, value in enumerate(row_data, start=1):
                ws.cell(row=start_row, column=col, value=value)
            start_row += 1

        return start_row

    def add_progression_analysis(self, ws, data, start_row):
        """Adds detailed progression analysis."""
        ws.cell(row=start_row, column=1, value="Training Progression Analysis").font = Font(bold=True, size=12)
        start_row += 2

        run_numbers = sorted(data.keys(), key=int)
        progression_data = self.analyze_training_progression(data)

        # Add run-by-run analysis
        headers = ['Run', 'Completion Rate', 'Avg Time', 'Time Range', 'Notable Events']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
        start_row += 1

        for run_num in run_numbers:
            run_stats = progression_data[run_num]
            row_data = [
                f"Run {run_num}",
                f"{run_stats['completion_rate']:.1f}%",
                self.format_time(run_stats['avg_time']),
                self.format_time(run_stats['time_range']),
                run_stats['notes']
            ]

            for col, value in enumerate(row_data, start=1):
                ws.cell(row=start_row, column=col, value=value)
            start_row += 1

        return start_row

    def analyze_training_progression(self, data):
        """Analyzes progression throughout the training session."""
        progression = {}
        
        for run_num, run_data in data.items():
            valid_times = []
            total_athletes = 0
            completed_runs = 0
            notable_events = []

            for entry in run_data:
                total_athletes += 1
                if entry['status'].upper() not in ['DNS', 'ERR', 'DNF', 'DSQ']:
                    if entry['finish'] is not None and entry['finish'] > 0:
                        valid_times.append(entry['finish'])
                        completed_runs += 1

                # Check for notable events
                if entry['error_details']:
                    notable_events.extend(entry['error_details'])

            # Calculate statistics
            completion_rate = (completed_runs / total_athletes * 100) if total_athletes > 0 else 0
            avg_time = sum(valid_times) / len(valid_times) if valid_times else None
            time_range = max(valid_times) - min(valid_times) if len(valid_times) >= 2 else None

            progression[run_num] = {
                'completion_rate': completion_rate,
                'avg_time': avg_time,
                'time_range': time_range,
                'notes': '; '.join(notable_events) if notable_events else 'No significant issues'
            }

        return progression

    def export_error_log(self, output_path, data):
        """Exports detailed error and validation log."""
        try:
            with open(output_path.replace('.xlsx', '_validation.log'), 'w') as f:
                f.write(f"Validation Log - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Event: {self.event_var.get()}\n")
                f.write(f"Session: {self.session_var.get()}\n")
                f.write("-" * 80 + "\n\n")

                for run_num, run_data in sorted(data.items()):
                    f.write(f"Run {run_num}:\n")
                    f.write("-" * 40 + "\n")
                    
                    for entry in run_data:
                        if entry['error_details']:
                            f.write(f"Bib {entry['bib']}:\n")
                            for error in entry['error_details']:
                                f.write(f"  • {error}\n")
                    f.write("\n")

                # Add statistical anomalies
                f.write("\nStatistical Analysis:\n")
                f.write("-" * 40 + "\n")
                anomalies = self.find_statistical_anomalies(data)
                for anomaly in anomalies:
                    f.write(f"• {anomaly}\n")

        except Exception as e:
            print(f"Error writing validation log: {str(e)}")

    def find_statistical_anomalies(self, data):
        """Identifies statistical anomalies across the entire session."""
        anomalies = []
        
        # Analyze progression across runs
        run_averages = {}
        for run_num, run_data in data.items():
            valid_times = []
            for entry in run_data:
                if (entry['status'].upper() not in ['DNS', 'ERR', 'DNF', 'DSQ'] and 
                    entry['finish'] is not None and entry['finish'] > 0):
                    valid_times.append(entry['finish'])
            
            if valid_times:
                run_averages[run_num] = sum(valid_times) / len(valid_times)

        # Check for unusual changes between runs
        if len(run_averages) >= 2:
            prev_avg = None
            for run_num in sorted(run_averages.keys(), key=int):
                if prev_avg is not None:
                    change = run_averages[run_num] - prev_avg
                    if abs(change) > 2.0:  # More than 2 seconds change
                        anomalies.append(
                            f"Large average time change between runs {int(run_num)-1} "
                            f"and {run_num}: {self.format_time(change, True)}"
                        )
                prev_avg = run_averages[run_num]

        # Check for unusual patterns in splits
        for run_num, run_data in data.items():
            split_stats = self.calculate_split_statistics(run_data)
            if split_stats and split_stats['coefficient_of_variation'] > 15:
                anomalies.append(
                    f"High variation in split times for Run {run_num} "
                    f"(CV: {split_stats['coefficient_of_variation']:.1f}%)"
                )

        return anomalies

    def get_best_run_number(self, metrics):
        """Determines the run number with the best performance."""
        if not metrics['finish_times']:
            return "N/A"
        
        best_time = min(metrics['finish_times'])
        run_index = metrics['finish_times'].index(best_time)
        return run_index + 1

    def calculate_improvement_rate(self, metrics):
        """Calculates overall improvement rate as percentage."""
        if not metrics['finish_times'] or len(metrics['finish_times']) < 2:
            return 0.0
        
        first_time = metrics['finish_times'][0]
        best_time = min(metrics['finish_times'])
        return ((first_time - best_time) / first_time) * 100

    def generate_performance_notes(self, metrics):
        """Generates concise performance notes based on metrics."""
        notes = []
        
        # Check improvement trend
        if metrics['progression_rate']['finish'] is not None:
            if metrics['progression_rate']['finish'] < 0:
                notes.append("Consistent improvement")
            elif metrics['progression_rate']['finish'] > 0:
                notes.append("Performance declining")
        
        # Check consistency
        if metrics['consistency_scores']['finish'] is not None:
            if metrics['consistency_scores']['finish'] > 90:
                notes.append("Very consistent")
            elif metrics['consistency_scores']['finish'] < 70:
                notes.append("Inconsistent")
        
        return "; ".join(notes) if notes else "No significant patterns"
    
    def build_filter_controls(self, parent_frame):
        """Creates filtering and view controls for timing data."""
        filter_frame = tk.LabelFrame(parent_frame, text="Data Filters", padx=10, pady=5)
        filter_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)

        # Time threshold filter
        tk.Label(filter_frame, text="Time Threshold:").grid(row=0, column=0, padx=5)
        self.threshold_var = tk.StringVar(value="0.0")
        threshold_entry = tk.Entry(filter_frame, textvariable=self.threshold_var, width=8)
        threshold_entry.grid(row=0, column=1, padx=5)
        
        # Status filter
        tk.Label(filter_frame, text="Status Filter:").grid(row=0, column=2, padx=5)
        self.status_var = tk.StringVar(value="ALL")
        status_combo = Combobox(filter_frame, textvariable=self.status_var, 
                              values=["ALL", "VALID", "DNF", "DSQ", "DNS", "ERR"])
        status_combo.grid(row=0, column=3, padx=5)
        
        # Apply filters button
        tk.Button(filter_frame, text="Apply Filters", 
                 command=self.apply_filters).grid(row=0, column=4, padx=5)

    def build_validation_controls(self, parent_frame):
        """Creates validation controls and indicators."""
        validation_frame = tk.LabelFrame(parent_frame, text="Validation", padx=10, pady=5)
        validation_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)

        # Validation level selector
        tk.Label(validation_frame, text="Validation Level:").grid(row=0, column=0, padx=5)
        self.validation_level = tk.StringVar(value="NORMAL")
        validation_combo = Combobox(validation_frame, textvariable=self.validation_level,
                                  values=["STRICT", "NORMAL", "RELAXED"])
        validation_combo.grid(row=0, column=1, padx=5)
        
        # Auto-validation toggle
        self.auto_validate = tk.BooleanVar(value=True)
        tk.Checkbutton(validation_frame, text="Auto-validate", 
                      variable=self.auto_validate).grid(row=0, column=2, padx=5)
        
        # Manual validation button
        tk.Button(validation_frame, text="Validate Data",
                 command=self.validate_current_data).grid(row=0, column=3, padx=5)

    def create_status_indicators(self, parent_frame):
        """Creates status indicators for data validation and processing."""
        status_frame = tk.LabelFrame(parent_frame, text="Status", padx=10, pady=5)
        status_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)

        # Data status indicator
        self.data_status_var = tk.StringVar(value="No Data Loaded")
        self.data_status_label = tk.Label(status_frame, textvariable=self.data_status_var)
        self.data_status_label.grid(row=0, column=0, padx=5)
        
        # Validation status indicator
        self.validation_status_var = tk.StringVar(value="Not Validated")
        self.validation_status_label = tk.Label(status_frame, textvariable=self.validation_status_var)
        self.validation_status_label.grid(row=0, column=1, padx=5)
        
        # Error counter
        self.error_count_var = tk.StringVar(value="Errors: 0")
        self.error_count_label = tk.Label(status_frame, textvariable=self.error_count_var)
        self.error_count_label.grid(row=0, column=2, padx=5)

    def create_progress_indicators(self):
        """Creates progress indicators for long operations."""
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.root, 
            variable=self.progress_var,
            maximum=100,
            mode='determinate'
        )
        self.progress_bar.grid(row=998, column=0, columnspan=3, sticky="ew", padx=5, pady=5)
        self.progress_bar.grid_remove()  # Hidden by default
        
        self.progress_label_var = tk.StringVar()
        self.progress_label = tk.Label(self.root, textvariable=self.progress_label_var)
        self.progress_label.grid(row=997, column=0, columnspan=3)
        self.progress_label.grid_remove()  # Hidden by default

    def show_error_details(self, errors):
        """Shows detailed error information in a separate window."""
        error_window = tk.Toplevel(self.root)
        error_window.title("Validation Details")
        error_window.geometry("600x400")

        # Create scrolled text widget
        text_widget = scrolledtext.ScrolledText(error_window, wrap=tk.WORD, width=70, height=20)
        text_widget.pack(expand=True, fill='both', padx=10, pady=10)

        # Add error information
        text_widget.insert(tk.END, "Validation Results\n")
        text_widget.insert(tk.END, "=================\n\n")
        
        for category, category_errors in errors.items():
            text_widget.insert(tk.END, f"{category}:\n")
            text_widget.insert(tk.END, "-" * len(category) + "\n")
            for error in category_errors:
                text_widget.insert(tk.END, f"• {error}\n")
            text_widget.insert(tk.END, "\n")

        text_widget.configure(state='disabled')  # Make read-only

    def update_validation_status(self, status, error_count=0):
        """Updates the validation status indicators."""
        self.validation_status_var.set(status)
        self.error_count_var.set(f"Errors: {error_count}")
        
        # Update label colors based on status
        if status == "Valid":
            self.validation_status_label.configure(fg="green")
        elif status == "Invalid":
            self.validation_status_label.configure(fg="red")
        else:
            self.validation_status_label.configure(fg="black")

    def show_progress(self, show=True, label=None):
        """Shows or hides the progress indicators."""
        if show:
            self.progress_bar.grid()
            if label:
                self.progress_label_var.set(label)
                self.progress_label.grid()
            self.progress_var.set(0)
            self.root.update_idletasks()
        else:
            self.progress_bar.grid_remove()
            self.progress_label.grid_remove()
            self.root.update_idletasks()

    def update_progress(self, value, label=None):
        """Updates the progress bar and label."""
        self.progress_var.set(value)
        if label:
            self.progress_label_var.set(label)
        self.root.update_idletasks()

    def apply_filters(self):
        """Applies the current filters to the timing data."""
        try:
            threshold = float(self.threshold_var.get())
            status_filter = self.status_var.get()
            
            if self.timing_data:
                filtered_data = self.filter_timing_data(
                    self.timing_data,
                    threshold,
                    status_filter
                )
                self.update_display(filtered_data)
                
        except ValueError:
            messagebox.showerror("Error", "Invalid threshold value")

    def filter_timing_data(self, data, threshold, status_filter):
        """Filters timing data based on threshold and status."""
        filtered = {}
        
        for run_num, run_data in data.items():
            filtered_run = []
            for entry in run_data:
                # Apply status filter
                if status_filter != "ALL" and entry['status'] != status_filter:
                    continue
                    
                # Apply threshold filter
                if threshold > 0:
                    if entry['finish'] is not None and entry['finish'] < threshold:
                        continue
                    
                filtered_run.append(entry)
            
            if filtered_run:
                filtered[run_num] = filtered_run
                
        return filtered

    def validate_current_data(self):
        """Performs validation on the current data."""
        if not self.timing_data:
            messagebox.showwarning("Warning", "No data to validate")
            return
            
        self.show_progress(True, "Validating data...")
        
        # Collect all validation results
        errors = {
            "Time Validation": [],
            "Progression Errors": [],
            "Statistical Anomalies": [],
            "Data Consistency": []
        }
        
        total_steps = len(self.timing_data) * 2  # Multiple validation passes
        current_step = 0
        
        # Validate each run
        for run_num, run_data in self.timing_data.items():
            # Time validation
            validation_errors = self.validate_run_consistency(run_data)
            if validation_errors:
                errors["Time Validation"].extend(
                    [f"Run {run_num}: {error}" for error in validation_errors]
                )
            
            current_step += 1
            self.update_progress(
                (current_step / total_steps) * 100,
                f"Validating run {run_num}..."
            )
            
            # Statistical validation
            anomalies = self.analyze_split_relationships(run_data)
            if anomalies:
                errors["Statistical Anomalies"].extend(
                    [f"Run {run_num}: {anomaly}" for anomaly in anomalies]
                )
            
            current_step += 1
            self.update_progress(
                (current_step / total_steps) * 100,
                f"Analyzing run {run_num}..."
            )
        
        # Update status indicators
        total_errors = sum(len(err_list) for err_list in errors.values())
        self.update_validation_status(
            "Valid" if total_errors == 0 else "Invalid",
            total_errors
        )
        
        # Show error details if any
        if total_errors > 0:
            self.show_error_details(errors)
        
        self.show_progress(False)

    def create_menu_system(self):
        """Creates the main menu system."""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # File Menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Open CSV...", command=self.select_file)
        file_menu.add_command(label="Save Excel...", command=self.reformat_file)
        file_menu.add_separator()
        file_menu.add_command(label="Save Current State", command=self.save_current_state)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)

        # Edit Menu
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Settings...", command=self.open_settings)
        edit_menu.add_separator()
        edit_menu.add_command(label="Clear Recent Names", command=self.clear_recent_names)
        edit_menu.add_command(label="Clear Recent Hills", command=self.clear_recent_hills)

        # View Menu
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        view_menu.add_checkbutton(label="Show Error Details", 
                                variable=self.show_error_details_var,
                                command=self.toggle_error_details)
        view_menu.add_checkbutton(label="Show Statistics", 
                                variable=self.show_statistics_var,
                                command=self.toggle_statistics)

        # Tools Menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="Validate Data", command=self.validate_current_data)
        tools_menu.add_command(label="Export Error Log", command=self.export_error_log_dialog)
        tools_menu.add_separator()
        tools_menu.add_command(label="Analysis Options...", command=self.show_analysis_options)

        # Help Menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="User Guide", command=self.show_user_guide)
        help_menu.add_command(label="About", command=self.show_about)

    def create_context_menus(self):
        """Creates context menus for various widgets."""
        # Athlete listbox context menu
        self.athlete_context_menu = tk.Menu(self.root, tearoff=0)
        self.athlete_context_menu.add_command(label="Edit Athlete", 
                                            command=self.edit_selected_athlete)
        self.athlete_context_menu.add_command(label="Remove Athlete", 
                                            command=self.remove_selected_athlete)
        self.athlete_context_menu.add_separator()
        self.athlete_context_menu.add_command(label="View Statistics", 
                                            command=self.view_athlete_statistics)
        
        self.athlete_listbox.bind("<Button-3>", self.show_athlete_context_menu)

        # Guest listbox context menu
        self.guest_context_menu = tk.Menu(self.root, tearoff=0)
        self.guest_context_menu.add_command(label="Convert to Athlete", 
                                          command=self.convert_guest_to_athlete)
        self.guest_context_menu.add_command(label="Remove Guest", 
                                          command=self.remove_selected_guest)
        
        self.guest_listbox.bind("<Button-3>", self.show_guest_context_menu)

    def setup_keyboard_shortcuts(self):
        """Sets up keyboard shortcuts for common actions."""
        self.root.bind("<Control-o>", lambda e: self.select_file())
        self.root.bind("<Control-s>", lambda e: self.save_current_state())
        self.root.bind("<Control-r>", lambda e: self.reformat_file())
        self.root.bind("<Control-q>", lambda e: self.root.quit())
        
        # Additional shortcuts for athlete management
        self.athlete_name_entry.bind("<Control-Return>", lambda e: self.add_athlete())
        self.athlete_name_entry.bind("<Shift-Return>", lambda e: self.add_guest())
        
        # Navigation shortcuts
        self.root.bind("<Control-Tab>", self.cycle_teams_forward)
        self.root.bind("<Control-Shift-Tab>", self.cycle_teams_backward)

    def create_toolbar(self):
        """Creates a toolbar with common actions."""
        toolbar_frame = tk.Frame(self.root, bd=1, relief=tk.RAISED)
        toolbar_frame.grid(row=0, column=0, columnspan=3, sticky="ew", padx=5, pady=2)

        # Create toolbar buttons
        toolbar_buttons = [
            ("Open", self.select_file, "Open CSV file"),
            ("Save", self.save_current_state, "Save current state"),
            ("Export", self.reformat_file, "Export to Excel"),
            (None, None, None),  # Separator
            ("Validate", self.validate_current_data, "Validate timing data"),
            ("Settings", self.open_settings, "Open settings"),
            (None, None, None),  # Separator
            ("Help", self.show_user_guide, "Show user guide")
        ]

        for btn_text, command, tooltip in toolbar_buttons:
            if btn_text is None:
                # Add separator
                tk.Frame(toolbar_frame, width=2, bd=1, relief=tk.SUNKEN).pack(
                    side=tk.LEFT, padx=2, pady=2)
            else:
                btn = tk.Button(toolbar_frame, text=btn_text, command=command)
                btn.pack(side=tk.LEFT, padx=2, pady=2)
                self.create_tooltip(btn, tooltip)

    def create_status_bar(self):
        """Creates a status bar at the bottom of the window."""
        status_frame = tk.Frame(self.root, bd=1, relief=tk.SUNKEN)
        status_frame.grid(row=999, column=0, columnspan=3, sticky="ew")

        # Status message (left)
        self.status_message = tk.StringVar(value="Ready")
        status_label = tk.Label(status_frame, textvariable=self.status_message, 
                              anchor=tk.W, padx=5)
        status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Version (right)
        version_label = tk.Label(status_frame, text=f"v{self.VERSION}", padx=5)
        version_label.pack(side=tk.RIGHT)

    def create_tooltip(self, widget, text):
        """Creates a tooltip for a widget."""
        def show_tooltip(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            
            label = tk.Label(tooltip, text=text, background="lightyellow", 
                           relief=tk.SOLID, borderwidth=1)
            label.pack()
            
            def hide_tooltip():
                tooltip.destroy()
            
            tooltip.bind("<Leave>", lambda e: hide_tooltip())
            widget.bind("<Leave>", lambda e: hide_tooltip())

        widget.bind("<Enter>", show_tooltip)

    # Context Menu Event Handlers
    def show_athlete_context_menu(self, event):
        """Shows context menu for athlete listbox."""
        try:
            self.athlete_listbox.selection_clear(0, tk.END)
            self.athlete_listbox.selection_set(self.athlete_listbox.nearest(event.y))
            self.athlete_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.athlete_context_menu.grab_release()

    def show_guest_context_menu(self, event):
        """Shows context menu for guest listbox."""
        try:
            self.guest_listbox.selection_clear(0, tk.END)
            self.guest_listbox.selection_set(self.guest_listbox.nearest(event.y))
            self.guest_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.guest_context_menu.grab_release()

    # Team Navigation
    def cycle_teams_forward(self, event=None):
        """Cycles through teams forward."""
        teams = ["SQAH", "SQAF", "OTHER"]
        current_index = teams.index(self.current_team)
        next_index = (current_index + 1) % len(teams)
        self.set_team(teams[next_index])

    def cycle_teams_backward(self, event=None):
        """Cycles through teams backward."""
        teams = ["SQAH", "SQAF", "OTHER"]
        current_index = teams.index(self.current_team)
        prev_index = (current_index - 1) % len(teams)
        self.set_team(teams[prev_index])

    def show_user_guide(self):
        """Shows the user guide window."""
        guide_window = tk.Toplevel(self.root)
        guide_window.title("User Guide")
        guide_window.geometry("600x400")

        text_widget = scrolledtext.ScrolledText(guide_window, wrap=tk.WORD)
        text_widget.pack(expand=True, fill='both', padx=10, pady=10)

        guide_text = """
Timing System Data Formatter - User Guide

Keyboard Shortcuts:
• Ctrl+O: Open CSV file
• Ctrl+S: Save current state
• Ctrl+R: Reformat and export to Excel
• Ctrl+Tab: Cycle through teams forward
• Ctrl+Shift+Tab: Cycle through teams backward
• Ctrl+Enter: Add athlete
• Shift+Enter: Add guest

Managing Athletes:
1. Select team using the team buttons
2. Enter athlete name and bib number
3. Click "Add Athlete" or use Ctrl+Enter
4. Right-click on athlete for more options

Validation:
• Data is automatically validated on import
• Use Tools > Validate Data for manual validation
• Error details are shown in the validation window

Export:
1. Select CSV file
2. Fill in training parameters
3. Click "Reformat Selected File" or use Ctrl+R
4. Choose export location

Tips:
• Use auto-complete for faster name entry
• Right-click athletes for quick actions
• Check validation results before export
• Save current state regularly
        """

        text_widget.insert(tk.END, guide_text)
        text_widget.configure(state='disabled')

    def show_about(self):
        """Shows the about dialog."""
        about_text = f"""
Timing System Data Formatter v{self.VERSION}

Created by: {self.AUTHOR}
Last Updated: {datetime.now().strftime('%B %Y')}

A specialized tool for formatting and analyzing
timing data from skiing training sessions.

© 2024 All Rights Reserved
        """
        messagebox.showinfo("About", about_text)


    def edit_selected_athlete(self):
        """Opens dialog to edit selected athlete's details."""
        selection = self.athlete_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an athlete to edit")
            return

        athlete_index = selection[0]
        athlete = self.athletes[self.current_team][athlete_index]

        # Create edit dialog
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Athlete")
        edit_window.geometry("400x300")
        edit_window.transient(self.root)
        edit_window.grab_set()

        # Create form fields
        tk.Label(edit_window, text="Name:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        name_var = tk.StringVar(value=athlete['name'])
        name_entry = tk.Entry(edit_window, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(edit_window, text="Bib Number:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        bib_var = tk.StringVar(value=athlete['bib'])
        bib_entry = tk.Entry(edit_window, textvariable=bib_var, width=10)
        bib_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # Add additional fields for enhanced athlete management
        tk.Label(edit_window, text="Category:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        category_var = tk.StringVar(value=athlete.get('category', ''))
        category_combo = Combobox(edit_window, textvariable=category_var,
                                values=['U14', 'U16', 'U18', 'U21', 'Senior'])
        category_combo.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(edit_window, text="Notes:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        notes_text = tk.Text(edit_window, width=30, height=4)
        notes_text.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        notes_text.insert('1.0', athlete.get('notes', ''))

        def save_changes():
            # Validate bib number
            try:
                new_bib = int(bib_var.get())
            except ValueError:
                messagebox.showerror("Error", "Invalid bib number")
                return

            # Check for bib number conflicts
            if str(new_bib) != athlete['bib']:  # Only check if bib changed
                if self.check_bib_conflict(new_bib):
                    messagebox.showerror("Error", "Bib number already in use")
                    return

            # Update athlete data
            athlete['name'] = name_var.get()
            athlete['bib'] = str(new_bib)
            athlete['category'] = category_var.get()
            athlete['notes'] = notes_text.get('1.0', 'end-1c')

            # Save changes
            self.save_athletes_to_json()
            self.update_athlete_listbox()
            edit_window.destroy()

        # Add buttons
        button_frame = tk.Frame(edit_window)
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)

        tk.Button(button_frame, text="Save", command=save_changes).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Cancel", command=edit_window.destroy).pack(side=tk.LEFT, padx=5)

    def convert_guest_to_athlete(self):
        """Converts selected guest to permanent athlete."""
        selection = self.guest_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a guest to convert")
            return

        guest_index = selection[0]
        guest = self.temp_guests[guest_index]

        # Create conversion dialog
        convert_window = tk.Toplevel(self.root)
        convert_window.title("Convert Guest to Athlete")
        convert_window.geometry("400x250")
        convert_window.transient(self.root)
        convert_window.grab_set()

        # Form fields
        tk.Label(convert_window, text="Name:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        name_var = tk.StringVar(value=guest['name'])
        name_entry = tk.Entry(convert_window, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(convert_window, text="Bib Number:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        bib_var = tk.StringVar(value=guest['bib'])
        bib_entry = tk.Entry(convert_window, textvariable=bib_var, width=10)
        bib_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(convert_window, text="Team:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        team_var = tk.StringVar(value=self.current_team)
        team_combo = Combobox(convert_window, textvariable=team_var,
                            values=["SQAH", "SQAF", "OTHER"])
        team_combo.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        def perform_conversion():
            # Validate data
            try:
                new_bib = int(bib_var.get())
            except ValueError:
                messagebox.showerror("Error", "Invalid bib number")
                return

            if self.check_bib_conflict(new_bib, exclude_guest=guest['bib']):
                messagebox.showerror("Error", "Bib number already in use")
                return

            # Create new athlete
            new_athlete = {
                'name': name_var.get(),
                'bib': str(new_bib),
                'category': '',
                'notes': f'Converted from guest on {datetime.now().strftime("%Y-%m-%d")}'
            }

            # Add to selected team
            selected_team = team_var.get()
            self.athletes[selected_team].append(new_athlete)

            # Remove from guests
            self.temp_guests.pop(guest_index)

            # Update displays
            self.save_athletes_to_json()
            self.update_athlete_listbox()
            self.update_guest_listbox()
            convert_window.destroy()

        # Add buttons
        button_frame = tk.Frame(convert_window)
        button_frame.grid(row=3, column=0, columnspan=2, pady=20)

        tk.Button(button_frame, text="Convert", command=perform_conversion).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Cancel", command=convert_window.destroy).pack(side=tk.LEFT, padx=5)

    def view_athlete_statistics(self):
        """Shows detailed statistics for selected athlete."""
        selection = self.athlete_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an athlete to view statistics")
            return

        athlete = self.athletes[self.current_team][selection[0]]
        
        # Create statistics window
        stats_window = tk.Toplevel(self.root)
        stats_window.title(f"Statistics - {athlete['name']}")
        stats_window.geometry("600x400")
        stats_window.transient(self.root)

        # Create notebook for different stat views
        notebook = ttk.Notebook(stats_window)
        notebook.pack(expand=True, fill='both', padx=5, pady=5)

        # Summary tab
        summary_frame = ttk.Frame(notebook)
        notebook.add(summary_frame, text='Summary')
        self.create_summary_tab(summary_frame, athlete)

        # Detailed Stats tab
        details_frame = ttk.Frame(notebook)
        notebook.add(details_frame, text='Detailed Stats')
        self.create_details_tab(details_frame, athlete)

        # Performance Graph tab
        graph_frame = ttk.Frame(notebook)
        notebook.add(graph_frame, text='Performance Graph')
        self.create_graph_tab(graph_frame, athlete)

    def create_summary_tab(self, parent, athlete):
        """Creates the summary statistics tab."""
        if not hasattr(self, 'timing_data') or not self.timing_data:
            tk.Label(parent, text="No timing data available").pack(pady=20)
            return

        # Collect athlete's data
        athlete_stats = self.calculate_athlete_metrics(self.get_athlete_runs(athlete['bib']))

        # Create summary table
        summary_frame = tk.LabelFrame(parent, text="Performance Summary", padx=10, pady=5)
        summary_frame.pack(fill='x', padx=5, pady=5)

        row = 0
        for label, value in [
            ("Total Runs", len(athlete_stats['finish_times'])),
            ("Best Time", self.format_time(athlete_stats['best_times']['finish'])),
            ("Average Time", self.format_time(athlete_stats['average_times']['finish'])),
            ("Completion Rate", f"{athlete_stats['completion_rate']:.1f}%"),
            ("Consistency Score", f"{athlete_stats['consistency_scores']['finish']:.1f}%")
        ]:
            tk.Label(summary_frame, text=label + ":").grid(row=row, column=0, sticky='e', padx=5, pady=2)
            tk.Label(summary_frame, text=str(value)).grid(row=row, column=1, sticky='w', padx=5, pady=2)
            row += 1

    def create_details_tab(self, parent, athlete):
        """Creates the detailed statistics tab."""
        if not hasattr(self, 'timing_data') or not self.timing_data:
            tk.Label(parent, text="No timing data available").pack(pady=20)
            return

        # Create scrolled frame
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Add detailed statistics
        self.add_split_statistics(scrollable_frame, athlete)
        self.add_progression_statistics(scrollable_frame, athlete)
        self.add_comparison_statistics(scrollable_frame, athlete)

        # Pack the scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

    def create_graph_tab(self, parent, athlete):
        """Creates the performance graph tab."""
        if not hasattr(self, 'timing_data') or not self.timing_data:
            tk.Label(parent, text="No timing data available").pack(pady=20)
            return

        # Create frame for graph controls
        control_frame = tk.Frame(parent)
        control_frame.pack(fill='x', padx=5, pady=5)

        # Graph type selector
        tk.Label(control_frame, text="Graph Type:").pack(side=tk.LEFT, padx=5)
        graph_type = tk.StringVar(value="splits")
        Combobox(control_frame, textvariable=graph_type,
                values=["splits", "finish", "combined"],
                width=15).pack(side=tk.LEFT, padx=5)

        # Update button
        tk.Button(control_frame, text="Update Graph",
                 command=lambda: self.update_performance_graph(graph_frame, athlete, graph_type.get())
                 ).pack(side=tk.LEFT, padx=5)

        # Create frame for the graph
        graph_frame = tk.Frame(parent)
        graph_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Initial graph
        self.update_performance_graph(graph_frame, athlete, "splits")

    def parse_timing_data(self, file_path):
        """
        Fixed version that handles both single and multiple split formats.
        """
        timing_data = {}
        header_found = False
        column_indices = {}
        split_columns = []
        
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                separator = '>'
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Look for header line
                    if not header_found and "Bib#" in line:
                        header_found = True
                        headers = line.split(separator)
                        
                        # Get column indices
                        for i, header in enumerate(headers):
                            header = header.lower().strip()
                            if "bib" in header:
                                column_indices['bib'] = i
                            elif "run" in header:
                                column_indices['run'] = i
                            elif "split" in header:
                                split_columns.append(i)  # Store all split columns
                            elif "finish time" in header:
                                column_indices['finish'] = i
                            elif "status" in header:
                                column_indices['status'] = i
                        continue
                    
                    if header_found:
                        try:
                            data = line.split(separator)
                            if len(data) <= 1:
                                continue
                            
                            # Get run number - handle both formats
                            run_number = data[column_indices['run']].strip()
                            if not run_number:
                                continue
                            
                            if run_number not in timing_data:
                                timing_data[run_number] = []
                            
                            # Process all split times
                            splits = []
                            for split_col in split_columns:
                                if split_col < len(data):
                                    split_time = self.validate_time(data[split_col].strip())
                                    splits.append(split_time)
                            
                            # Create entry
                            entry = {
                                'bib': data[column_indices['bib']].strip(),
                                'splits': splits,
                                'finish': self.validate_time(
                                    data[column_indices['finish']].strip() 
                                    if 'finish' in column_indices and 
                                    column_indices['finish'] < len(data) 
                                    else ''
                                ),
                                'status': data[column_indices['status']].strip() 
                                        if 'status' in column_indices and 
                                            column_indices['status'] < len(data) 
                                        else '',
                                'run': run_number
                            }
                            
                            timing_data[run_number].append(entry)
                            
                        except IndexError as e:
                            print(f"Error processing line: {line}")
                            print(f"Error details: {str(e)}")
                            continue
                
                return timing_data
                
        except Exception as e:
            messagebox.showerror("Error", f"Error parsing CSV file: {str(e)}")
            return None

    def validate_time(self, time_str):
        """
        Fixed time validation that handles both formats.
        """
        if not time_str or not isinstance(time_str, str):
            return None
            
        time_str = time_str.strip().upper()
        if time_str in ['DNF', 'DSQ', '', 'DNS', '0']:
            return None
            
        try:
            if ':' in time_str:
                # Handle MM:SS.sss format
                if '.' in time_str:
                    minutes, rest = time_str.split(':')
                    seconds, milliseconds = rest.split('.')
                    return float(minutes) * 60 + float(seconds) + float(milliseconds) / 1000
                # Handle MM:SS format
                minutes, seconds = time_str.split(':')
                return float(minutes) * 60 + float(seconds)
            # Handle plain seconds format
            return float(time_str)
        except (ValueError, TypeError):
            return None

        
    def create_formatted_excel(self, output_path):
        """Creates a formatted Excel file with comprehensive timing data analysis."""
        if not hasattr(self, 'timing_data') or not self.timing_data:
            messagebox.showerror("Error", "No timing data to export")
            return False

        try:
            wb = Workbook()
            main_sheet = wb.active
            main_sheet.title = "Timing Data"
            
            # Style Definitions
            styles = self.create_excel_styles()
            
            # Set column widths for main sheet
            column_widths = self.calculate_column_widths()
            for col, width in column_widths.items():
                main_sheet.column_dimensions[col].width = width
            
            # Write header section
            current_row = self.write_header_section(main_sheet, styles)
            
            # Process each run's data
            for run_number in sorted(self.timing_data.keys(), key=int):
                run_data = self.timing_data[run_number]
                current_row = self.write_run_data(main_sheet, run_data, current_row, styles)
                current_row += 2  # Add spacing between runs
            
            # Create additional sheets
            self.create_athlete_analysis_sheet(wb, styles)
            self.create_category_sheet(wb, styles)
            self.create_statistics_sheet(wb, styles)
            
            # Save the workbook
            wb.save(output_path)
            return True
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create Excel file: {str(e)}")
            print(f"Excel creation error: {str(e)}")  # For debugging
            return False

    def create_excel_styles(self):
        """Creates and returns a dictionary of Excel styles."""
        styles = {
            'title': Font(name='Arial', size=18, bold=True),
            'header': Font(name='Arial', size=11, bold=True),
            'normal': Font(name='Arial', size=11),
            'error': Font(name='Arial', size=11, color="FF0000"),
            
            'borders': {
                'thick': Border(
                    left=Side(style='thick'),
                    right=Side(style='thick'),
                    top=Side(style='thick'),
                    bottom=Side(style='thick')
                ),
                'thin': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            
            'alignments': {
                'center': Alignment(horizontal='center', vertical='center'),
                'left': Alignment(horizontal='left', vertical='center')
            },
            
            'fills': {
                'header': PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
                'category': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            }
        }
        return styles

    def calculate_column_widths(self):
        """Calculates column widths based on content type and split count."""
        base_columns = {
            'A': 6,   # Index
            'B': 18,  # Bib #
            'C': 20,  # Name/Category
        }
        
        # Calculate widths for split columns
        current_col = 'D'
        split_columns = {}
        
        # For each split: Time(12), Diff(12), Rank(8)
        for _ in range(self.num_splits):
            split_columns[current_col] = 12  # Time
            split_columns[chr(ord(current_col) + 1)] = 12  # Diff
            split_columns[chr(ord(current_col) + 2)] = 8   # Rank
            current_col = chr(ord(current_col) + 3)
        
        # Add finish columns
        finish_columns = {
            current_col: 12,      # Finish Time
            chr(ord(current_col) + 1): 12,  # Finish Diff
            chr(ord(current_col) + 2): 10   # Status
        }
        
        return {**base_columns, **split_columns, **finish_columns}

    def write_header_section(self, ws, styles):
        """Writes the header section of the Excel sheet."""
        # Title
        ws['B1'] = self.excel_title
        ws['B1'].font = styles['title']
        ws['B1'].alignment = styles['alignments']['left']
        ws.row_dimensions[1].height = 30
        
        # Get season and format team name
        season = self.get_season(self.date_var.get())
        current_team_name = self.team_names[self.current_team]
        team_with_season = f"{current_team_name} - {season}" if season else current_team_name
        
        # Header information
        header_info = {
            ('B2', 'Team:'): team_with_season,
            ('B3', 'Session #:'): self.session_var.get(),
            ('B4', 'Event:'): self.event_var.get(),
            ('B5', 'Snow Condition:'): self.snow_condition_var.get(),
            ('E2', 'Date:'): self.date_var.get(),
            ('E3', 'Start Time:'): self.time_var.get(),
            ('E4', 'Hill:'): self.hill_var.get(),
            ('E5', 'Weather:'): (f"{self.sky_condition_var.get()}, "
                            f"{self.precipitation_var.get()}, "
                            f"{self.wind_condition_var.get()}")
        }
        
        # Write header info with consistent styling
        for (cell_coord, label), value in header_info.items():
            # Label
            ws[cell_coord] = label
            ws[cell_coord].font = styles['header']
            ws[cell_coord].alignment = styles['alignments']['left']
            
            # Value
            value_col = chr(ord(cell_coord[0]) + 1)
            value_cell = f"{value_col}{cell_coord[1]}"
            ws[value_cell] = value
            ws[value_cell].font = styles['normal']
            ws[value_cell].alignment = styles['alignments']['left']
        
        return 8  # Return next row for data


    def write_run_data(self, ws, run_data, start_row, styles):
        """Writes run data with comprehensive split handling and formatting."""
        current_row = start_row + 2  # Space before run section
        
        # Run header
        ws[f'B{current_row}'] = f"Run {run_data[0]['run']}"
        ws[f'B{current_row}'].font = styles['header']
        current_row += 1

        # Generate headers based on available splits
        headers = ['Bib #', 'Name']
        for i in range(self.num_splits):
            headers.extend([f'Split {i+1}', 'Diff.', 'Rank'])
        headers.extend(['Finish', 'Diff.', 'Status'])

        # Write column headers with consistent styling
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = styles['header']
            cell.border = styles['borders']['thick']
            cell.alignment = styles['alignments']['center']
            cell.fill = styles['fills']['header']
        current_row += 1

        # Process and collect valid times
        split_data = self.collect_valid_times(run_data)
        finish_data = self.collect_finish_times(run_data)

        # Write athlete data rows
        for entry in run_data:
            current_row = self.write_athlete_row(ws, entry, split_data, finish_data, 
                                            current_row, styles)
            
            # Add error details if present
            if entry['error_details']:
                current_row = self.write_error_details(ws, entry, current_row, styles)

        return current_row

    def collect_valid_times(self, run_data):
        """Collects and organizes valid split times for each split."""
        split_data = [[] for _ in range(self.num_splits)]
        
        for entry in run_data:
            if entry['status'].upper() not in ['DNS', 'ERR']:
                # Process each split
                for i, split in enumerate(entry['splits']):
                    if (split is not None and split > 0 and 
                        (i == 0 or self.is_valid_split_progression(entry['splits'][:i+1]))):
                        split_data[i].append((split, entry['bib']))
        
        # Sort each split's times
        for split_times in split_data:
            if split_times:
                split_times.sort(key=lambda x: x[0])
                
        return split_data

    def collect_finish_times(self, run_data):
        """Collects valid finish times."""
        finish_data = []
        
        for entry in run_data:
            if (entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR'] and 
                entry['finish'] is not None and entry['finish'] > 0):
                # Only include finish time if split progression is valid
                if self.is_valid_split_progression(entry['splits']):
                    finish_data.append((entry['finish'], entry['bib']))
        
        if finish_data:
            finish_data.sort(key=lambda x: x[0])
        
        return finish_data

    def is_valid_split_progression(self, splits):
        """Verifies that split times progress logically."""
        valid_splits = [s for s in splits if s is not None and s > 0]
        return all(valid_splits[i] > valid_splits[i-1] for i in range(1, len(valid_splits)))

    def write_athlete_row(self, ws, entry, split_data, finish_data, current_row, styles):
        """Writes a single athlete's row with proper formatting."""
        bib = entry['bib']
        status = entry['status'].upper()
        athlete_name = self.get_athlete_name(bib)
        athlete = self.find_athlete_by_bib(bib)  # Get full athlete data for gender info
        
        # Write base data
        ws.cell(row=current_row, column=2, value=bib).alignment = styles['alignments']['center']
        name_cell = ws.cell(row=current_row, column=3, value=athlete_name)
        name_cell.alignment = styles['alignments']['left']
        
        # Add gender indicator if available
        if athlete and 'gender' in athlete:
            name_cell.value = f"{athlete_name} ({athlete['gender']})"
        
        current_col = 4  # Start of split columns
        
        # Process each split
        for split_idx, split in enumerate(entry['splits']):
            if split is not None and split > 0:
                # Get best split and rank
                best_split = split_data[split_idx][0][0] if split_data[split_idx] else None
                split_diff = split - best_split if best_split is not None else None
                split_rank = next((i + 1 for i, (_, b) in enumerate(split_data[split_idx]) 
                                if b == bib), '')
                
                # Write split time
                time_cell = ws.cell(row=current_row, column=current_col, 
                                value=self.format_time(split))
                time_cell.alignment = styles['alignments']['center']
                
                # Write difference
                if split_diff is not None:
                    diff_cell = ws.cell(row=current_row, column=current_col + 1, 
                                    value=self.format_time(split_diff, True))
                    diff_cell.alignment = styles['alignments']['center']
                
                # Write rank
                rank_cell = ws.cell(row=current_row, column=current_col + 2, value=split_rank)
                rank_cell.alignment = styles['alignments']['center']
                
                # Apply gradient color based on rank
                if split_rank:
                    gradient_color = self.get_gradient_color(
                        split_rank - 1,
                        len(split_data[split_idx]),
                        opacity=0.6 if split_idx == 0 else 0.5
                    )
                    time_cell.fill = gradient_color
            
            current_col += 3  # Move to next split columns
        
        # Write finish data
        if status not in ['DNF', 'DSQ', 'DNS', 'ERR'] and entry['finish'] is not None:
            best_finish = finish_data[0][0] if finish_data else None
            finish_diff = entry['finish'] - best_finish if best_finish is not None else None
            finish_rank = next((i + 1 for i, (_, b) in enumerate(finish_data) 
                            if b == bib), '')
            
            # Finish time
            finish_cell = ws.cell(row=current_row, column=current_col, 
                                value=self.format_time(entry['finish']))
            finish_cell.alignment = styles['alignments']['center']
            
            # Difference
            if finish_diff is not None:
                ws.cell(row=current_row, column=current_col + 1, 
                    value=self.format_time(finish_diff, True))
            
            # Apply gradient color
            if finish_rank:
                finish_cell.fill = self.get_gradient_color(
                    finish_rank - 1,
                    len(finish_data),
                    opacity=0.5
                )
        
        # Status
        status_cell = ws.cell(row=current_row, column=current_col + 2, value=status)
        status_cell.alignment = styles['alignments']['center']
        
        # Apply borders to entire row
        for col in range(2, current_col + 3):
            cell = ws.cell(row=current_row, column=col)
            cell.border = styles['borders']['thin']
        
        return current_row + 1

    def write_error_details(self, ws, entry, current_row, styles):
        """Writes validation error details under athlete row."""
        # Error header
        error_cell = ws.cell(row=current_row, column=2, 
                            value=f"Validation Issues for Bib {entry['bib']}:")
        error_cell.font = styles['error']
        current_row += 1
        
        # Individual errors
        for error in entry['error_details']:
            error_detail = ws.cell(row=current_row, column=2, value=f"  • {error}")
            error_detail.font = styles['error']
            current_row += 1
        
        return current_row

    def create_athlete_analysis_sheet(self, wb, styles):
        """Creates detailed athlete analysis sheet with individual performance tracking."""
        analysis_sheet = wb.create_sheet(title="Athlete Analysis")
        
        # Set column widths
        self.set_analysis_sheet_columns(analysis_sheet)
        
        current_row = 1
        
        # Sheet title
        analysis_sheet['B1'] = "Athlete Performance Analysis"
        analysis_sheet['B1'].font = styles['title']
        current_row += 2

        # Process each athlete in the current team
        for athlete in sorted(self.athletes[self.current_team], key=lambda x: int(x['bib'])):
            current_row = self.write_athlete_analysis(analysis_sheet, athlete, current_row, styles)
            current_row += 2  # Space between athletes

    def set_analysis_sheet_columns(self, ws):
        """Sets up column structure for athlete analysis sheet."""
        base_widths = {
            'A': 6,    # Index
            'B': 18,   # Bib/Run
            'C': 25,   # Name/Description
        }
        
        # Dynamic split columns (Time, Diff, Rank, Progress for each split)
        current_col = 'D'
        for _ in range(self.num_splits):
            for width in [12, 12, 8, 8]:  # Time, Diff, Rank, Progress
                ws.column_dimensions[current_col] = width
                current_col = chr(ord(current_col) + 1)
        
        # Finish columns
        finish_widths = {
            current_col: 12,      # Finish Time
            chr(ord(current_col) + 1): 12,  # Diff
            chr(ord(current_col) + 2): 10,  # Status
            chr(ord(current_col) + 3): 8    # Progress
        }
        
        # Apply all widths
        for col, width in {**base_widths, **finish_widths}.items():
            ws.column_dimensions[col].width = width

    def write_athlete_analysis(self, ws, athlete, start_row, styles):
        """Writes comprehensive analysis for a single athlete."""
        athlete_bib = int(athlete['bib'])
        athlete_name = athlete['name']
        gender = athlete.get('gender', 'N/A')
        
        # Athlete header with gender
        header = f"{athlete_name} (Bib {athlete_bib}, {gender})"
        header_cell = ws.cell(row=start_row, column=2, value=header)
        header_cell.font = styles['header']
        header_cell.fill = styles['fills']['header']
        
        current_row = start_row + 1
        
        # Generate dynamic headers based on splits
        headers = ['Run #']
        for i in range(self.num_splits):
            headers.extend([f'Split {i+1}', 'Diff', 'Rank', '▲/▼'])
        headers.extend(['Finish', 'Diff', 'Status', 'Trend'])
        
        # Write headers
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = styles['header']
            cell.border = styles['borders']['thick']
            cell.alignment = styles['alignments']['center']
        current_row += 1
        
        # Collect and analyze athlete's data across all runs
        athlete_data = self.collect_athlete_data(athlete_bib)
        if not athlete_data:
            ws.cell(row=current_row, column=2, value="No timing data available")
            return current_row + 1
        
        # Write detailed performance data
        current_row = self.write_athlete_performance(ws, athlete_data, current_row, styles)
        
        # Add performance summary
        current_row = self.write_performance_summary(ws, athlete_data, current_row, styles)
        
        return current_row

    def collect_athlete_data(self, bib):
        """Collects and preprocesses all data for an athlete."""
        athlete_data = {
            'runs': [],
            'best_splits': [None] * self.num_splits,
            'best_finish': None,
            'improvements': {
                'splits': [0] * self.num_splits,
                'finish': 0
            },
            'trends': {
                'splits': [[] for _ in range(self.num_splits)],
                'finish': []
            }
        }
        
        for run_num, run_data in sorted(self.timing_data.items()):
            entry = next((e for e in run_data if int(e['bib']) == bib), None)
            if entry:
                run_info = {
                    'run': int(run_num),
                    'splits': entry['splits'],
                    'finish': entry['finish'],
                    'status': entry['status'],
                    'comparisons': {'splits': [], 'finish': None}
                }
                
                # Process splits
                for i, split in enumerate(entry['splits']):
                    if split is not None and split > 0:
                        # Update best time
                        if (athlete_data['best_splits'][i] is None or 
                            split < athlete_data['best_splits'][i]):
                            athlete_data['best_splits'][i] = split
                            athlete_data['improvements']['splits'][i] += 1
                        
                        # Record for trend analysis
                        athlete_data['trends']['splits'][i].append(split)
                
                # Process finish
                if (entry['status'].upper() not in ['DNF', 'DSQ'] and 
                    entry['finish'] is not None):
                    if (athlete_data['best_finish'] is None or 
                        entry['finish'] < athlete_data['best_finish']):
                        athlete_data['best_finish'] = entry['finish']
                        athlete_data['improvements']['finish'] += 1
                    athlete_data['trends']['finish'].append(entry['finish'])
                
                athlete_data['runs'].append(run_info)
        
        return athlete_data

    def write_athlete_performance(self, ws, athlete_data, start_row, styles):
        """Writes detailed performance data for each run."""
        current_row = start_row
        
        for run_info in athlete_data['runs']:
            # Base data
            ws.cell(row=current_row, column=2, value=f"Run {run_info['run']}")
            
            current_col = 4  # Start of split columns
            
            # Process splits
            for i, split in enumerate(run_info['splits']):
                if split is not None and split > 0:
                    # Calculate comparison to best
                    best = athlete_data['best_splits'][i]
                    diff = split - best if best is not None else None
                    
                    # Calculate progression indicator
                    if len(athlete_data['trends']['splits'][i]) > 1:
                        prev_split = athlete_data['trends']['splits'][i][-2]
                        trend = '▼' if split < prev_split else '▲' if split > prev_split else '='
                    else:
                        trend = '-'
                    
                    # Write split data
                    ws.cell(row=current_row, column=current_col, 
                        value=self.format_time(split)).alignment = styles['alignments']['center']
                    
                    if diff is not None:
                        ws.cell(row=current_row, column=current_col + 1,
                            value=self.format_time(diff, True)).alignment = styles['alignments']['center']
                    
                    ws.cell(row=current_row, column=current_col + 3,
                        value=trend).alignment = styles['alignments']['center']
                
                current_col += 4  # Move to next split section
            
            # Write finish data
            if (run_info['status'].upper() not in ['DNF', 'DSQ'] and 
                run_info['finish'] is not None):
                
                finish_diff = (run_info['finish'] - athlete_data['best_finish'] 
                            if athlete_data['best_finish'] is not None else None)
                
                # Calculate finish trend
                if len(athlete_data['trends']['finish']) > 1:
                    prev_finish = athlete_data['trends']['finish'][-2]
                    finish_trend = '▼' if run_info['finish'] < prev_finish else '▲'
                else:
                    finish_trend = '-'
                
                ws.cell(row=current_row, column=current_col,
                    value=self.format_time(run_info['finish'])).alignment = styles['alignments']['center']
                
                if finish_diff is not None:
                    ws.cell(row=current_row, column=current_col + 1,
                        value=self.format_time(finish_diff, True)).alignment = styles['alignments']['center']
            
            ws.cell(row=current_row, column=current_col + 2,
                value=run_info['status']).alignment = styles['alignments']['center']
            
            ws.cell(row=current_row, column=current_col + 3,
                value=finish_trend).alignment = styles['alignments']['center']
            
            # Apply row styling
            for col in range(2, current_col + 4):
                cell = ws.cell(row=current_row, column=col)
                cell.border = styles['borders']['thin']
            
            current_row += 1
        
        return current_row

    def write_performance_summary(self, ws, athlete_data, start_row, styles):
        """Writes performance summary and statistics."""
        current_row = start_row + 1
        
        # Summary header
        summary_cell = ws.cell(row=current_row, column=2, value="Performance Summary")
        summary_cell.font = styles['header']
        summary_cell.fill = styles['fills']['header']
        current_row += 1
        
        # Calculate summary statistics
        summary_stats = self.calculate_athlete_summary(athlete_data)
        
        # Write statistics
        stats_to_write = [
            ("Total Runs", len(athlete_data['runs'])),
            ("Completion Rate", f"{summary_stats['completion_rate']:.1f}%"),
            ("Best Finish", self.format_time(athlete_data['best_finish'])),
            ("Average Finish", self.format_time(summary_stats['avg_finish'])),
            ("Improvement Rate", f"{summary_stats['improvement_rate']:.1f}%"),
            ("Consistency Score", f"{summary_stats['consistency_score']:.1f}%")
        ]
        
        for label, value in stats_to_write:
            ws.cell(row=current_row, column=2, value=label).font = styles['header']
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1
        
        return current_row

    def calculate_athlete_summary(self, athlete_data):
        """Calculates comprehensive summary statistics for an athlete."""
        completed_runs = len([r for r in athlete_data['runs'] 
                            if r['status'].upper() not in ['DNF', 'DSQ']])
        total_runs = len(athlete_data['runs'])
        
        valid_finishes = [r['finish'] for r in athlete_data['runs'] 
                        if r['status'].upper() not in ['DNF', 'DSQ'] and r['finish'] is not None]
        
        summary = {
            'completion_rate': (completed_runs / total_runs * 100) if total_runs > 0 else 0,
            'avg_finish': sum(valid_finishes) / len(valid_finishes) if valid_finishes else None,
            'improvement_rate': 0,
            'consistency_score': 0
        }
        
        # Calculate improvement rate
        if len(valid_finishes) >= 2:
            first_time = valid_finishes[0]
            best_time = min(valid_finishes)
            summary['improvement_rate'] = ((first_time - best_time) / first_time * 100)
            
            # Calculate consistency score
            mean = sum(valid_finishes) / len(valid_finishes)
            variance = sum((t - mean) ** 2 for t in valid_finishes) / len(valid_finishes)
            summary['consistency_score'] = 100 / (1 + variance)
        
        return summary

    def create_category_sheet(self, wb, styles):
        """Creates sheet with gender-based category splits and analysis."""
        category_sheet = wb.create_sheet(title="Category Analysis")
        
        # Set initial column widths
        self.set_category_sheet_columns(category_sheet)
        
        current_row = 1
        
        # Sheet title
        title_cell = category_sheet['B1'] = "Performance Analysis by Category"
        title_cell.font = styles['title']
        current_row += 2
        
        # Process each run with gender separation
        for run_number in sorted(self.timing_data.keys(), key=int):
            current_row = self.write_category_run_data(
                category_sheet, 
                self.timing_data[run_number],
                run_number,
                current_row,
                styles
            )
            current_row += 2  # Space between runs

    def set_category_sheet_columns(self, ws):
        """Sets up columns for category analysis sheet."""
        column_widths = {
            'A': 6,    # Index
            'B': 18,   # Bib
            'C': 25,   # Name
            'D': 12,   # Split 1
            'E': 12,   # Split Diff
            'F': 8,    # Split Rank
            'G': 12,   # Finish
            'H': 12,   # Finish Diff
            'I': 8,    # Status
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def write_category_run_data(self, ws, run_data, run_number, start_row, styles):
        """Writes run data separated by gender categories."""
        current_row = start_row
        
        # Run header
        run_header = ws.cell(row=current_row, column=1, value=f"Run {run_number}")
        run_header.font = styles['header']
        run_header.fill = styles['fills']['header']
        current_row += 2
        
        # Split data by gender
        women_data = []
        men_data = []
        
        for entry in run_data:
            athlete = self.find_athlete_by_bib(entry['bib'])
            if athlete and 'gender' in athlete:
                if athlete['gender'] == 'F':
                    women_data.append(entry)
                else:
                    men_data.append(entry)
        
        # Write women's section
        if women_data:
            current_row = self.write_category_section(
                ws, women_data, "Women", current_row, styles)
            current_row += 2
        
        # Write men's section
        if men_data:
            current_row = self.write_category_section(
                ws, men_data, "Men", current_row, styles)
            current_row += 2
        
        # Add category comparison
        current_row = self.write_category_comparison(
            ws, women_data, men_data, current_row, styles)
        
        return current_row

    def write_category_section(self, ws, data, category, start_row, styles):
        """Writes data for a specific gender category."""
        current_row = start_row
        
        # Category header
        category_cell = ws.cell(row=current_row, column=1, value=category)
        category_cell.font = styles['header']
        category_cell.fill = styles['fills']['category']
        ws.merge_cells(start_row=current_row, start_column=1, 
                    end_row=current_row, end_column=9)
        current_row += 1
        
        # Column headers
        headers = ['Bib', 'Name', 'Split 1', 'Split Diff', 'Rank',
                'Finish', 'Diff', 'Status']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = styles['header']
            cell.fill = styles['fills']['header']
            cell.border = styles['borders']['thick']
            cell.alignment = styles['alignments']['center']
        current_row += 1
        
        # Sort data by finish time
        valid_data = [entry for entry in data 
                    if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']]
        valid_data.sort(key=lambda x: x['finish'] if x['finish'] is not None else float('inf'))
        
        # Get best times for comparisons
        best_split = min((entry['splits'][0] for entry in valid_data 
                        if entry['splits'] and entry['splits'][0] is not None),
                        default=None)
        best_finish = min((entry['finish'] for entry in valid_data 
                        if entry['finish'] is not None),
                        default=None)
        
        # Write athlete data
        for rank, entry in enumerate(valid_data, 1):
            current_row = self.write_category_athlete_row(
                ws, entry, rank, best_split, best_finish, current_row, styles)
        
        # Write DNF/DSQ entries
        other_data = [entry for entry in data if entry not in valid_data]
        for entry in other_data:
            current_row = self.write_category_athlete_row(
                ws, entry, None, best_split, best_finish, current_row, styles)
        
        return current_row

    def write_category_athlete_row(self, ws, entry, rank, best_split, best_finish, row, styles):
        """Writes a single athlete row in the category section."""
        bib = entry['bib']
        name = self.get_athlete_name(bib)
        status = entry['status'].upper()
        
        # Write basic info
        ws.cell(row=row, column=1, value=bib).alignment = styles['alignments']['center']
        ws.cell(row=row, column=2, value=name).alignment = styles['alignments']['left']
        
        # Write split data if valid
        if entry['splits'] and entry['splits'][0] is not None:
            split_cell = ws.cell(row=row, column=3, 
                            value=self.format_time(entry['splits'][0]))
            split_cell.alignment = styles['alignments']['center']
            
            if best_split is not None:
                diff = entry['splits'][0] - best_split
                ws.cell(row=row, column=4,
                    value=self.format_time(diff, True)).alignment = styles['alignments']['center']
            
            if rank is not None:
                ws.cell(row=row, column=5, value=rank).alignment = styles['alignments']['center']
        
        # Write finish data
        if status not in ['DNF', 'DSQ', 'DNS', 'ERR'] and entry['finish'] is not None:
            finish_cell = ws.cell(row=row, column=6, 
                                value=self.format_time(entry['finish']))
            finish_cell.alignment = styles['alignments']['center']
            
            if best_finish is not None:
                diff = entry['finish'] - best_finish
                ws.cell(row=row, column=7,
                    value=self.format_time(diff, True)).alignment = styles['alignments']['center']
        
        # Write status
        ws.cell(row=row, column=8, value=status).alignment = styles['alignments']['center']
        
        # Apply borders
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = styles['borders']['thin']
        
        return row + 1

    def write_category_comparison(self, ws, women_data, men_data, start_row, styles):
        """Adds statistical comparison between gender categories."""
        current_row = start_row + 1
        
        comparison_cell = ws.cell(row=current_row, column=1, value="Category Comparison")
        comparison_cell.font = styles['header']
        comparison_cell.fill = styles['fills']['header']
        ws.merge_cells(start_row=current_row, start_column=1, 
                    end_row=current_row, end_column=9)
        current_row += 2
        
        # Calculate statistics for both categories
        women_stats = self.calculate_category_stats(women_data)
        men_stats = self.calculate_category_stats(men_data)
        
        # Write comparison table
        headers = ['Metric', 'Women', 'Men', 'Difference']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = styles['header']
            cell.fill = styles['fills']['header']
        current_row += 1
        
        # Compare metrics
        metrics_to_compare = [
            ('Athletes', lambda s: s['total_athletes'], '{:d}'),
            ('Completion Rate', lambda s: s['completion_rate'], '{:.1f}%'),
            ('Average Finish', lambda s: s['avg_finish'], self.format_time),
            ('Best Time', lambda s: s['best_time'], self.format_time),
            ('Time Range', lambda s: s['time_range'], self.format_time)
        ]
        
        for label, getter, formatter in metrics_to_compare:
            women_val = getter(women_stats) if women_stats else None
            men_val = getter(men_stats) if men_stats else None
            
            row_data = [label]
            
            # Format values
            if women_val is not None:
                row_data.append(formatter(women_val))
            else:
                row_data.append('N/A')
                
            if men_val is not None:
                row_data.append(formatter(men_val))
            else:
                row_data.append('N/A')
            
            # Calculate difference where applicable
            if women_val is not None and men_val is not None and label != 'Athletes':
                diff = women_val - men_val
                row_data.append(self.format_time(diff, True) if isinstance(women_val, float)
                            else f"{diff:+.1f}%")
            else:
                row_data.append('N/A')
            
            # Write row
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.alignment = styles['alignments']['center']
                cell.border = styles['borders']['thin']
            
            current_row += 1
        
        return current_row

    def calculate_category_stats(self, data):
        """Calculates statistics for a gender category."""
        if not data:
            return None
            
        valid_times = [entry['finish'] for entry in data 
                    if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']
                    and entry['finish'] is not None]
        
        if not valid_times:
            return None
        
        return {
            'total_athletes': len(data),
            'completion_rate': len(valid_times) / len(data) * 100,
            'avg_finish': sum(valid_times) / len(valid_times),
            'best_time': min(valid_times),
            'time_range': max(valid_times) - min(valid_times)
        }

    def create_statistics_sheet(self, wb, styles):
        """Creates comprehensive statistics and analysis sheet."""
        stats_sheet = wb.create_sheet(title="Training Statistics")
        
        # Set column widths
        self.set_statistics_sheet_columns(stats_sheet)
        
        current_row = 1
        
        # Write main sections
        current_row = self.write_session_overview(stats_sheet, current_row, styles)
        current_row = self.write_split_analysis(stats_sheet, current_row, styles)
        current_row = self.write_athlete_rankings(stats_sheet, current_row, styles)
        current_row = self.write_trend_analysis(stats_sheet, current_row, styles)
        
        # Add final notes or warnings if any
        if hasattr(self, 'analysis_warnings'):
            current_row = self.write_analysis_warnings(stats_sheet, current_row, styles)

    def set_statistics_sheet_columns(self, ws):
        """Sets column widths for statistics sheet."""
        column_widths = {
            'A': 6,    # Index
            'B': 25,   # Description/Name
            'C': 15,   # Value/Time
            'D': 15,   # Additional Info
            'E': 15,   # Extra Stats
            'F': 15,   # Notes
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def write_session_overview(self, ws, start_row, styles):
        """Writes overall session statistics and information."""
        current_row = start_row
        
        # Section header
        header_cell = ws.cell(row=current_row, column=1, value="Session Overview")
        header_cell.font = styles['title']
        current_row += 2
        
        # Calculate overview statistics
        overview_stats = self.calculate_session_statistics()
        
        # Write basic information
        info_rows = [
            ("Event", self.event_var.get()),
            ("Date", self.date_var.get()),
            ("Hill", self.hill_var.get()),
            ("Conditions", f"{self.snow_condition_var.get()}, {self.sky_condition_var.get()}"),
            ("Total Runs", str(len(self.timing_data))),
            ("Total Athletes", str(overview_stats['total_athletes'])),
            ("Completion Rate", f"{overview_stats['completion_rate']:.1f}%"),
            ("DNF Rate", f"{overview_stats['dnf_rate']:.1f}%"),
            ("Average Athletes per Run", f"{overview_stats['avg_athletes_per_run']:.1f}"),
        ]
        
        for label, value in info_rows:
            ws.cell(row=current_row, column=2, value=label).font = styles['header']
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1
        
        current_row += 1
        return current_row

    def write_split_analysis(self, ws, start_row, styles):
        """Writes detailed analysis for each split."""
        current_row = start_row
        
        # Section header
        header_cell = ws.cell(row=current_row, column=1, value="Split Time Analysis")
        header_cell.font = styles['title']
        current_row += 2
        
        # Analyze each split
        for split_idx in range(self.num_splits):
            current_row = self.write_single_split_analysis(
                ws, split_idx, current_row, styles)
            current_row += 1
        
        return current_row

    def write_single_split_analysis(self, ws, split_idx, start_row, styles):
        """Analyzes and writes statistics for a single split."""
        current_row = start_row
        
        # Split header
        split_header = ws.cell(row=current_row, column=2, 
                            value=f"Split {split_idx + 1} Statistics")
        split_header.font = styles['header']
        split_header.fill = styles['fills']['header']
        current_row += 1
        
        # Collect split times across all runs
        split_stats = self.calculate_split_statistics(split_idx)
        
        # Write statistics
        stats_rows = [
            ("Best Time", self.format_time(split_stats['best_time'])),
            ("Average Time", self.format_time(split_stats['avg_time'])),
            ("Standard Deviation", f"{split_stats['std_dev']:.3f}s"),
            ("Consistency Score", f"{split_stats['consistency_score']:.1f}%"),
            ("Total Valid Times", str(split_stats['valid_count'])),
            ("Invalid Rate", f"{split_stats['invalid_rate']:.1f}%")
        ]
        
        for label, value in stats_rows:
            ws.cell(row=current_row, column=2, value=label)
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1
        
        # Add progression analysis if this is a normal split (not acceleration)
        if not split_stats.get('is_acceleration_split', False):
            current_row = self.write_split_progression(
                ws, split_idx, current_row, styles)
        
        return current_row

    def write_athlete_rankings(self, ws, start_row, styles):
        """Writes athlete performance rankings and analysis."""
        current_row = start_row
        
        # Section header
        header_cell = ws.cell(row=current_row, column=1, value="Athlete Rankings")
        header_cell.font = styles['title']
        current_row += 2
        
        # Calculate athlete performance metrics
        athlete_metrics = self.calculate_athlete_metrics()
        
        # Write ranking tables for different categories
        ranking_categories = [
            ("Best Overall Time", lambda x: x['best_time']),
            ("Most Consistent", lambda x: -x['consistency_score']),  # Negative for proper sorting
            ("Most Improved", lambda x: -x['improvement_rate']),
            ("Highest Completion Rate", lambda x: -x['completion_rate'])
        ]
        
        for category, sort_key in ranking_categories:
            current_row = self.write_ranking_category(
                ws, category, athlete_metrics, sort_key, current_row, styles)
            current_row += 2
        
        return current_row

    def write_trend_analysis(self, ws, start_row, styles):
        """Writes trend analysis for the training session."""
        current_row = start_row
        
        # Section header
        header_cell = ws.cell(row=current_row, column=1, value="Trend Analysis")
        header_cell.font = styles['title']
        current_row += 2
        
        # Calculate trends
        trends = self.analyze_session_trends()
        
        # Write trend information
        trend_sections = [
            ("Overall Session Trends", trends['overall']),
            ("Split Time Trends", trends['splits']),
            ("Performance Patterns", trends['patterns']),
            ("Weather Impact", trends['weather_impact'])
        ]
        
        for section, trend_data in trend_sections:
            section_cell = ws.cell(row=current_row, column=2, value=section)
            section_cell.font = styles['header']
            current_row += 1
            
            for trend in trend_data:
                ws.cell(row=current_row, column=2, value=f"• {trend}")
                current_row += 1
            
            current_row += 1
        
        return current_row

    def calculate_session_statistics(self):
        """Calculates overall session statistics."""
        total_athletes = 0
        completed_runs = 0
        total_runs = 0
        athletes_per_run = []
        
        for run_data in self.timing_data.values():
            run_athletes = len(run_data)
            total_athletes = max(total_athletes, run_athletes)
            athletes_per_run.append(run_athletes)
            
            completed = sum(1 for entry in run_data 
                        if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR'])
            completed_runs += completed
            total_runs += run_athletes
        
        return {
            'total_athletes': total_athletes,
            'completion_rate': (completed_runs / total_runs * 100) if total_runs > 0 else 0,
            'dnf_rate': ((total_runs - completed_runs) / total_runs * 100) 
                        if total_runs > 0 else 0,
            'avg_athletes_per_run': sum(athletes_per_run) / len(athletes_per_run) 
                                if athletes_per_run else 0
        }

    def analyze_session_trends(self):
        """Analyzes various trends throughout the session."""
        trends = {
            'overall': [],
            'splits': [],
            'patterns': [],
            'weather_impact': []
        }
        
        # Analyze time progression
        run_averages = []
        for run_data in self.timing_data.values():
            valid_times = [entry['finish'] for entry in run_data 
                        if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']
                        and entry['finish'] is not None]
            if valid_times:
                run_averages.append(sum(valid_times) / len(valid_times))
        
        if run_averages:
            # Overall trend
            if all(run_averages[i] <= run_averages[i-1] for i in range(1, len(run_averages))):
                trends['overall'].append("Consistent improvement throughout session")
            elif all(run_averages[i] >= run_averages[i-1] for i in range(1, len(run_averages))):
                trends['overall'].append("Times gradually increased - possible fatigue effect")
            
            # Variation analysis
            avg_variation = sum(abs(run_averages[i] - run_averages[i-1]) 
                            for i in range(1, len(run_averages))) / (len(run_averages) - 1)
            if avg_variation < 0.5:
                trends['patterns'].append("Very consistent performance across runs")
            elif avg_variation > 2.0:
                trends['patterns'].append("High variation between runs - may indicate changing conditions")
        
        # Weather impact analysis
        if hasattr(self, 'snow_condition_var'):
            snow_cond = self.snow_condition_var.get()
            if "Hard" in snow_cond:
                trends['weather_impact'].append("Hard snow conditions - typically faster times")
            elif "Soft" in snow_cond:
                trends['weather_impact'].append("Soft snow conditions - may affect consistency")
        
        return trends

    def write_analysis_warnings(self, ws, start_row, styles):
        """Writes any analysis warnings or notes."""
        current_row = start_row + 1
        
        warning_cell = ws.cell(row=current_row, column=1, value="Analysis Notes")
        warning_cell.font = styles['header']
        warning_cell.fill = styles['fills']['header']
        current_row += 1
        
        for warning in self.analysis_warnings:
            cell = ws.cell(row=current_row, column=2, value=f"• {warning}")
            cell.font = styles['error']
            current_row += 1
        
        return current_row

    def get_gradient_color(self, index, total, opacity=0.5):
        """
        Enhanced gradient color generator with better color ranges and opacity control.
        
        Args:
            index: Position in sequence (0 = best)
            total: Total number of items
            opacity: Opacity level (0.0 to 1.0)
        """
        if total <= 1:
            # Single entry - light green
            base_color = (144, 238, 144)  # Light green
            bg_color = (255, 255, 255)    # White background
            r = int(base_color[0] * opacity + bg_color[0] * (1 - opacity))
            g = int(base_color[1] * opacity + bg_color[1] * (1 - opacity))
            b = int(base_color[2] * opacity + bg_color[2] * (1 - opacity))
            hex_color = f"{r:02x}{g:02x}{b:02x}".upper()
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        # Color definitions for different types of splits
        color_schemes = {
            'normal': [
                (144, 238, 144),  # Light green
                (255, 230, 102),  # Warm yellow
                (255, 165, 0),    # Orange
                (255, 99, 71),    # Tomato red
                (255, 64, 64)     # Bright red
            ],
            'acceleration': [
                (135, 206, 250),  # Light blue
                (100, 149, 237),  # Cornflower blue
                (65, 105, 225),   # Royal blue
                (0, 0, 205),      # Medium blue
                (0, 0, 139)       # Dark blue
            ]
        }

        # Select color scheme based on opacity
        # Use acceleration scheme for first split (opacity > 0.55)
        colors = color_schemes['acceleration'] if opacity > 0.55 else color_schemes['normal']
        
        # Calculate position in gradient
        position = index / (total - 1)
        num_segments = len(colors) - 1
        segment_length = 1.0 / num_segments
        segment_index = min(int(position / segment_length), num_segments - 1)
        segment_position = (position - (segment_index * segment_length)) / segment_length
        
        # Interpolate between colors
        color1 = colors[segment_index]
        color2 = colors[segment_index + 1]
        
        # Calculate interpolated color
        r = int(color1[0] + (color2[0] - color1[0]) * segment_position)
        g = int(color1[1] + (color2[1] - color1[1]) * segment_position)
        b = int(color1[2] + (color2[2] - color1[2]) * segment_position)
        
        # Apply opacity
        bg_color = (255, 255, 255)  # White background
        r = int(r * opacity + bg_color[0] * (1 - opacity))
        g = int(g * opacity + bg_color[0] * (1 - opacity))
        b = int(b * opacity + bg_color[0] * (1 - opacity))
        
        hex_color = f"{r:02x}{g:02x}{b:02x}".upper()
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    def get_category_color(self, category_type):
        """Returns consistent colors for different categories."""
        colors = {
            'men': 'CCE5FF',      # Light blue
            'women': 'FFE5CC',    # Light orange
            'header': 'E6E6E6',   # Light gray
            'alert': 'FFD9D9',    # Light red
            'success': 'D9FFD9',  # Light green
            'neutral': 'F2F2F2'   # Very light gray
        }
        return colors.get(category_type, colors['neutral'])

    def apply_cell_style(self, cell, style_type, styles, value=None):
        """
        Applies consistent styling to cells based on type.
        
        Args:
            cell: The cell to style
            style_type: Type of styling to apply ('header', 'data', 'alert', etc.)
            styles: Dictionary of style definitions
            value: Optional value to write to cell
        """
        if value is not None:
            cell.value = value
        
        # Base styles
        cell.alignment = styles['alignments']['center']
        cell.border = styles['borders']['thin']
        
        # Apply specific styles based on type
        if style_type == 'header':
            cell.font = styles['header']
            cell.fill = PatternFill(start_color=self.get_category_color('header'), 
                                end_color=self.get_category_color('header'), 
                                fill_type='solid')
            cell.border = styles['borders']['thick']
        elif style_type == 'data':
            cell.font = styles['normal']
        elif style_type == 'alert':
            cell.font = styles['error']
            cell.fill = PatternFill(start_color=self.get_category_color('alert'), 
                                end_color=self.get_category_color('alert'), 
                                fill_type='solid')
        elif style_type == 'time':
            cell.font = styles['normal']
            cell.number_format = '[h]:mm:ss.000'
        elif style_type == 'diff':
            cell.font = styles['normal']
            cell.number_format = '+[h]:mm:ss.000;-[h]:mm:ss.000'

    def apply_row_style(self, ws, row, start_col, end_col, style_type, styles):
        """Applies consistent styling to an entire row."""
        for col in range(start_col, end_col + 1):
            self.apply_cell_style(ws.cell(row=row, column=col), style_type, styles)

    def merge_and_style_cells(self, ws, start_row, start_col, end_row, end_col, value, 
                            style_type, styles):
        """Merges cells and applies consistent styling."""
        ws.merge_cells(start_row=start_row, start_column=start_col,
                    end_row=end_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=value)
        self.apply_cell_style(cell, style_type, styles)

    def format_time_cell(self, time_value, as_difference=False):
        """
        Returns formatted time value for Excel cells.
        
        Args:
            time_value: Time in seconds
            as_difference: Whether to format as a time difference
        """
        if time_value is None:
            return ""
            
        if as_difference:
            sign = "+" if time_value >= 0 else "-"
            abs_time = abs(time_value)
        else:
            sign = ""
            abs_time = time_value
        
        # Handle different time ranges
        if abs_time < 1:  # Less than 1 second
            formatted = f"{abs_time:.3f}"
        elif abs_time < 60:  # Less than 1 minute
            formatted = f"{abs_time:.2f}"
        else:  # 1 minute or more
            minutes = int(abs_time // 60)
            seconds = abs_time % 60
            formatted = f"{minutes}:{seconds:05.2f}"
        
        return f"{sign}{formatted}" if as_difference else formatted

    def create_excel_styles(self):
        """Creates comprehensive style dictionary for Excel formatting."""
        styles = {
            'fonts': {
                'title': Font(name='Arial', size=18, bold=True),
                'header': Font(name='Arial', size=11, bold=True),
                'normal': Font(name='Arial', size=11),
                'error': Font(name='Arial', size=11, color="FF0000"),
                'alert': Font(name='Arial', size=11, color="FF6B00"),
                'success': Font(name='Arial', size=11, color="008000")
            },
            'borders': {
                'thick': Border(
                    left=Side(style='thick'),
                    right=Side(style='thick'),
                    top=Side(style='thick'),
                    bottom=Side(style='thick')
                ),
                'thin': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                'none': Border(
                    left=Side(style='none'),
                    right=Side(style='none'),
                    top=Side(style='none'),
                    bottom=Side(style='none')
                )
            },
            'alignments': {
                'center': Alignment(horizontal='center', vertical='center'),
                'left': Alignment(horizontal='left', vertical='center'),
                'right': Alignment(horizontal='right', vertical='center')
            },
            'fills': {
                'header': PatternFill(start_color="E6E6E6", 
                                    end_color="E6E6E6", 
                                    fill_type="solid"),
                'alert': PatternFill(start_color="FFD9D9", 
                                end_color="FFD9D9", 
                                fill_type="solid"),
                'success': PatternFill(start_color="D9FFD9", 
                                    end_color="D9FFD9", 
                                    fill_type="solid"),
                'neutral': PatternFill(start_color="F2F2F2", 
                                    end_color="F2F2F2", 
                                    fill_type="solid")
            }
        }
        return styles

    def format_excel_range(self, ws, start_row, start_col, end_row, end_col, 
                        format_type='data'):
        """
        Applies consistent formatting to a range of cells.
        
        Args:
            ws: Worksheet
            start_row: Starting row
            start_col: Starting column
            end_row: Ending row
            end_col: Ending column
            format_type: Type of formatting to apply
        """
        styles = self.create_excel_styles()
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                self.apply_cell_style(cell, format_type, styles)

    def add_cell_comment(self, cell, comment_text):
        """Adds a formatted comment to a cell."""
        from openpyxl.comments import Comment
        comment = Comment(comment_text, "Timing System")
        cell.comment = comment




    def autocomplete_athlete_name(self, event):
        """
        Fixed autocomplete that works with new UI.
        """
        input_text = self.athlete_name_entry.get().lower()
        
        # Update entry width if needed
        if not self.name_entry_width:
            self.name_entry_width = self.athlete_name_entry.winfo_width()
            self.line_height = self.athlete_name_entry.winfo_reqheight()
        
        if not input_text:
            self.autocomplete_listbox.grid_remove()
            return
        
        # Filter suggestions
        suggestions = [name for name in reversed(self.recent_names) 
                    if name.lower().startswith(input_text)]
        suggestions = suggestions[:3]  # Limit to 3 suggestions
        
        if suggestions:
            self.autocomplete_listbox.delete(0, tk.END)
            for suggestion in suggestions:
                self.autocomplete_listbox.insert(tk.END, suggestion)
                
            # Position and show listbox
            self.autocomplete_listbox.grid(
                row=2, column=1, padx=(0, 5), pady=(0, 5), sticky="w")
            # Set height based on number of suggestions
            self.autocomplete_listbox.configure(height=len(suggestions))
        else:
            self.autocomplete_listbox.grid_remove()
        

    def parse_csv_file(self, file_path):
        """
        Initial CSV file parsing to get session info and detect splits structure.
        """
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                separator = '>'
                
                # Initialize variables
                session_info = {
                    'session': "",
                    'date': "",
                    'time': "",
                    'splits': []  # Will store detected split columns
                }
                
                # First pass: Find header and detect splits
                header_line_idx = -1
                for i, line in enumerate(lines[:20]):  # Check first 20 lines
                    line = line.strip()
                    if not line:
                        continue
                        
                    parts = line.split(separator)
                    
                    # Look for session info
                    if "Session #" in line:
                        session_info['session'] = parts[1].strip('#').strip()
                    elif "Date" == parts[0]:
                        # Convert date format if needed
                        date_str = parts[1].strip()
                        try:
                            date_obj = datetime.strptime(date_str, '%m/%d/%y')
                            session_info['date'] = date_obj.strftime('%d/%m/%Y')
                        except:
                            session_info['date'] = date_str
                    elif "Time" == parts[0]:
                        session_info['time'] = parts[1].strip()
                    
                    # Detect header line and split columns
                    if "Bib#" in line:
                        header_line_idx = i
                        headers = [h.lower().strip() for h in parts]
                        
                        # Find all split columns
                        for idx, header in enumerate(headers):
                            if "split" in header:
                                split_num = ''.join(filter(str.isdigit, header))
                                if split_num:  # If there's a number in the split header
                                    session_info['splits'].append({
                                        'index': idx,
                                        'number': int(split_num)
                                    })
                        
                        # Sort splits by their number
                        session_info['splits'].sort(key=lambda x: x['number'])
                        break
                
                # Set the number of splits for the app
                self.num_splits = len(session_info['splits'])
                
                # Update UI with session info
                self.date_var.set(session_info['date'])
                self.time_var.set(session_info['time'])
                self.session_var.set(session_info['session'])
                
                return session_info
                
        except Exception as e:
            print(f"Error reading file: {str(e)}")
            messagebox.showerror("Error", f"Error reading file: {str(e)}")
            return None

    def parse_timing_data(self, file_path):
        """
        Enhanced parsing of timing data handling multiple splits.
        """
        timing_data = {}
        header_found = False
        column_indices = {}
        split_columns = []
        
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()
                separator = '>'
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Find header and map columns
                    if not header_found and "Bib#" in line:
                        header_found = True
                        headers = [h.lower().strip() for h in line.split(separator)]
                        
                        for i, header in enumerate(headers):
                            if "bib" in header:
                                column_indices['bib'] = i
                            elif "run" in header:
                                column_indices['run'] = i
                            elif "finish time" in header:
                                column_indices['finish'] = i
                            elif "status" in header:
                                column_indices['status'] = i
                            elif "split" in header:
                                split_columns.append(i)
                        
                        self.num_splits = len(split_columns)  # Update number of splits
                        continue
                    
                    # Process data lines
                    if header_found:
                        data = line.split(separator)
                        if len(data) <= 1:
                            continue
                        
                        try:
                            # Get run number
                            run_number = data[column_indices['run']].strip()
                            if not run_number:
                                continue
                            
                            # Initialize run in timing data
                            if run_number not in timing_data:
                                timing_data[run_number] = []
                            
                            # Process split times
                            splits = []
                            valid_split_count = 0
                            
                            for split_col in split_columns:
                                if split_col < len(data):
                                    split_time = self.validate_time(data[split_col].strip())
                                    splits.append(split_time)
                                    if split_time is not None and split_time > 0:
                                        valid_split_count += 1
                                else:
                                    splits.append(None)
                            
                            # Create entry
                            entry = {
                                'bib': data[column_indices['bib']].strip(),
                                'splits': splits,
                                'valid_splits': valid_split_count,
                                'finish': self.validate_time(
                                    data[column_indices['finish']].strip() 
                                    if 'finish' in column_indices and 
                                    column_indices['finish'] < len(data) 
                                    else ''
                                ),
                                'status': data[column_indices.get('status', -1)].strip() 
                                        if 'status' in column_indices and 
                                            column_indices['status'] < len(data) 
                                        else '',
                                'run': run_number
                            }
                            
                            # Add entry only if it has valid data
                            if entry['valid_splits'] > 0 or entry['finish'] is not None:
                                timing_data[run_number].append(entry)
                            
                        except IndexError as e:
                            print(f"Error processing line: {line}")
                            print(f"Error details: {str(e)}")
                            continue
                
                if not timing_data:
                    messagebox.showerror("Error", "No valid timing data found in the file.")
                    return None
                
                return timing_data
                
        except Exception as e:
            messagebox.showerror("Error", f"Error parsing CSV file: {str(e)}")
            return None

    def validate_time(self, time_str):
        """
        Enhanced time validation handling multiple time formats.
        """
        if not time_str or not isinstance(time_str, str):
            return None
        
        time_str = time_str.strip().upper()
        if time_str in ['DNF', 'DSQ', '', 'DNS', '0']:
            return None
        
        try:
            if ':' in time_str:
                # Handle MM:SS.sss format
                parts = time_str.split(':')
                if len(parts) != 2:
                    return None
                    
                minutes = float(parts[0])
                
                # Handle seconds with potential milliseconds
                if '.' in parts[1]:
                    seconds, milliseconds = parts[1].split('.')
                    return (minutes * 60) + float(seconds) + (float(milliseconds) / 1000)
                else:
                    return (minutes * 60) + float(parts[1])
            else:
                # Handle plain seconds format
                return float(time_str)
                
        except (ValueError, TypeError):
            return None

    def create_formatted_excel(self, output_path):
        """Creates a formatted Excel file handling multiple splits properly."""
        wb = Workbook()
        ws = wb.active
        
        # Define styles
        try:
            title_font = Font(name='Avenir Next LT Pro', size=18, bold=True)
        except:
            title_font = Font(name='Arial', size=18, bold=True)
        
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=11)
        
        # Define borders
        header_border = Border(
            left=Side(style='thick'), 
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        normal_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set column widths dynamically based on number of splits
        base_columns = {
            'A': 6,   # Index
            'B': 18,  # Bib #
            'C': 20,  # Name
        }
        
        # Calculate column positions for splits and finish
        current_col = 'D'
        split_columns = []  # Store split column positions
        
        # Add columns for each split (Time, Diff, Rank)
        for split_num in range(self.num_splits):
            split_info = {
                'time_col': current_col,
                'diff_col': chr(ord(current_col) + 1),
                'rank_col': chr(ord(current_col) + 2)
            }
            split_columns.append(split_info)
            
            # Set width for all three columns
            for _ in range(3):
                ws.column_dimensions[current_col] = 12
                current_col = chr(ord(current_col) + 1)
        
        # Add finish columns
        finish_columns = {
            current_col: 12,      # Finish Time
            chr(ord(current_col) + 1): 12,  # Finish Diff
            chr(ord(current_col) + 2): 10   # Status
        }
        
        # Apply all column widths
        for col, width in {**base_columns, **finish_columns}.items():
            ws.column_dimensions[col].width = width
        
        # Write title and headers
        current_row = 1
        ws['B1'] = self.excel_title
        ws['B1'].font = title_font
        ws['B1'].alignment = Alignment(vertical='center', horizontal='left')
        ws.row_dimensions[1].height = 30
        
        # Write header information
        season = self.get_season(self.date_var.get())
        current_team_name = self.team_names[self.current_team]
        team_with_season = f"{current_team_name} - {season}" if season else current_team_name
        
        header_info = {
            ('B2', 'Team:'): team_with_season,
            ('B3', 'Session #:'): self.session_var.get(),
            ('B4', 'Event:'): self.event_var.get(),
            ('B5', 'Snow Condition:'): self.snow_condition_var.get(),
            ('E2', 'Date:'): self.date_var.get(),
            ('E3', 'Start Time:'): self.time_var.get(),
            ('E4', 'Hill:'): self.hill_var.get(),
            ('E5', 'Weather:'): f"{self.sky_condition_var.get()}, {self.precipitation_var.get()}, {self.wind_condition_var.get()}"
        }
        
        for (cell_coord, label), value in header_info.items():
            ws[cell_coord] = label
            ws[cell_coord].font = header_font
            ws[cell_coord].alignment = Alignment(horizontal='left')
            
            value_col = chr(ord(cell_coord[0]) + 1)
            value_cell = f"{value_col}{cell_coord[1]}"
            ws[value_cell] = value
            ws[value_cell].font = normal_font
            ws[value_cell].alignment = Alignment(horizontal='left')
        
        # Process timing data
        if self.selected_file:
            timing_data = self.parse_timing_data(self.selected_file)
            if timing_data:
                current_row = 8  # Start after header section
                
                # Process each run
                for run_number in sorted(timing_data.keys(), key=int):
                    current_row = self.write_run_data(
                        ws, 
                        timing_data[run_number], 
                        current_row, 
                        split_columns
                    )
                    current_row += 1  # Space between runs
                
        try:
            wb.save(output_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {str(e)}")
            return False

    def write_run_data(self, ws, run_data, start_row, split_columns):
        """Writes run data to Excel with proper split handling."""
        # Write run header
        current_row = start_row + 2
        ws[f'B{current_row}'] = f"Run {run_data[0]['run']}"
        current_row += 1
        
        # Generate dynamic headers based on available splits
        headers = ['Bib #', 'Name']
        for split_idx in range(self.num_splits):
            headers.extend([f'Split {split_idx + 1}', 'Diff.', 'Rank'])
        headers.extend(['Finish', 'Diff.', 'Status'])
        
        # Write headers
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Process and collect valid times for each split
        split_data = []
        for split_idx in range(self.num_splits):
            valid_times = []
            for entry in run_data:
                if (entry['status'].upper() not in ['DNS', 'ERR'] and 
                    len(entry['splits']) > split_idx and 
                    entry['splits'][split_idx] is not None and 
                    entry['splits'][split_idx] > 0):
                    valid_times.append((entry['splits'][split_idx], entry['bib']))
            
            if valid_times:
                valid_times.sort(key=lambda x: x[0])  # Sort by time
            split_data.append(valid_times)
        
        # Collect finish times
        finish_data = []
        for entry in run_data:
            if (entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR'] and 
                entry['finish'] is not None and 
                entry['finish'] > 0):
                finish_data.append((entry['finish'], entry['bib']))
        
        if finish_data:
            finish_data.sort(key=lambda x: x[0])
        
        # Write data rows
        for entry in run_data:
            bib = entry['bib']
            status = entry['status'].upper()
            
            # Base row data
            row_data = [bib, self.get_athlete_name(bib)]
            col = 4  # Start after Name column
            
            # Process each split
            for split_idx in range(self.num_splits):
                if (split_idx < len(entry['splits']) and 
                    entry['splits'][split_idx] is not None and 
                    entry['splits'][split_idx] > 0):
                    
                    split_time = entry['splits'][split_idx]
                    best_split = split_data[split_idx][0][0] if split_data[split_idx] else None
                    split_diff = split_time - best_split if best_split is not None else None
                    
                    # Find rank
                    rank = next((i + 1 for i, (_, b) in enumerate(split_data[split_idx]) 
                            if b == bib), '')
                    
                    # Write split data
                    ws.cell(row=current_row, column=col, value=self.format_time(split_time))
                    ws.cell(row=current_row, column=col + 1, 
                        value=self.format_time(split_diff, True) if split_diff is not None else '')
                    ws.cell(row=current_row, column=col + 2, value=rank)
                    
                    # Apply color gradient for valid times
                    if rank and rank != '':
                        ws.cell(row=current_row, column=col).fill = self.get_gradient_color(
                            rank - 1,  # Zero-based index for gradient
                            len(split_data[split_idx]),
                            0.6 if split_idx == 0 else 0.5  # Stronger color for first split
                        )
                else:
                    # Write empty cells for missing split
                    for i in range(3):
                        ws.cell(row=current_row, column=col + i, value='')
                
                col += 3  # Move to next split columns
            
            # Write finish data
            if status not in ['DNF', 'DSQ', 'DNS', 'ERR'] and entry['finish'] is not None:
                best_finish = finish_data[0][0] if finish_data else None
                finish_diff = entry['finish'] - best_finish if best_finish is not None else None
                
                finish_rank = next((i + 1 for i, (_, b) in enumerate(finish_data) 
                                if b == bib), '')
                
                ws.cell(row=current_row, column=col, value=self.format_time(entry['finish']))
                ws.cell(row=current_row, column=col + 1, 
                    value=self.format_time(finish_diff, True) if finish_diff is not None else '')
                ws.cell(row=current_row, column=col + 2, value=status)
                
                if finish_rank and finish_rank != '':
                    ws.cell(row=current_row, column=col).fill = self.get_gradient_color(
                        finish_rank - 1,
                        len(finish_data),
                        0.5
                    )
            else:
                ws.cell(row=current_row, column=col + 2, value=status)
            
            # Apply borders to all cells in row
            for column in range(2, col + 3):
                cell = ws.cell(row=current_row, column=column)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
        
        return current_row

    def build_athlete_management(self, parent_frame):
        """Creates the athlete management section with ttk styling."""
        # Right Side Frame with modern styling
        right_frame = ttk.LabelFrame(parent_frame, text="Athlete Management")
        right_frame.grid(row=1, column=2, rowspan=3, padx=10, pady=10, sticky="nsew")

        # Team Selection Section
        team_frame = ttk.LabelFrame(right_frame, text="Team Selection")
        team_frame.pack(fill="x", padx=5, pady=5)

        team_button_frame = ttk.Frame(team_frame)
        team_button_frame.pack(pady=5)

        # Team buttons with consistent styling
        self.sqah_button = ttk.Button(
            team_button_frame,
            text="SQAH",
            command=lambda: self.set_team("SQAH"),
            style="Accent.TButton"
        )
        self.sqah_button.pack(side=tk.LEFT, padx=5)

        self.sqaf_button = ttk.Button(
            team_button_frame,
            text="SQAF",
            command=lambda: self.set_team("SQAF")
        )
        self.sqaf_button.pack(side=tk.LEFT, padx=5)

        self.other_button = ttk.Button(
            team_button_frame,
            text="OTHER",
            command=lambda: self.set_team("OTHER")
        )
        self.other_button.pack(side=tk.LEFT, padx=5)

        # Athletes List Section
        athletes_frame = ttk.LabelFrame(right_frame, text="Current Athletes")
        athletes_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Use ttk.Treeview instead of Listbox for better styling
        self.athlete_tree = ttk.Treeview(
            athletes_frame,
            columns=("bib", "name"),
            show="headings",
            height=10
        )
        self.athlete_tree.heading("bib", text="Bib")
        self.athlete_tree.heading("name", text="Name")
        self.athlete_tree.heading("gender", text="Gender")
        self.athlete_tree.column("bib", width=50)
        self.athlete_tree.column("name", width=150)
        self.athlete_tree.column("gender", width=50)
        self.athlete_tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Add gender selection to the Add Athlete section
        add_frame = ttk.LabelFrame(right_frame, text="Add Athlete/Guest")
        add_frame.pack(fill="x", padx=5, pady=5)

        # Add gender selection
        gender_frame = ttk.Frame(add_frame)
        gender_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(gender_frame, text="Gender:").pack(side=tk.LEFT)
        self.gender_var = tk.StringVar(value="M")
        gender_combo = ttk.Combobox(
            gender_frame, 
            textvariable=self.gender_var,
            values=["M", "F"],
            width=5,
            state="readonly"
        )
        gender_combo.pack(side=tk.LEFT, padx=5)

        # Add Split Categories button
        split_button = ttk.Button(
            right_frame,
            text="Split Categories",
            command=self.create_split_category_excel,
            style="Accent.TButton"
        )
        split_button.pack(fill="x", padx=5, pady=5)

        # Add scrollbar to treeview
        athlete_scrollbar = ttk.Scrollbar(
            athletes_frame,
            orient="vertical",
            command=self.athlete_tree.yview
        )
        athlete_scrollbar.pack(side="right", fill="y")
        self.athlete_tree.configure(yscrollcommand=athlete_scrollbar.set)

        # Temporary Guests Section
        guests_frame = ttk.LabelFrame(right_frame, text="Temporary Guests")
        guests_frame.pack(fill="both", expand=True, padx=5, pady=5)

        self.guest_tree = ttk.Treeview(
            guests_frame,
            columns=("bib", "name", "status"),
            show="headings",
            height=5
        )
        self.guest_tree.heading("bib", text="Bib")
        self.guest_tree.heading("name", text="Name")
        self.guest_tree.heading("status", text="Status")
        self.guest_tree.column("bib", width=50)
        self.guest_tree.column("name", width=150)
        self.guest_tree.column("status", width=50)
        self.guest_tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Add scrollbar to guest treeview
        guest_scrollbar = ttk.Scrollbar(
            guests_frame,
            orient="vertical",
            command=self.guest_tree.yview
        )
        guest_scrollbar.pack(side="right", fill="y")
        self.guest_tree.configure(yscrollcommand=guest_scrollbar.set)

        # Add Athlete Section
        add_frame = ttk.LabelFrame(right_frame, text="Add Athlete/Guest")
        add_frame.pack(fill="x", padx=5, pady=5)

        # Name entry with autocomplete
        name_frame = ttk.Frame(add_frame)
        name_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(name_frame, text="Name:").pack(side=tk.LEFT)
        self.athlete_name_entry = ttk.Entry(name_frame)
        self.athlete_name_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=5)

        # Bib entry
        bib_frame = ttk.Frame(add_frame)
        bib_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(bib_frame, text="Bib #:").pack(side=tk.LEFT)
        self.bib_number_entry = ttk.Entry(bib_frame)
        self.bib_number_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=5)

        # Buttons frame
        button_frame = ttk.Frame(add_frame)
        button_frame.pack(fill="x", padx=5, pady=5)

        ttk.Button(
            button_frame,
            text="Add Athlete",
            command=self.add_athlete,
            style="Accent.TButton"
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Add Guest",
            command=self.add_guest
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Remove Selected",
            command=self.remove_selected_athlete,
            style="Danger.TButton"
        ).pack(side=tk.LEFT, padx=5)

        # Bind events
        self.athlete_name_entry.bind('<KeyRelease>', self.autocomplete_athlete_name)
        self.athlete_tree.bind('<Delete>', self.remove_selected_athlete)
        self.guest_tree.bind('<Delete>', self.remove_selected_athlete)
        
        # Context menu setup
        self.create_context_menus()

    def create_context_menus(self):
        """Creates right-click context menus for athletes and guests."""
        self.athlete_menu = tk.Menu(self.root, tearoff=0)
        self.athlete_menu.add_command(
            label="Edit",
            command=self.edit_selected_athlete
        )
        self.athlete_menu.add_command(
            label="Remove",
            command=self.remove_selected_athlete
        )
        self.athlete_menu.add_separator()
        self.athlete_menu.add_command(
            label="View Statistics",
            command=self.view_athlete_statistics
        )

        self.guest_menu = tk.Menu(self.root, tearoff=0)
        self.guest_menu.add_command(
            label="Convert to Athlete",
            command=self.convert_guest_to_athlete
        )
        self.guest_menu.add_command(
            label="Remove",
            command=self.remove_selected_guest
        )

        # Bind right-click events
        self.athlete_tree.bind("<Button-3>", self.show_athlete_context_menu)
        self.guest_tree.bind("<Button-3>", self.show_guest_context_menu)

    def update_athlete_display(self):
        """Updates the athlete treeview with current team's athletes."""
        # Clear current display
        for item in self.athlete_tree.get_children():
            self.athlete_tree.delete(item)

        # Add sorted athletes
        for athlete in sorted(self.athletes[self.current_team], key=lambda x: int(x['bib'])):
            self.athlete_tree.insert(
                "",
                "end",
                values=(athlete['bib'], athlete['name'])
            )

    def update_guest_display(self):
        """Updates the guest treeview."""
        # Clear current display
        for item in self.guest_tree.get_children():
            self.guest_tree.delete(item)

        # Add all guests
        for guest in self.temp_guests:
            status = "Inactive" if guest.get("inactive", False) else "Active"
            self.guest_tree.insert(
                "",
                "end",
                values=(guest['bib'], guest['name'], status)
            )

    def add_athlete(self):
        """Enhanced add_athlete method with gender support."""
        athlete_name = self.athlete_name_entry.get()
        bib_number = self.bib_number_entry.get()
        gender = self.gender_var.get()

        if athlete_name and bib_number and self.current_team:
            if not bib_number.isdigit():
                messagebox.showwarning("Input Error", "Bib invalid")
                return

            bib_number = int(bib_number)

            # Check for duplicate bib
            for athlete in self.athletes[self.current_team]:
                if int(athlete['bib']) == bib_number:
                    messagebox.showwarning("Duplicate Bib", 
                                         "This bib number is already assigned.")
                    return

            # Add the new athlete with gender
            new_athlete = {
                "name": athlete_name,
                "bib": str(bib_number),
                "gender": gender
            }
            self.athletes[self.current_team].append(new_athlete)
            self.save_athletes_to_json()
            self.update_athlete_display()

            # Add to recent names
            self.add_athlete_to_memory(athlete_name)

            # Clear inputs
            self.athlete_name_entry.delete(0, tk.END)
            self.bib_number_entry.delete(0, tk.END)
            self.gender_var.set("M")  # Reset to default
        else:
            messagebox.showwarning("Input Error", "Please fill all fields.")

    def create_split_category_excel(self):
        """Creates an Excel file with men and women split into separate sections."""
        if not hasattr(self, 'timing_data') or not self.timing_data:
            messagebox.showwarning("Warning", "No timing data available")
            return

        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Split Categories"

        # Style definitions
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=11)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", 
                                fill_type="solid")

        current_row = 1

        # Process each run
        for run_number in sorted(self.timing_data.keys(), key=int):
            run_data = self.timing_data[run_number]
            
            # Split athletes by gender
            men_data = []
            women_data = []
            
            for entry in run_data:
                bib = entry['bib']
                athlete = self.find_athlete_by_bib(bib)
                if athlete:
                    if athlete['gender'] == 'M':
                        men_data.append(entry)
                    else:
                        women_data.append(entry)

            # Write run header
            ws.cell(row=current_row, column=1, value=f"Run {run_number}")
            ws.merge_cells(start_row=current_row, start_column=1, 
                         end_row=current_row, end_column=6)
            current_row += 1

            # Write women's section
            current_row = self.write_category_data(
                ws, women_data, current_row, "Women", header_fill)
            
            current_row += 1  # Add spacing

            # Write men's section
            current_row = self.write_category_data(
                ws, men_data, current_row, "Men", header_fill)
            
            current_row += 2  # Add spacing between runs

        # Save the file
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="split_categories.xlsx"
            )
            if filename:
                wb.save(filename)
                messagebox.showinfo("Success", "Split categories file created successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {str(e)}")

    def write_category_data(self, ws, data, start_row, category, header_fill):
        """Writes category-specific data to the worksheet."""
        current_row = start_row

        # Category header
        ws.cell(row=current_row, column=1, value=category)
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=6)
        header_cell = ws.cell(row=current_row, column=1)
        header_cell.fill = header_fill
        header_cell.font = Font(bold=True)
        current_row += 1

        # Column headers
        headers = ['Bib', 'Name', 'Split 1', 'Split Diff', 'Finish', 'Diff']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        current_row += 1

        # Sort data by finish time
        valid_data = [entry for entry in data 
                     if entry['status'].upper() not in ['DNF', 'DSQ', 'DNS', 'ERR']]
        valid_data.sort(key=lambda x: x['finish'] if x['finish'] is not None else float('inf'))

        # Get best times for diffs
        best_split = min((entry['splits'][0] for entry in valid_data 
                         if entry['splits'] and entry['splits'][0] is not None), 
                        default=None)
        best_finish = min((entry['finish'] for entry in valid_data 
                          if entry['finish'] is not None), 
                         default=None)

        # Write data
        for entry in valid_data:
            row = [
                entry['bib'],
                self.get_athlete_name(entry['bib']),
                self.format_time(entry['splits'][0]) if entry['splits'] else '',
                self.format_time(entry['splits'][0] - best_split, True) 
                    if entry['splits'] and entry['splits'][0] and best_split else '',
                self.format_time(entry['finish']) if entry['finish'] else '',
                self.format_time(entry['finish'] - best_finish, True) 
                    if entry['finish'] and best_finish else ''
            ]
            
            for col, value in enumerate(row, start=1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1

        return current_row

    def find_athlete_by_bib(self, bib):
        """Finds an athlete across all teams by bib number."""
        for team in self.athletes.values():
            for athlete in team:
                if athlete['bib'] == str(bib):
                    return athlete
        # Check temporary guests
        for guest in self.temp_guests:
            if guest['bib'] == str(bib):
                return guest
        return None

    def build_gui(self):
        """Builds the complete GUI with ttk styling and enhanced functionality."""
        # Configure grid weights for proper layout
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(2, weight=1)
        
        # Create menu and toolbar
        self.create_menu_system()
        self.create_toolbar()
        
        # Main title
        title_frame = ttk.Frame(self.root)
        title_frame.grid(row=0, column=0, columnspan=3, pady=(10, 0))
        
        title_label = ttk.Label(
            title_frame,
            text="Brower Timing Reformatted",
            font=("Arial", 20, "bold")
        )
        title_label.pack()

        # Build main sections
        self.build_file_section()
        self.build_training_parameters()
        self.build_athlete_management()
        self.build_bottom_section()
        
        # Set up keyboard shortcuts
        self.setup_keyboard_shortcuts()
        
        # Create status bar
        self.create_status_bar()

    def build_file_section(self):
        """Creates the file selection and management section."""
        left_frame = ttk.Frame(self.root)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="n")
        
        # File Selection Section
        file_frame = ttk.LabelFrame(left_frame, text="File Selection")
        file_frame.pack(fill="x", padx=5, pady=5)
        
        file_content = ttk.Frame(file_frame)
        file_content.pack(padx=10, pady=5)
        
        self.select_file_button = ttk.Button(
            file_content,
            text="Select Brower CSV",
            command=self.select_file,
            style="Accent.TButton"
        )
        self.select_file_button.pack(side=tk.LEFT, padx=5)
        
        self.file_label = ttk.Label(
            file_content,
            text="No file selected",
            wraplength=300
        )
        self.file_label.pack(side=tk.LEFT, padx=5)
        
        # File Details Section
        details_frame = ttk.LabelFrame(left_frame, text="File Details")
        details_frame.pack(fill="x", padx=5, pady=5)
        
        # Date
        date_frame = ttk.Frame(details_frame)
        date_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(date_frame, text="Date:").pack(side=tk.LEFT)
        self.date_entry = ttk.Entry(
            date_frame,
            textvariable=self.date_var,
            state="readonly"
        )
        self.date_entry.pack(side=tk.LEFT, padx=5, fill="x", expand=True)
        
        # Time
        time_frame = ttk.Frame(details_frame)
        time_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(time_frame, text="Start Time:").pack(side=tk.LEFT)
        self.time_entry = ttk.Entry(
            time_frame,
            textvariable=self.time_var,
            state="readonly"
        )
        self.time_entry.pack(side=tk.LEFT, padx=5, fill="x", expand=True)
        
        # Session
        session_frame = ttk.Frame(details_frame)
        session_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(session_frame, text="Session #:").pack(side=tk.LEFT)
        self.session_entry = ttk.Entry(
            session_frame,
            textvariable=self.session_var,
            state="readonly"
        )
        self.session_entry.pack(side=tk.LEFT, padx=5, fill="x", expand=True)
        
        # Reformat Button Section
        reformat_frame = ttk.LabelFrame(left_frame, text="Export")
        reformat_frame.pack(fill="x", padx=5, pady=10)
        
        self.reformat_button = ttk.Button(
            reformat_frame,
            text="Reformat Selected File",
            command=self.reformat_file,
            style="Accent.TButton"
        )
        self.reformat_button.pack(padx=20, pady=10)

    def build_training_parameters(self):
        """Creates the training parameters section with event and weather settings."""
        params_frame = ttk.LabelFrame(self.root)
        params_frame.grid(row=1, column=1, padx=10, pady=10, sticky="n")
        
        # Event and Location Frame
        event_loc_frame = ttk.LabelFrame(params_frame, text="Event & Location")
        event_loc_frame.pack(fill="x", padx=5, pady=5)
        
        # Event Selection
        event_frame = ttk.Frame(event_loc_frame)
        event_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(event_frame, text="Event:").pack(side=tk.LEFT)
        self.event_combo = ttk.Combobox(
            event_frame,
            textvariable=self.event_var,
            values=["SL", "GS", "SG", "DH", "SX"],
            width=10,
            state="readonly"
        )
        self.event_combo.pack(side=tk.LEFT, padx=5)
        
        # Hill Selection with Autocomplete
        hill_frame = ttk.Frame(event_loc_frame)
        hill_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(hill_frame, text="Hill:").pack(side=tk.LEFT)
        self.hill_entry = ttk.Entry(
            hill_frame,
            textvariable=self.hill_var
        )
        self.hill_entry.pack(side=tk.LEFT, padx=5, fill="x", expand=True)
        
        # Hill Autocomplete Listbox
        self.hill_autocomplete_listbox = tk.Listbox(
            event_loc_frame,
            height=0,
            borderwidth=1,
            relief="solid"
        )
        self.hill_autocomplete_listbox.pack_forget()
        
        # Weather Conditions Section
        weather_frame = ttk.LabelFrame(params_frame, text="Weather Conditions")
        weather_frame.pack(fill="x", padx=5, pady=5)
        
        # Snow Conditions
        snow_frame = ttk.Frame(weather_frame)
        snow_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(snow_frame, text="Snow:").pack(side=tk.LEFT)
        self.snow_combo = ttk.Combobox(
            snow_frame,
            textvariable=self.snow_condition_var,
            values=["Very Soft", "Soft", "Medium", "Hard", "Injected"],
            width=15,
            state="readonly"
        )
        self.snow_combo.pack(side=tk.LEFT, padx=5)
        self.snow_combo.set("Hard")
        
        # Sky Conditions
        sky_frame = ttk.Frame(weather_frame)
        sky_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(sky_frame, text="Sky:").pack(side=tk.LEFT)
        self.sky_combo = ttk.Combobox(
            sky_frame,
            textvariable=self.sky_condition_var,
            values=["Clear Sky", "Partly Cloudy", "Overcast", "Low Visibility", "Fog"],
            width=15,
            state="readonly"
        )
        self.sky_combo.pack(side=tk.LEFT, padx=5)
        self.sky_combo.set("Clear Sky")
        
        # Precipitation
        precip_frame = ttk.Frame(weather_frame)
        precip_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(precip_frame, text="Precipitation:").pack(side=tk.LEFT)
        self.precip_combo = ttk.Combobox(
            precip_frame,
            textvariable=self.precipitation_var,
            values=["No Precipitation", "Light Snow", "Medium Snow", "Snowstorm",
                    "Light Rain", "Moderate Rain", "Downpour"],
            width=15,
            state="readonly"
        )
        self.precip_combo.pack(side=tk.LEFT, padx=5)
        self.precip_combo.set("No Precipitation")
        
        # Wind Conditions
        wind_frame = ttk.Frame(weather_frame)
        wind_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(wind_frame, text="Wind:").pack(side=tk.LEFT)
        self.wind_combo = ttk.Combobox(
            wind_frame,
            textvariable=self.wind_condition_var,
            values=["No Wind", "Light Wind", "Moderate Wind", "Heavy Wind"],
            width=15,
            state="readonly"
        )
        self.wind_combo.pack(side=tk.LEFT, padx=5)
        self.wind_combo.set("No Wind")
        
        # Bind hill autocomplete events
        self.hill_entry.bind('<KeyRelease>', self.autocomplete_hill_name)
        self.hill_autocomplete_listbox.bind("<<ListboxSelect>>", self.on_hill_suggestion_select)

    def build_athlete_management(self):
        """Creates the athlete management section with enhanced team and gender support."""
        # Right Side Frame with modern styling
        right_frame = ttk.Frame(self.root)
        right_frame.grid(row=1, column=2, rowspan=3, padx=10, pady=10, sticky="n")
        
        # Team Selection Section
        team_frame = ttk.LabelFrame(right_frame, text="Team Selection")
        team_frame.pack(fill="x", padx=5, pady=5)
        
        team_buttons = ttk.Frame(team_frame)
        team_buttons.pack(pady=5)
        
        # Team Selection Buttons
        self.sqah_button = ttk.Button(
            team_buttons,
            text=self.team_names["SQAH"],
            command=lambda: self.set_team("SQAH"),
            style="Accent.TButton",
            width=12
        )
        self.sqah_button.pack(side=tk.LEFT, padx=2)
        
        self.sqaf_button = ttk.Button(
            team_buttons,
            text=self.team_names["SQAF"],
            command=lambda: self.set_team("SQAF"),
            width=12
        )
        self.sqaf_button.pack(side=tk.LEFT, padx=2)
        
        self.other_button = ttk.Button(
            team_buttons,
            text="OTHER",
            command=lambda: self.set_team("OTHER"),
            width=12
        )
        self.other_button.pack(side=tk.LEFT, padx=2)
        
        # Athletes Treeview Section
        athletes_frame = ttk.LabelFrame(right_frame, text="Current Athletes")
        athletes_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Create athletes treeview with scrollbar
        athlete_tree_frame = ttk.Frame(athletes_frame)
        athlete_tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.athlete_tree = ttk.Treeview(
            athlete_tree_frame,
            columns=("bib", "name", "gender"),
            show="headings",
            height=10
        )
        
        # Configure treeview columns
        self.athlete_tree.heading("bib", text="Bib")
        self.athlete_tree.heading("name", text="Name")
        self.athlete_tree.heading("gender", text="Gender")
        
        self.athlete_tree.column("bib", width=50)
        self.athlete_tree.column("name", width=200)
        self.athlete_tree.column("gender", width=60)
        
        # Add scrollbar to treeview
        athlete_scrollbar = ttk.Scrollbar(
            athlete_tree_frame,
            orient="vertical",
            command=self.athlete_tree.yview
        )
        self.athlete_tree.configure(yscrollcommand=athlete_scrollbar.set)
        
        # Pack treeview and scrollbar
        self.athlete_tree.pack(side=tk.LEFT, fill="both", expand=True)
        athlete_scrollbar.pack(side=tk.RIGHT, fill="y")
        
        # Temporary Guests Section
        guests_frame = ttk.LabelFrame(right_frame, text="Temporary Guests")
        guests_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Create guests treeview with scrollbar
        guest_tree_frame = ttk.Frame(guests_frame)
        guest_tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.guest_tree = ttk.Treeview(
            guest_tree_frame,
            columns=("bib", "name", "status"),
            show="headings",
            height=5
        )
        
        # Configure guest treeview columns
        self.guest_tree.heading("bib", text="Bib")
        self.guest_tree.heading("name", text="Name")
        self.guest_tree.heading("status", text="Status")
        
        self.guest_tree.column("bib", width=50)
        self.guest_tree.column("name", width=200)
        self.guest_tree.column("status", width=60)
        
        # Add scrollbar to guest treeview
        guest_scrollbar = ttk.Scrollbar(
            guest_tree_frame,
            orient="vertical",
            command=self.guest_tree.yview
        )
        self.guest_tree.configure(yscrollcommand=guest_scrollbar.set)
        
        # Pack guest treeview and scrollbar
        self.guest_tree.pack(side=tk.LEFT, fill="both", expand=True)
        guest_scrollbar.pack(side=tk.RIGHT, fill="y")
        
        # Add Athlete Section
        add_frame = ttk.LabelFrame(right_frame, text="Add Athlete/Guest")
        add_frame.pack(fill="x", padx=5, pady=5)
        
        # Name entry with autocomplete
        name_frame = ttk.Frame(add_frame)
        name_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(name_frame, text="Name:").pack(side=tk.LEFT)
        self.athlete_name_entry = ttk.Entry(name_frame)
        self.athlete_name_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=5)
        
        # Bib entry
        bib_frame = ttk.Frame(add_frame)
        bib_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(bib_frame, text="Bib #:").pack(side=tk.LEFT)
        self.bib_number_entry = ttk.Entry(bib_frame)
        self.bib_number_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=5)
        
        # Gender selection
        gender_frame = ttk.Frame(add_frame)
        gender_frame.pack(fill="x", padx=5, pady=2)
        ttk.Label(gender_frame, text="Gender:").pack(side=tk.LEFT)
        self.gender_combo = ttk.Combobox(
            gender_frame,
            values=["M", "F"],
            width=5,
            state="readonly"
        )
        self.gender_combo.pack(side=tk.LEFT, padx=5)
        self.gender_combo.set("M")
        
        # Buttons frame
        button_frame = ttk.Frame(add_frame)
        button_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Button(
            button_frame,
            text="Add Athlete",
            command=self.add_athlete,
            style="Accent.TButton"
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            button_frame,
            text="Add Guest",
            command=self.add_guest
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            button_frame,
            text="Remove Selected",
            command=self.remove_selected_athlete,
            style="Danger.TButton"
        ).pack(side=tk.LEFT, padx=2)
        
        # Create autocomplete listbox for athlete names
        self.autocomplete_listbox = tk.Listbox(
            add_frame,
            height=0,
            borderwidth=1,
            relief="solid"
        )
        self.autocomplete_listbox.pack_forget()
        
        # Bind events
        self.setup_athlete_bindings()

    def setup_athlete_bindings(self):
        """Sets up event bindings for athlete management."""
        # Athlete tree bindings
        self.athlete_tree.bind("<Delete>", self.remove_selected_athlete)
        self.athlete_tree.bind("<BackSpace>", self.remove_selected_athlete)
        self.athlete_tree.bind("<Double-1>", self.edit_selected_athlete)
        self.athlete_tree.bind("<Button-3>", self.show_athlete_context_menu)
        
        # Guest tree bindings
        self.guest_tree.bind("<Delete>", self.remove_selected_guest)
        self.guest_tree.bind("<BackSpace>", self.remove_selected_guest)
        self.guest_tree.bind("<Double-1>", self.edit_selected_guest)
        self.guest_tree.bind("<Button-3>", self.show_guest_context_menu)
        
        # Entry bindings
        self.athlete_name_entry.bind('<KeyRelease>', self.autocomplete_athlete_name)
        self.autocomplete_listbox.bind("<<ListboxSelect>>", self.on_suggestion_select)
        
        # Quick add bindings
        self.athlete_name_entry.bind('<Control-Return>', lambda e: self.add_athlete())
        self.athlete_name_entry.bind('<Shift-Return>', lambda e: self.add_guest())
        
        # Update bindings
        self.athlete_name_entry.bind('<Configure>', self.update_suggestion_box_width)

    def create_context_menus(self):
        """Creates right-click context menus for athletes and guests."""
        # Athlete context menu
        self.athlete_menu = tk.Menu(self.root, tearoff=0)
        self.athlete_menu.add_command(label="Edit", command=self.edit_selected_athlete)
        self.athlete_menu.add_command(label="Remove", command=self.remove_selected_athlete)
        self.athlete_menu.add_separator()
        self.athlete_menu.add_command(label="View Statistics", command=self.view_athlete_statistics)
        
        # Guest context menu
        self.guest_menu = tk.Menu(self.root, tearoff=0)
        self.guest_menu.add_command(label="Convert to Athlete", command=self.convert_guest_to_athlete)
        self.guest_menu.add_command(label="Remove", command=self.remove_selected_guest)

#todo : creation of ui















